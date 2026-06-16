from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse, Response
from sqlmodel import Session, select
from typing import Optional
from datetime import datetime, date, time as dtime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from backend.database import get_session
from backend.models import (
    Landlord,
    LeaseParcel,
    LeasePeriod,
    LeasePeriodGrainItem,
    LeasePayment,
    LeasePaymentGrainItem,
    GrainCulture,
    GrainStock,
    CashRegister,
    Transaction,
    StockAdjustmentLog,
    StockAdjustmentType,
    Currency,
    TransactionType,
    User,
)
from backend.schemas import (
    LandlordCreate,
    LandlordResponse,
    LandlordUpdate,
    LeaseParcelCreate,
    LeaseParcelUpdate,
    LeaseParcelResponse,
    LeasePeriodCreate,
    LeasePeriodUpdate,
    LeasePeriodResponse,
    LeasePeriodGrainItemResponse,
    LeasePaymentCreate,
    LeasePaymentResponse,
    LeasePaymentGrainItemResponse,
)
from backend.auth import get_current_user, get_current_super_admin

router = APIRouter()

EPS = 0.01
VALID_TERMS = {"grain", "cash", "grain_cash"}


# ===== Helpers =====

def add_one_year(dt):
    """Додати 1 рік до дати"""
    d = dt.date() if isinstance(dt, datetime) else dt
    try:
        return d.replace(year=d.year + 1)
    except ValueError:
        return d.replace(year=d.year + 1, day=28)


def _to_dt(d):
    """date|datetime -> datetime (00:00)"""
    if isinstance(d, datetime):
        return d
    return datetime.combine(d, dtime.min)


def current_price(parcel_id: int, culture_id: int, session: Session) -> float:
    """Поточна ціна культури на ділянці = ціна з найновішого (max year) періоду,
    де є ця культура. Якщо ніде немає — 0."""
    row = session.exec(
        select(LeasePeriodGrainItem.price_per_kg_uah)
        .join(LeasePeriod, LeasePeriod.id == LeasePeriodGrainItem.period_id)
        .where(
            LeasePeriod.parcel_id == parcel_id,
            LeasePeriodGrainItem.culture_id == culture_id,
        )
        .order_by(LeasePeriod.year.desc(), LeasePeriod.id.desc())
    ).first()
    return float(row) if row is not None else 0.0


def grain_paid_kg_map(period_id: int, session: Session) -> dict:
    """culture_id -> сплачено кг по періоду (усі не скасовані виплати, що мають
    зернові позиції: і зерном, і грошима-в-рахунок-зерна)."""
    rows = session.exec(
        select(LeasePaymentGrainItem)
        .join(LeasePayment, LeasePayment.id == LeasePaymentGrainItem.payment_id)
        .where(
            LeasePayment.period_id == period_id,
            LeasePayment.is_cancelled == False,
        )
    ).all()
    res: dict = {}
    for it in rows:
        res[it.culture_id] = res.get(it.culture_id, 0.0) + float(it.quantity_kg or 0.0)
    return res


def cash_paid_uah(period_id: int, session: Session) -> float:
    """Скільки грн сплачено в рахунок грошової частини періоду."""
    payments = session.exec(
        select(LeasePayment).where(
            LeasePayment.period_id == period_id,
            LeasePayment.is_cancelled == False,
            LeasePayment.applies_to == "cash",
        )
    ).all()
    return sum(float(p.amount_uah or 0.0) for p in payments)


def build_period_response(period: LeasePeriod, parcel: LeaseParcel, session: Session) -> LeasePeriodResponse:
    paid_map = grain_paid_kg_map(period.id, session)
    grain_items = session.exec(
        select(LeasePeriodGrainItem).where(LeasePeriodGrainItem.period_id == period.id)
    ).all()
    items_out = []
    grain_remaining_cash = 0.0
    for gi in grain_items:
        culture = session.get(GrainCulture, gi.culture_id)
        paid = float(paid_map.get(gi.culture_id, 0.0))
        remaining_kg = max(0.0, float(gi.quantity_kg or 0.0) - paid)
        cur_price = current_price(parcel.id, gi.culture_id, session)
        rem_cash = remaining_kg * cur_price
        grain_remaining_cash += rem_cash
        items_out.append(LeasePeriodGrainItemResponse(
            id=gi.id,
            period_id=gi.period_id,
            culture_id=gi.culture_id,
            culture_name=culture.name if culture else None,
            quantity_kg=gi.quantity_kg,
            price_per_kg_uah=gi.price_per_kg_uah,
            paid_kg=round(paid, 2),
            remaining_kg=round(remaining_kg, 2),
            current_price_per_kg_uah=round(cur_price, 4),
            remaining_cash_uah=round(rem_cash, 2),
        ))

    cash_oblig_uah = float(period.cash_amount or 0.0) * float(getattr(period, "cash_rate", 1.0) or 1.0)
    paid_cash = cash_paid_uah(period.id, session)
    cash_remaining = max(0.0, cash_oblig_uah - paid_cash)
    remaining_total = round(grain_remaining_cash + cash_remaining, 2)

    return LeasePeriodResponse(
        id=period.id,
        parcel_id=period.parcel_id,
        year=period.year,
        period_start=period.period_start,
        period_end=period.period_end,
        cash_amount=period.cash_amount,
        cash_currency=period.cash_currency,
        cash_rate=float(getattr(period, "cash_rate", 1.0) or 1.0),
        note=period.note,
        grain_items=items_out,
        cash_paid_uah=round(paid_cash, 2),
        cash_remaining_uah=round(cash_remaining, 2),
        grain_remaining_cash_uah=round(grain_remaining_cash, 2),
        remaining_cash_uah=remaining_total,
        created_at=period.created_at,
    )


def build_parcel_response(parcel: LeaseParcel, session: Session) -> LeaseParcelResponse:
    periods = session.exec(
        select(LeasePeriod).where(LeasePeriod.parcel_id == parcel.id).order_by(LeasePeriod.year.asc())
    ).all()
    period_out = [build_period_response(p, parcel, session) for p in periods]
    cumulative = round(sum(p.remaining_cash_uah for p in period_out), 2)
    return LeaseParcelResponse(
        id=parcel.id,
        landlord_id=parcel.landlord_id,
        landlord_full_name=parcel.landlord_full_name,
        area_ha=parcel.area_ha,
        label=parcel.label,
        payment_terms=parcel.payment_terms,
        start_date=parcel.start_date,
        is_active=parcel.is_active,
        note=parcel.note,
        periods=period_out,
        cumulative_balance_uah=cumulative,
        created_at=parcel.created_at,
        updated_at=parcel.updated_at,
    )


def build_payment_response(payment: LeasePayment, session: Session) -> LeasePaymentResponse:
    parcel = session.get(LeaseParcel, payment.parcel_id)
    period = session.get(LeasePeriod, payment.period_id)
    grain_items = session.exec(
        select(LeasePaymentGrainItem).where(LeasePaymentGrainItem.payment_id == payment.id)
    ).all()
    grain_out = []
    for gi in grain_items:
        culture = session.get(GrainCulture, gi.culture_id)
        gi_dict = gi.model_dump()
        gi_dict["culture_name"] = culture.name if culture else None
        grain_out.append(LeasePaymentGrainItemResponse(**gi_dict))

    payment_dict = payment.model_dump()
    payment_dict["grain_items"] = grain_out or None
    payment_dict["landlord_full_name"] = parcel.landlord_full_name if parcel else None
    payment_dict["area_ha"] = parcel.area_ha if parcel else None
    payment_dict["label"] = parcel.label if parcel else None
    payment_dict["period_year"] = period.year if period else None
    if payment.created_by_user_id:
        user = session.get(User, payment.created_by_user_id)
        payment_dict["created_by_user_full_name"] = user.full_name if user else None
    return LeasePaymentResponse(**payment_dict)


# ===== Landlords =====

@router.get("/landlords", response_model=list[LandlordResponse])
async def list_landlords(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    q: Optional[str] = Query(None, description="Пошук за ПІБ")
):
    """Список орендодавців (пошук для автодоповнення)"""
    query = select(Landlord)
    if q:
        query = query.where(Landlord.full_name.ilike(f"%{q}%"))
    return session.exec(query.order_by(Landlord.full_name)).all()


@router.post("/landlords", response_model=LandlordResponse)
async def create_landlord(
    payload: LandlordCreate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Створення орендодавця"""
    cleaned_name = " ".join(payload.full_name.split()).strip()
    if not cleaned_name:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Вкажіть ПІБ орендодавця"
        )

    landlord = Landlord(
        full_name=cleaned_name,
        phone=payload.phone.strip() if payload.phone else None
    )
    session.add(landlord)
    session.commit()
    session.refresh(landlord)
    return landlord


@router.patch("/landlords/{landlord_id}", response_model=LandlordResponse)
async def update_landlord(
    landlord_id: int,
    payload: LandlordUpdate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Оновлення орендодавця (синхронізує знімок ПІБ на ділянках)"""
    landlord = session.get(Landlord, landlord_id)
    if not landlord:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Орендодавця не знайдено"
        )

    if payload.full_name is not None:
        cleaned_name = " ".join(payload.full_name.split()).strip()
        if not cleaned_name:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Вкажіть ПІБ орендодавця"
            )
        landlord.full_name = cleaned_name
        # оновлюємо знімок ПІБ на всіх ділянках орендодавця
        for parcel in session.exec(
            select(LeaseParcel).where(LeaseParcel.landlord_id == landlord_id)
        ).all():
            parcel.landlord_full_name = cleaned_name
            session.add(parcel)

    if payload.phone is not None:
        landlord.phone = payload.phone.strip() if payload.phone else None

    landlord.updated_at = datetime.utcnow()
    session.add(landlord)
    session.commit()
    session.refresh(landlord)
    return landlord


@router.delete("/landlords/{landlord_id}", response_model=LandlordResponse)
async def delete_landlord(
    landlord_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_super_admin)
):
    """Видалення орендодавця (лише без ділянок)"""
    landlord = session.get(Landlord, landlord_id)
    if not landlord:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Орендодавця не знайдено"
        )

    has_parcels = session.exec(
        select(LeaseParcel).where(LeaseParcel.landlord_id == landlord_id)
    ).first()
    if has_parcels:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Неможливо видалити орендодавця з ділянками. Спершу видаліть ділянки."
        )

    session.delete(landlord)
    session.commit()
    return landlord


# ===== Parcels (ділянки) =====

@router.get("/parcels", response_model=list[LeaseParcelResponse])
async def list_parcels(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    landlord_id: Optional[int] = Query(None, description="Фільтр за орендодавцем"),
    is_active: Optional[bool] = Query(None, description="Фільтр за активністю"),
):
    """Список ділянок з періодами та накопичувальним балансом"""
    query = select(LeaseParcel)
    if landlord_id:
        query = query.where(LeaseParcel.landlord_id == landlord_id)
    if is_active is not None:
        query = query.where(LeaseParcel.is_active == is_active)
    parcels = session.exec(query.order_by(LeaseParcel.created_at.desc())).all()
    return [build_parcel_response(p, session) for p in parcels]


@router.get("/parcels/{parcel_id}", response_model=LeaseParcelResponse)
async def get_parcel(
    parcel_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    parcel = session.get(LeaseParcel, parcel_id)
    if not parcel:
        raise HTTPException(status_code=404, detail="Ділянку не знайдено")
    return build_parcel_response(parcel, session)


@router.post("/parcels", response_model=LeaseParcelResponse)
async def create_parcel(
    payload: LeaseParcelCreate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_super_admin)
):
    """Створення ділянки для орендодавця"""
    landlord = session.get(Landlord, payload.landlord_id)
    if not landlord:
        raise HTTPException(status_code=404, detail="Орендодавця не знайдено")

    terms = (payload.payment_terms or "grain").strip()
    if terms not in VALID_TERMS:
        raise HTTPException(status_code=400, detail="Невідомі умови оплати")
    if not payload.area_ha or payload.area_ha <= 0:
        raise HTTPException(status_code=400, detail="Вкажіть кількість га")

    parcel = LeaseParcel(
        landlord_id=landlord.id,
        landlord_full_name=landlord.full_name,
        area_ha=payload.area_ha,
        label=(payload.label.strip() if payload.label else None),
        payment_terms=terms,
        start_date=_to_dt(payload.start_date),
        is_active=True,
        note=payload.note,
    )
    session.add(parcel)
    session.flush()

    # Перший рік разом зі створенням ділянки (атомарно) — за бажанням
    if payload.first_period is not None:
        _build_period(parcel, payload.first_period, session)

    session.commit()
    session.refresh(parcel)
    return build_parcel_response(parcel, session)


@router.patch("/parcels/{parcel_id}", response_model=LeaseParcelResponse)
async def update_parcel(
    parcel_id: int,
    payload: LeaseParcelUpdate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_super_admin)
):
    parcel = session.get(LeaseParcel, parcel_id)
    if not parcel:
        raise HTTPException(status_code=404, detail="Ділянку не знайдено")

    data = payload.model_dump(exclude_unset=True)
    if "payment_terms" in data and data["payment_terms"] not in VALID_TERMS:
        raise HTTPException(status_code=400, detail="Невідомі умови оплати")
    if "area_ha" in data and (not data["area_ha"] or data["area_ha"] <= 0):
        raise HTTPException(status_code=400, detail="Вкажіть кількість га")
    if "label" in data and data["label"]:
        data["label"] = data["label"].strip()
    if "start_date" in data and data["start_date"]:
        data["start_date"] = _to_dt(data["start_date"])

    for field, value in data.items():
        setattr(parcel, field, value)
    parcel.updated_at = datetime.utcnow()
    session.add(parcel)
    session.commit()
    session.refresh(parcel)
    return build_parcel_response(parcel, session)


@router.delete("/parcels/{parcel_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_parcel(
    parcel_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_super_admin)
):
    """Видалення ділянки разом з періодами (лише якщо немає активних виплат)"""
    parcel = session.get(LeaseParcel, parcel_id)
    if not parcel:
        raise HTTPException(status_code=404, detail="Ділянку не знайдено")

    active_payment = session.exec(
        select(LeasePayment).where(
            LeasePayment.parcel_id == parcel_id,
            LeasePayment.is_cancelled == False,
        )
    ).first()
    if active_payment:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Неможливо видалити ділянку з активними виплатами. Спершу скасуйте виплати."
        )

    # Видаляємо у порядку залежностей (без ORM relationships UOW не гарантує порядок —
    # тому явно прибираємо дітей і робимо flush перед батьками).
    payment_ids = [p.id for p in session.exec(
        select(LeasePayment).where(LeasePayment.parcel_id == parcel_id)
    ).all()]
    if payment_ids:
        for gi in session.exec(
            select(LeasePaymentGrainItem).where(LeasePaymentGrainItem.payment_id.in_(payment_ids))
        ).all():
            session.delete(gi)
        session.flush()
        for p in session.exec(
            select(LeasePayment).where(LeasePayment.parcel_id == parcel_id)
        ).all():
            session.delete(p)
        session.flush()

    period_ids = [pe.id for pe in session.exec(
        select(LeasePeriod).where(LeasePeriod.parcel_id == parcel_id)
    ).all()]
    if period_ids:
        for gi in session.exec(
            select(LeasePeriodGrainItem).where(LeasePeriodGrainItem.period_id.in_(period_ids))
        ).all():
            session.delete(gi)
        session.flush()
        for period in session.exec(
            select(LeasePeriod).where(LeasePeriod.parcel_id == parcel_id)
        ).all():
            session.delete(period)
        session.flush()

    session.delete(parcel)
    session.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


# ===== Periods (роки) =====

@router.get("/parcels/{parcel_id}/periods", response_model=list[LeasePeriodResponse])
async def list_periods(
    parcel_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    parcel = session.get(LeaseParcel, parcel_id)
    if not parcel:
        raise HTTPException(status_code=404, detail="Ділянку не знайдено")
    periods = session.exec(
        select(LeasePeriod).where(LeasePeriod.parcel_id == parcel_id).order_by(LeasePeriod.year.asc())
    ).all()
    return [build_period_response(p, parcel, session) for p in periods]


def _build_period(parcel: LeaseParcel, payload: LeasePeriodCreate, session: Session) -> LeasePeriod:
    """Створює період (рік) для ділянки БЕЗ commit. Валідує умови, дублікат року,
    дати та зернові позиції. Використовується open_period та create_parcel (перший рік)."""
    existing = session.exec(
        select(LeasePeriod).where(
            LeasePeriod.parcel_id == parcel.id,
            LeasePeriod.year == payload.year,
        )
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"Рік {payload.year} вже відкрито на цій ділянці")

    terms = parcel.payment_terms
    needs_grain = terms in ("grain", "grain_cash")
    needs_cash = terms in ("cash", "grain_cash")

    if needs_grain and not payload.grain_items:
        raise HTTPException(status_code=400, detail="Додайте хоча б одну зернову позицію")
    if needs_cash and (not payload.cash_amount or payload.cash_amount <= 0):
        raise HTTPException(status_code=400, detail="Вкажіть річну грошову суму")

    if payload.period_start:
        p_start = _to_dt(payload.period_start)
    else:
        base = parcel.start_date.date() if isinstance(parcel.start_date, datetime) else parcel.start_date
        try:
            p_start = datetime.combine(base.replace(year=payload.year), dtime.min)
        except ValueError:
            p_start = datetime.combine(base.replace(year=payload.year, day=28), dtime.min)
    p_end = _to_dt(add_one_year(p_start))

    period = LeasePeriod(
        parcel_id=parcel.id,
        year=payload.year,
        period_start=p_start,
        period_end=p_end,
        cash_amount=(payload.cash_amount if needs_cash else 0.0),
        cash_currency=(payload.cash_currency if needs_cash else "UAH"),
        cash_rate=(payload.cash_rate if needs_cash else 1.0),
        note=payload.note,
    )
    session.add(period)
    session.flush()

    if needs_grain:
        for item in payload.grain_items:
            culture = session.get(GrainCulture, item.culture_id)
            if not culture:
                raise HTTPException(status_code=404, detail=f"Культуру з ID {item.culture_id} не знайдено")
            price = item.price_per_kg_uah if item.price_per_kg_uah and item.price_per_kg_uah > 0 else float(culture.price_per_kg or 1.0)
            if not item.quantity_kg or item.quantity_kg <= 0:
                raise HTTPException(status_code=400, detail=f"Вкажіть кількість для '{culture.name}'")
            session.add(LeasePeriodGrainItem(
                period_id=period.id,
                culture_id=item.culture_id,
                quantity_kg=item.quantity_kg,
                price_per_kg_uah=price,
            ))
    return period


@router.post("/parcels/{parcel_id}/periods", response_model=LeasePeriodResponse)
async def open_period(
    parcel_id: int,
    payload: LeasePeriodCreate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_super_admin)
):
    """Відкрити новий рік (період) на ділянці.
    Зернові ціни беруться зі знімка (price_per_kg_uah), за замовч. — поточна ціна культури.
    """
    parcel = session.get(LeaseParcel, parcel_id)
    if not parcel:
        raise HTTPException(status_code=404, detail="Ділянку не знайдено")

    period = _build_period(parcel, payload, session)
    session.commit()
    session.refresh(period)
    return build_period_response(period, parcel, session)


@router.patch("/periods/{period_id}", response_model=LeasePeriodResponse)
async def update_period(
    period_id: int,
    payload: LeasePeriodUpdate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_super_admin)
):
    period = session.get(LeasePeriod, period_id)
    if not period:
        raise HTTPException(status_code=404, detail="Період не знайдено")
    parcel = session.get(LeaseParcel, period.parcel_id)

    data = payload.model_dump(exclude_unset=True)
    grain_items = data.pop("grain_items", None)

    for field, value in data.items():
        setattr(period, field, value)
    session.add(period)

    if grain_items is not None:
        # перевірка: не можна прибрати позицію, по якій вже є виплати
        paid_map = grain_paid_kg_map(period_id, session)
        new_culture_ids = {gi["culture_id"] if isinstance(gi, dict) else gi.culture_id for gi in grain_items}
        for cid, paid in paid_map.items():
            if paid > EPS and cid not in new_culture_ids:
                culture = session.get(GrainCulture, cid)
                raise HTTPException(
                    status_code=400,
                    detail=f"Не можна прибрати культуру '{culture.name if culture else cid}' — по ній вже є виплати"
                )
        for old in session.exec(
            select(LeasePeriodGrainItem).where(LeasePeriodGrainItem.period_id == period_id)
        ).all():
            session.delete(old)
        session.flush()
        for gi in grain_items:
            cid = gi["culture_id"] if isinstance(gi, dict) else gi.culture_id
            qty = gi["quantity_kg"] if isinstance(gi, dict) else gi.quantity_kg
            price = gi.get("price_per_kg_uah") if isinstance(gi, dict) else gi.price_per_kg_uah
            culture = session.get(GrainCulture, cid)
            if not culture:
                raise HTTPException(status_code=404, detail=f"Культуру з ID {cid} не знайдено")
            if not price or price <= 0:
                price = float(culture.price_per_kg or 1.0)
            session.add(LeasePeriodGrainItem(
                period_id=period_id,
                culture_id=cid,
                quantity_kg=qty,
                price_per_kg_uah=price,
            ))

    session.commit()
    session.refresh(period)
    return build_period_response(period, parcel, session)


@router.delete("/periods/{period_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_period(
    period_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_super_admin)
):
    period = session.get(LeasePeriod, period_id)
    if not period:
        raise HTTPException(status_code=404, detail="Період не знайдено")

    has_payment = session.exec(
        select(LeasePayment).where(LeasePayment.period_id == period_id)
    ).first()
    if has_payment:
        raise HTTPException(
            status_code=400,
            detail="Неможливо видалити рік з виплатами. Спершу скасуйте/видаліть виплати."
        )
    for gi in session.exec(
        select(LeasePeriodGrainItem).where(LeasePeriodGrainItem.period_id == period_id)
    ).all():
        session.delete(gi)
    session.flush()
    session.delete(period)
    session.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@router.get("/periods/{period_id}/balance", response_model=LeasePeriodResponse)
async def get_period_balance(
    period_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    period = session.get(LeasePeriod, period_id)
    if not period:
        raise HTTPException(status_code=404, detail="Період не знайдено")
    parcel = session.get(LeaseParcel, period.parcel_id)
    return build_period_response(period, parcel, session)


# ===== Payments =====

@router.get("/payments", response_model=list[LeasePaymentResponse])
async def list_payments(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    parcel_id: Optional[int] = Query(None),
    period_id: Optional[int] = Query(None),
    landlord_id: Optional[int] = Query(None),
):
    query = select(LeasePayment)
    if parcel_id:
        query = query.where(LeasePayment.parcel_id == parcel_id)
    if period_id:
        query = query.where(LeasePayment.period_id == period_id)
    if landlord_id:
        query = query.where(
            LeasePayment.parcel_id.in_(
                select(LeaseParcel.id).where(LeaseParcel.landlord_id == landlord_id)
            )
        )
    payments = session.exec(
        query.order_by(LeasePayment.is_cancelled.asc(), LeasePayment.payment_date.desc())
    ).all()
    return [build_payment_response(p, session) for p in payments]


def _deduct_grain_from_stock(culture_id, qty, contract_label, current_user, session):
    """Списання зерна зі складу (як при виплаті оренди): спочатку віртуально
    з-під невикупленого у фермерів (farmer_quantity_kg — лічильник боргу, не
    змінюється), решта — з own_quantity_kg. Повертає (from_own, from_farmer_virtual)."""
    culture = session.get(GrainCulture, culture_id)
    stock = session.exec(select(GrainStock).where(GrainStock.culture_id == culture_id)).first()
    if not stock:
        raise HTTPException(status_code=400, detail=f"На складі немає культури '{culture.name if culture else '?'}'")
    if float(stock.quantity_kg or 0.0) < qty - EPS:
        raise HTTPException(
            status_code=400,
            detail=f"Недостатньо '{culture.name}' на складі: є {stock.quantity_kg:.2f} кг, потрібно {qty:.2f} кг"
        )
    quantity_before = stock.quantity_kg
    remaining = float(qty)
    farmer_virtual = 0.0
    fq = float(stock.farmer_quantity_kg or 0.0)
    if fq > 0:
        farmer_virtual = min(remaining, fq)
        remaining -= farmer_virtual
    oq = float(stock.own_quantity_kg or 0.0)
    take_own = min(remaining, oq)
    if remaining - take_own > EPS:
        raise HTTPException(
            status_code=400,
            detail=f"Недостатньо власного зерна '{culture.name}' на складі для цієї виплати "
                   f"(після частки {farmer_virtual:.2f} кг під «невикуплене» потрібно ще "
                   f"{remaining:.2f} кг з own, є {oq:.2f} кг)."
        )
    stock.own_quantity_kg = max(0.0, oq - take_own)
    stock.quantity_kg = max(0.0, float(stock.quantity_kg or 0.0) - float(qty))
    session.add(stock)

    dest = f"Виплата орендодавцю: {contract_label}"
    if farmer_virtual > EPS:
        dest += f" (фізично з-під невикупленого у фермерів: {farmer_virtual:.2f} кг)"
    session.add(StockAdjustmentLog(
        stock_type=StockAdjustmentType.GRAIN,
        culture_id=culture_id,
        item_name=culture.name if culture else "?",
        transaction_type=TransactionType.SUBTRACT,
        amount=float(qty),
        quantity_before=quantity_before,
        quantity_after=stock.quantity_kg,
        user_id=current_user.id,
        user_full_name=current_user.full_name,
        source="lease_payment",
        destination=dest,
    ))
    return round(take_own, 2), 0.0


def _deduct_cash(currency_str, amount, description, current_user, session):
    cash_register = session.exec(select(CashRegister)).first()
    if not cash_register:
        raise HTTPException(status_code=400, detail="Касу не знайдено")
    balance_field_map = {"UAH": "uah_balance", "USD": "usd_balance", "EUR": "eur_balance"}
    balance_field = balance_field_map.get(currency_str, "uah_balance")
    current_balance = getattr(cash_register, balance_field)
    setattr(cash_register, balance_field, current_balance - amount)
    cash_register.updated_at = datetime.utcnow()
    session.add(cash_register)
    try:
        currency_enum = Currency(currency_str)
    except ValueError:
        currency_enum = Currency.UAH
    session.add(Transaction(
        currency=currency_enum,
        amount=amount,
        transaction_type=TransactionType.SUBTRACT,
        user_id=current_user.id,
        description=description,
        uah_balance_after=cash_register.uah_balance,
        usd_balance_after=cash_register.usd_balance,
        eur_balance_after=cash_register.eur_balance,
    ))


@router.post("/payments", response_model=LeasePaymentResponse)
async def create_payment(
    payload: LeasePaymentCreate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_super_admin)
):
    """Виплата орендодавцю (на ділянку + рік)."""
    parcel = session.get(LeaseParcel, payload.parcel_id)
    if not parcel:
        raise HTTPException(status_code=404, detail="Ділянку не знайдено")
    period = session.get(LeasePeriod, payload.period_id)
    if not period or period.parcel_id != parcel.id:
        raise HTTPException(status_code=404, detail="Період не знайдено")

    if payload.payment_type not in ("grain", "cash"):
        raise HTTPException(status_code=400, detail="Тип виплати має бути 'grain' або 'cash'")

    terms = parcel.payment_terms
    # визначаємо, що гасимо
    if payload.payment_type == "grain":
        applies_to = "grain"
        if terms == "cash":
            raise HTTPException(status_code=400, detail="Ця ділянка оплачується лише грошима")
    else:  # cash
        if terms == "grain":
            applies_to = "grain"
        elif terms == "cash":
            applies_to = "cash"
        else:  # grain_cash — обираємо
            applies_to = payload.applies_to if payload.applies_to in ("grain", "cash") else "grain"

    label = f"{parcel.landlord_full_name} ({parcel.area_ha:g} га, {period.year})"

    payment = LeasePayment(
        parcel_id=parcel.id,
        period_id=period.id,
        payment_type=payload.payment_type,
        applies_to=applies_to,
        payment_date=payload.payment_date,
        note=payload.note,
        created_by_user_id=current_user.id,
    )

    # ── Зерном (фізична видача зі складу) ──
    if payload.payment_type == "grain":
        if not payload.grain_items:
            raise HTTPException(status_code=400, detail="Додайте хоча б одну позицію виплати")
        paid_map = grain_paid_kg_map(period.id, session)
        oblig = {gi.culture_id: gi for gi in session.exec(
            select(LeasePeriodGrainItem).where(LeasePeriodGrainItem.period_id == period.id)
        ).all()}
        # валідація
        clean_items = []
        for item in payload.grain_items:
            if not item.quantity_kg or item.quantity_kg <= 0:
                continue
            gi = oblig.get(item.culture_id)
            if not gi:
                culture = session.get(GrainCulture, item.culture_id)
                raise HTTPException(status_code=400, detail=f"Культура '{culture.name if culture else item.culture_id}' не входить у цей рік")
            remaining = max(0.0, float(gi.quantity_kg) - float(paid_map.get(item.culture_id, 0.0)))
            if item.quantity_kg > remaining + EPS:
                culture = session.get(GrainCulture, item.culture_id)
                raise HTTPException(
                    status_code=400,
                    detail=f"Перевищено залишок для '{culture.name if culture else item.culture_id}': залишок {remaining:.2f} кг, запит {item.quantity_kg:.2f} кг"
                )
            clean_items.append(item)
        if not clean_items:
            raise HTTPException(status_code=400, detail="Вкажіть кількість хоча б для однієї культури")

        session.add(payment)
        session.flush()
        for item in clean_items:
            from_own, from_farmer = _deduct_grain_from_stock(
                item.culture_id, float(item.quantity_kg), label, current_user, session
            )
            session.add(LeasePaymentGrainItem(
                payment_id=payment.id,
                culture_id=item.culture_id,
                quantity_kg=item.quantity_kg,
                from_own_kg=from_own,
                from_farmer_kg=from_farmer,
            ))

    # ── Грошима ──
    else:
        if not payload.amount or payload.amount <= 0:
            raise HTTPException(status_code=400, detail="Вкажіть суму виплати")
        currency_str = (payload.currency or "UAH").upper()
        rate = float(payload.exchange_rate or 1.0) if currency_str != "UAH" else 1.0
        if currency_str != "UAH" and (not payload.exchange_rate or payload.exchange_rate <= 0):
            raise HTTPException(status_code=400, detail=f"Вкажіть курс {currency_str} до грн")
        amount_uah = float(payload.amount) * rate

        payment.currency = currency_str
        payment.amount = float(payload.amount)
        payment.exchange_rate = rate
        payment.amount_uah = round(amount_uah, 2)

        if applies_to == "cash":
            cash_oblig = float(period.cash_amount or 0.0) * float(getattr(period, "cash_rate", 1.0) or 1.0)
            remaining = max(0.0, cash_oblig - cash_paid_uah(period.id, session))
            if amount_uah > remaining + EPS:
                raise HTTPException(
                    status_code=400,
                    detail=f"Сума перевищує залишок грошової частини: залишок {remaining:.2f} грн"
                )
            session.add(payment)
            session.flush()
        else:  # cash -> grain: конвертуємо у кг по поточній ціні
            if not payload.grain_items:
                raise HTTPException(status_code=400, detail="Виберіть культуру для зарахування")
            culture_id = payload.grain_items[0].culture_id
            gi = session.exec(
                select(LeasePeriodGrainItem).where(
                    LeasePeriodGrainItem.period_id == period.id,
                    LeasePeriodGrainItem.culture_id == culture_id,
                )
            ).first()
            if not gi:
                culture = session.get(GrainCulture, culture_id)
                raise HTTPException(status_code=400, detail=f"Культура '{culture.name if culture else culture_id}' не входить у цей рік")
            price = current_price(parcel.id, culture_id, session)
            if price <= 0:
                raise HTTPException(status_code=400, detail="Невідома поточна ціна культури")
            equiv_kg = amount_uah / price
            paid = float(grain_paid_kg_map(period.id, session).get(culture_id, 0.0))
            remaining_kg = max(0.0, float(gi.quantity_kg) - paid)
            if equiv_kg > remaining_kg + EPS:
                culture = session.get(GrainCulture, culture_id)
                raise HTTPException(
                    status_code=400,
                    detail=f"Сума перевищує залишок для '{culture.name if culture else culture_id}': залишок {remaining_kg:.2f} кг"
                )
            session.add(payment)
            session.flush()
            session.add(LeasePaymentGrainItem(
                payment_id=payment.id,
                culture_id=culture_id,
                quantity_kg=round(equiv_kg, 2),
                from_own_kg=0.0,
                from_farmer_kg=0.0,
            ))

        _deduct_cash(
            currency_str, float(payload.amount),
            f"Виплата орендодавцю: {label}",
            current_user, session
        )

    session.commit()
    session.refresh(payment)
    return build_payment_response(payment, session)


@router.post("/payments/{payment_id}/cancel", response_model=LeasePaymentResponse)
async def cancel_payment(
    payment_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_super_admin)
):
    """Скасування виплати з поверненням на склад / в касу"""
    payment = session.get(LeasePayment, payment_id)
    if not payment:
        raise HTTPException(status_code=404, detail="Виплату не знайдено")
    if payment.is_cancelled:
        raise HTTPException(status_code=400, detail="Виплату вже було скасовано")

    parcel = session.get(LeaseParcel, payment.parcel_id)
    label = parcel.landlord_full_name if parcel else "?"

    grain_items = session.exec(
        select(LeasePaymentGrainItem).where(LeasePaymentGrainItem.payment_id == payment.id)
    ).all()

    if payment.payment_type == "grain":
        for item in grain_items:
            if not item.quantity_kg or item.quantity_kg <= 0:
                continue
            culture = session.get(GrainCulture, item.culture_id)
            stock = session.exec(select(GrainStock).where(GrainStock.culture_id == item.culture_id)).first()
            if stock:
                quantity_before = stock.quantity_kg
                from_own = float(item.from_own_kg or 0.0)
                from_farmer = float(item.from_farmer_kg or 0.0)
                total = float(item.quantity_kg or 0.0)
                if from_own <= 0 and from_farmer <= 0:
                    from_own = total
                stock.own_quantity_kg += from_own
                stock.farmer_quantity_kg += from_farmer
                stock.quantity_kg += total
                session.add(stock)
                session.add(StockAdjustmentLog(
                    stock_type=StockAdjustmentType.GRAIN,
                    culture_id=item.culture_id,
                    item_name=culture.name if culture else "?",
                    transaction_type=TransactionType.ADD,
                    amount=item.quantity_kg,
                    quantity_before=quantity_before,
                    quantity_after=stock.quantity_kg,
                    user_id=current_user.id,
                    user_full_name=current_user.full_name,
                    source="lease_payment_cancel",
                    destination=f"Скасування виплати #{payment.id} ({label})",
                ))
    elif payment.payment_type == "cash" and payment.amount:
        cash_register = session.exec(select(CashRegister)).first()
        if cash_register:
            currency_str = payment.currency or "UAH"
            balance_field_map = {"UAH": "uah_balance", "USD": "usd_balance", "EUR": "eur_balance"}
            balance_field = balance_field_map.get(currency_str, "uah_balance")
            current_balance = getattr(cash_register, balance_field)
            setattr(cash_register, balance_field, current_balance + payment.amount)
            cash_register.updated_at = datetime.utcnow()
            session.add(cash_register)
            try:
                currency_enum = Currency(currency_str)
            except ValueError:
                currency_enum = Currency.UAH
            session.add(Transaction(
                currency=currency_enum,
                amount=payment.amount,
                transaction_type=TransactionType.ADD,
                user_id=current_user.id,
                description=f"Скасування виплати #{payment.id} ({label})",
                uah_balance_after=cash_register.uah_balance,
                usd_balance_after=cash_register.usd_balance,
                eur_balance_after=cash_register.eur_balance,
            ))

    payment.is_cancelled = True
    payment.updated_at = datetime.utcnow()
    session.add(payment)
    session.commit()
    session.refresh(payment)
    return build_payment_response(payment, session)


# ===== Excel exports =====

def _excel_header_style():
    return {
        "fill": PatternFill("solid", fgColor="1F2937"),
        "font": Font(color="FFFFFF", bold=True),
        "alignment": Alignment(horizontal="center", vertical="center"),
        "border": Border(
            left=Side(style="thin", color="E5E7EB"),
            right=Side(style="thin", color="E5E7EB"),
            top=Side(style="thin", color="E5E7EB"),
            bottom=Side(style="thin", color="E5E7EB")
        )
    }


def _apply_header(sheet, headers):
    sheet.append(headers)
    style = _excel_header_style()
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = style["fill"]
        cell.font = style["font"]
        cell.alignment = style["alignment"]
        cell.border = style["border"]


def _apply_body_style(sheet, num_cols):
    alt_fill = PatternFill("solid", fgColor="F8FAFC")
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB")
    )
    for row in range(2, sheet.max_row + 1):
        row_fill = alt_fill if row % 2 == 0 else None
        for col in range(1, num_cols + 1):
            cell = sheet.cell(row=row, column=col)
            if row_fill:
                cell.fill = row_fill
            cell.border = thin_border


def _to_response(workbook, filename):
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


TERMS_LABEL = {"grain": "Зерно", "cash": "Гроші", "grain_cash": "Гроші + зерно"}


@router.get("/landlords/export")
async def export_landlords(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    search: str | None = None
):
    """Експорт орендодавців у Excel"""
    query = select(Landlord).order_by(Landlord.full_name)
    if search:
        query = query.where(Landlord.full_name.ilike(f"%{search}%"))
    landlords = session.exec(query).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Орендодавці"
    headers = ["ПІБ", "Телефон", "Дата додавання"]
    _apply_header(ws, headers)
    for l in landlords:
        ws.append([
            l.full_name,
            l.phone or "-",
            l.created_at.strftime("%d.%m.%Y") if l.created_at else "-"
        ])
    _apply_body_style(ws, len(headers))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:C{ws.max_row}"
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    return _to_response(wb, f"landlords_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx")


@router.get("/parcels/export")
async def export_parcels(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    landlord_id: int | None = None,
):
    """Експорт ділянок із накопичувальним балансом"""
    query = select(LeaseParcel).order_by(LeaseParcel.landlord_full_name, LeaseParcel.id)
    if landlord_id:
        query = query.where(LeaseParcel.landlord_id == landlord_id)
    parcels = session.exec(query).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Ділянки"
    headers = ["Орендодавець", "Кількість, га", "Примітка", "Умови", "Роки", "Накопич. борг, грн", "Статус"]
    _apply_header(ws, headers)
    for p in parcels:
        resp = build_parcel_response(p, session)
        years = ", ".join(str(per.year) for per in resp.periods) or "-"
        ws.append([
            p.landlord_full_name,
            p.area_ha,
            p.label or "-",
            TERMS_LABEL.get(p.payment_terms, p.payment_terms),
            years,
            resp.cumulative_balance_uah,
            "Активна" if p.is_active else "Неактивна",
        ])
    _apply_body_style(ws, len(headers))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{ws.max_row}"
    for idx, w in enumerate([32, 14, 24, 16, 22, 18, 12], 1):
        ws.column_dimensions[chr(64 + idx)].width = w
    return _to_response(wb, f"parcels_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx")


@router.get("/parcels/debt-export")
async def export_parcels_debt(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Excel: лише ділянки з невиплаченим залишком (накопичувальний борг)."""
    parcels = session.exec(
        select(LeaseParcel).order_by(LeaseParcel.landlord_full_name, LeaseParcel.id)
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Борг оренди"
    headers = ["Орендодавець", "Кількість, га", "Примітка", "Умови", "Залишок до виплати, грн", "Деталізація по роках"]
    _apply_header(ws, headers)
    debt_fill = PatternFill("solid", fgColor="FEF3C7")

    for p in parcels:
        resp = build_parcel_response(p, session)
        if resp.cumulative_balance_uah <= EPS:
            continue
        details = []
        for per in resp.periods:
            if per.remaining_cash_uah > EPS:
                details.append(f"{per.year}: {per.remaining_cash_uah:.2f} грн")
        ws.append([
            p.landlord_full_name,
            p.area_ha,
            p.label or "-",
            TERMS_LABEL.get(p.payment_terms, p.payment_terms),
            resp.cumulative_balance_uah,
            "; ".join(details) or "-",
        ])
    if ws.max_row == 1:
        ws.append(["Немає ділянок із невиплаченим залишком", "", "", "", "", ""])
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=5)
        if isinstance(cell.value, (int, float)) and float(cell.value) > 0:
            cell.fill = debt_fill
    _apply_body_style(ws, len(headers))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{ws.max_row}"
    for idx, w in enumerate([32, 14, 24, 16, 22, 50], 1):
        ws.column_dimensions[chr(64 + idx)].width = w
    return _to_response(wb, f"lease_debt_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx")


@router.get("/payments/export")
async def export_payments(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    landlord_id: int | None = None,
    payment_type: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    show_cancelled: bool = False
):
    """Експорт виплат у Excel"""
    query = select(LeasePayment).order_by(LeasePayment.payment_date.desc())
    if landlord_id:
        query = query.where(
            LeasePayment.parcel_id.in_(
                select(LeaseParcel.id).where(LeaseParcel.landlord_id == landlord_id)
            )
        )
    if payment_type:
        query = query.where(LeasePayment.payment_type == payment_type)
    if not show_cancelled:
        query = query.where(LeasePayment.is_cancelled == False)
    if start_date:
        try:
            sd = datetime.combine(date.fromisoformat(start_date), dtime.min)
            query = query.where(LeasePayment.payment_date >= sd)
        except ValueError:
            pass
    if end_date:
        try:
            ed = datetime.combine(date.fromisoformat(end_date), dtime(23, 59, 59))
            query = query.where(LeasePayment.payment_date <= ed)
        except ValueError:
            pass

    payments = session.exec(query).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Виплати"
    headers = ["Дата", "Орендодавець", "Ділянка (га)", "Рік", "Тип", "Сума / зерно", "Статус", "Примітка"]
    _apply_header(ws, headers)
    grain_fill = PatternFill("solid", fgColor="BBF7D0")
    cash_fill = PatternFill("solid", fgColor="DBEAFE")
    cancelled_fill = PatternFill("solid", fgColor="FECACA")

    for p in payments:
        parcel = session.get(LeaseParcel, p.parcel_id)
        period = session.get(LeasePeriod, p.period_id)
        grain_items = session.exec(
            select(LeasePaymentGrainItem).where(LeasePaymentGrainItem.payment_id == p.id)
        ).all()
        if p.payment_type == "cash":
            sum_text = f"{p.amount:.2f} {p.currency or 'UAH'}"
        elif grain_items:
            parts = []
            for gi in grain_items:
                culture = session.get(GrainCulture, gi.culture_id)
                parts.append(f"{culture.name if culture else '?'}: {gi.quantity_kg:.2f} кг")
            sum_text = "; ".join(parts)
        else:
            sum_text = "-"
        ws.append([
            p.payment_date.strftime("%d.%m.%Y") if p.payment_date else "-",
            parcel.landlord_full_name if parcel else "-",
            f"{parcel.area_ha:g}" if parcel else "-",
            period.year if period else "-",
            "Зерном" if p.payment_type == "grain" else "Грошима",
            sum_text,
            "Скасовано" if p.is_cancelled else "Активна",
            p.note or "-",
        ])
    for row in range(2, ws.max_row + 1):
        type_cell = ws.cell(row=row, column=5)
        status_cell = ws.cell(row=row, column=7)
        if status_cell.value == "Скасовано":
            status_cell.fill = cancelled_fill
        else:
            type_cell.fill = grain_fill if type_cell.value == "Зерном" else cash_fill
        type_cell.alignment = Alignment(horizontal="center")
        status_cell.alignment = Alignment(horizontal="center")
    _apply_body_style(ws, len(headers))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{ws.max_row}"
    for idx, w in enumerate([14, 30, 14, 8, 12, 40, 14, 30], 1):
        ws.column_dimensions[chr(64 + idx)].width = w
    return _to_response(wb, f"payments_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx")
