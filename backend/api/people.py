from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select
from sqlalchemy import or_
from typing import Optional
from datetime import datetime, date, time as dtime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from backend.database import get_session
from backend.models import (
    Person,
    User,
    FarmerContract,
    FarmerContractPayment,
    FarmerGrainMovement,
    GrainCulture,
    GrainOwner,
)
from backend.schemas import (
    PersonCreate,
    PersonUpdate,
    PersonResponse,
    PersonActionResponse,
)
from backend.auth import get_current_user

router = APIRouter()


# ── Стиль для Excel-звітів (узгоджений з іншими експортами) ──
_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_THIN_BORDER = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)
_ALT_FILL = PatternFill("solid", fgColor="F8FAFC")


def _style_header(sheet, headers):
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


def _style_body(sheet, headers, number_columns=()):
    for row in range(2, sheet.max_row + 1):
        row_fill = _ALT_FILL if row % 2 == 0 else None
        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=row, column=col)
            cell.border = _THIN_BORDER
            if row_fill:
                cell.fill = row_fill
            if col in number_columns:
                cell.number_format = "#,##0.00"


def _clean_name(value: str) -> str:
    return " ".join(value.split()).strip()


@router.get("", response_model=list[PersonResponse])
async def list_people(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    q: Optional[str] = Query(default=None, description="Пошук за ПІБ або телефоном"),
):
    """Список людей (не фермерів)."""
    query = select(Person).order_by(Person.full_name)
    if q:
        like = f"%{q.strip()}%"
        query = query.where(or_(Person.full_name.ilike(like), Person.phone.ilike(like)))
    return session.exec(query).all()


@router.post("", response_model=PersonResponse)
async def create_person(
    payload: PersonCreate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Створення людини. Доступно будь-якому залогіненому (як і створення фермера)."""
    name = _clean_name(payload.full_name)
    if not name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Вкажіть ПІБ")

    phone = (payload.phone or "").strip() or None
    person = Person(full_name=name, phone=phone)
    session.add(person)
    session.commit()
    session.refresh(person)
    return person


@router.patch("/{person_id}", response_model=PersonResponse)
async def update_person(
    person_id: int,
    payload: PersonUpdate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Оновлення людини."""
    person = session.get(Person, person_id)
    if not person:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Людину не знайдено")

    if payload.full_name is not None:
        name = _clean_name(payload.full_name)
        if not name:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Вкажіть ПІБ")
        person.full_name = name

    if payload.phone is not None:
        phone = payload.phone.strip() or None
        person.phone = phone

    person.updated_at = datetime.utcnow()
    session.add(person)
    session.commit()
    session.refresh(person)
    return person


@router.get("/{person_id}/actions", response_model=list[PersonActionResponse])
async def list_person_actions(
    person_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Журнал дій по людині: контракти, виплати по них, трансфери зерна на людину."""
    person = session.get(Person, person_id)
    if not person:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Людину не знайдено")

    cultures_map = {c.id: c.name for c in session.exec(select(GrainCulture)).all()}
    owners_map = {o.id: o.full_name for o in session.exec(select(GrainOwner)).all()}

    actions: list[PersonActionResponse] = []

    contracts = session.exec(
        select(FarmerContract).where(FarmerContract.person_id == person_id)
    ).all()
    contract_ids = [c.id for c in contracts]
    for c in contracts:
        actions.append(PersonActionResponse(
            id=c.id,
            action_type="contract",
            description=f"Контракт #{c.id} ({c.contract_type})",
            amount_uah=c.total_value_uah,
            quantity_kg=None,
            culture_name=None,
            created_at=c.created_at,
            related_id=c.id,
        ))

    if contract_ids:
        payments = session.exec(
            select(FarmerContractPayment)
            .where(FarmerContractPayment.contract_id.in_(contract_ids))
        ).all()
        for p in payments:
            label = p.item_name or p.payment_type
            actions.append(PersonActionResponse(
                id=p.id,
                action_type="contract_payment",
                description=f"Оплата по контракту #{p.contract_id}: {label}"
                            + (" (скасовано)" if p.is_cancelled else ""),
                amount_uah=p.amount_uah,
                quantity_kg=p.quantity_kg,
                culture_name=cultures_map.get(p.culture_id) if p.culture_id else None,
                created_at=p.payment_date or p.created_at,
                related_id=p.contract_id,
            ))

    transfers = session.exec(
        select(FarmerGrainMovement).where(FarmerGrainMovement.to_person_id == person_id)
    ).all()
    for m in transfers:
        from_label = owners_map.get(m.from_owner_id, "?")
        actions.append(PersonActionResponse(
            id=m.id,
            action_type="transfer",
            description=f"Переказ зерна від фермера: {from_label}"
                        + (f" — {m.note}" if m.note else ""),
            amount_uah=None,
            quantity_kg=m.quantity_kg,
            culture_name=cultures_map.get(m.culture_id),
            created_at=m.created_at,
            related_id=m.id,
        ))

    actions.sort(key=lambda a: a.created_at or datetime.min, reverse=True)
    return actions


@router.get("/export")
async def export_people(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Експорт списку людей у Excel."""
    people = session.exec(select(Person).order_by(Person.full_name)).all()

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Люди"
    headers = ["ПІБ", "Телефон", "Дата створення"]
    sheet.append(headers)
    _style_header(sheet, headers)

    for p in people:
        sheet.append([
            p.full_name,
            p.phone or "-",
            p.created_at.strftime("%Y-%m-%d %H:%M") if p.created_at else "-",
        ])

    _style_body(sheet, headers)
    sheet.freeze_panes = "A2"
    if sheet.max_row > 1:
        sheet.auto_filter.ref = f"A1:C{sheet.max_row}"
    for idx, width in enumerate([28, 22, 18], start=1):
        sheet.column_dimensions[chr(64 + idx)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"people_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/actions/export")
async def export_people_actions(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    person_id: Optional[int] = Query(default=None, description="Фільтр по конкретній людині"),
    action_type: Optional[str] = Query(default=None, description="contract | contract_payment | transfer"),
    start_date: Optional[str] = Query(default=None),
    end_date: Optional[str] = Query(default=None),
):
    """Експорт журналу дій по людях у Excel.

    Підтримує фільтри по людині, типу дії та діапазону дат.
    Без фільтрів — експортуються всі дії по всіх людях, від найновіших до найстаріших.
    """
    start_dt = None
    end_dt = None
    if start_date:
        try:
            start_dt = datetime.combine(date.fromisoformat(start_date), dtime.min)
        except ValueError:
            raise HTTPException(status_code=400, detail="Некоректний формат дати початку")
    if end_date:
        try:
            end_dt = datetime.combine(date.fromisoformat(end_date), dtime.max)
        except ValueError:
            raise HTTPException(status_code=400, detail="Некоректний формат дати завершення")

    people_query = select(Person)
    if person_id:
        people_query = people_query.where(Person.id == person_id)
    people = session.exec(people_query).all()
    if not people:
        # Все ж поверне порожній звіт зі заголовками
        people = []

    cultures_map = {c.id: c.name for c in session.exec(select(GrainCulture)).all()}
    owners_map = {o.id: o.full_name for o in session.exec(select(GrainOwner)).all()}
    people_map = {p.id: p.full_name for p in people}

    rows: list[dict] = []

    if not action_type or action_type == "contract":
        contracts_q = select(FarmerContract)
        if person_id:
            contracts_q = contracts_q.where(FarmerContract.person_id == person_id)
        else:
            contracts_q = contracts_q.where(FarmerContract.person_id.is_not(None))
        contracts = session.exec(contracts_q).all()
        for c in contracts:
            ts = c.created_at
            if start_dt and ts and ts < start_dt:
                continue
            if end_dt and ts and ts > end_dt:
                continue
            rows.append({
                "ts": ts,
                "person": people_map.get(c.person_id, f"#{c.person_id}"),
                "type": "Контракт",
                "description": f"Контракт #{c.id} ({c.contract_type})",
                "qty": None,
                "amount": c.total_value_uah or 0.0,
                "culture": "",
            })

    if not action_type or action_type == "contract_payment":
        # Усі платежі по контрактах, де контракт прив'язаний до людини
        person_contracts = session.exec(
            select(FarmerContract).where(FarmerContract.person_id.is_not(None))
            if not person_id
            else select(FarmerContract).where(FarmerContract.person_id == person_id)
        ).all()
        person_contract_ids = [c.id for c in person_contracts]
        person_contract_to_person = {c.id: c.person_id for c in person_contracts}
        if person_contract_ids:
            payments = session.exec(
                select(FarmerContractPayment).where(FarmerContractPayment.contract_id.in_(person_contract_ids))
            ).all()
            for p in payments:
                ts = p.payment_date or p.created_at
                if start_dt and ts and ts < start_dt:
                    continue
                if end_dt and ts and ts > end_dt:
                    continue
                pid = person_contract_to_person.get(p.contract_id)
                label = p.item_name or p.payment_type
                desc = f"Оплата по контракту #{p.contract_id}: {label}"
                if p.is_cancelled:
                    desc += " (скасовано)"
                rows.append({
                    "ts": ts,
                    "person": people_map.get(pid, f"#{pid}"),
                    "type": "Оплата",
                    "description": desc,
                    "qty": p.quantity_kg,
                    "amount": p.amount_uah or 0.0,
                    "culture": cultures_map.get(p.culture_id, "") if p.culture_id else "",
                })

    if not action_type or action_type == "transfer":
        transfers_q = select(FarmerGrainMovement).where(FarmerGrainMovement.to_person_id.is_not(None))
        if person_id:
            transfers_q = transfers_q.where(FarmerGrainMovement.to_person_id == person_id)
        transfers = session.exec(transfers_q).all()
        for m in transfers:
            ts = m.created_at
            if start_dt and ts and ts < start_dt:
                continue
            if end_dt and ts and ts > end_dt:
                continue
            from_label = owners_map.get(m.from_owner_id, "?")
            desc = f"Переказ зерна від фермера: {from_label}"
            if m.note:
                desc += f" — {m.note}"
            rows.append({
                "ts": ts,
                "person": people_map.get(m.to_person_id, f"#{m.to_person_id}"),
                "type": "Переказ зерна",
                "description": desc,
                "qty": m.quantity_kg,
                "amount": None,
                "culture": cultures_map.get(m.culture_id, ""),
            })

    rows.sort(key=lambda r: r["ts"] or datetime.min, reverse=True)

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Дії по людях"
    headers = ["Дата", "Людина", "Тип", "Опис", "Культура", "Кількість, кг", "Сума, грн"]
    sheet.append(headers)
    _style_header(sheet, headers)

    for r in rows:
        sheet.append([
            r["ts"].strftime("%Y-%m-%d %H:%M") if r["ts"] else "-",
            r["person"],
            r["type"],
            r["description"],
            r["culture"] or "-",
            r["qty"] if r["qty"] is not None else "",
            r["amount"] if r["amount"] is not None else "",
        ])

    _style_body(sheet, headers, number_columns=(6, 7))
    sheet.freeze_panes = "A2"
    if sheet.max_row > 1:
        sheet.auto_filter.ref = f"A1:G{sheet.max_row}"
    for idx, width in enumerate([18, 24, 14, 44, 18, 14, 14], start=1):
        sheet.column_dimensions[chr(64 + idx)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"people_actions_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
