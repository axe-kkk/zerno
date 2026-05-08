from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select
from typing import Optional
from datetime import datetime, date
from io import BytesIO

from backend.database import get_session
from backend.models import AgriField, User
from backend.schemas import AgriFieldCreate, AgriFieldResponse
from backend.auth import get_current_user, get_current_super_admin
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

router = APIRouter()


@router.get("", response_model=list[AgriFieldResponse])
async def list_fields(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Список полів"""
    fields = session.exec(select(AgriField).order_by(AgriField.name)).all()
    return fields


@router.post("", response_model=AgriFieldResponse)
async def create_field(
    payload: AgriFieldCreate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Створення поля вручну (власник — підприємство)"""
    name = " ".join(payload.name.split()).strip()
    if not name:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Вкажіть назву поля"
        )

    field = AgriField(
        name=name,
        owner_name="Підприємство",
        note=payload.note
    )
    session.add(field)
    session.commit()
    session.refresh(field)
    return field


@router.patch("/{field_id}", response_model=AgriFieldResponse)
async def update_field(
    field_id: int,
    payload: AgriFieldCreate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Оновлення поля"""
    field = session.get(AgriField, field_id)
    if not field:
        raise HTTPException(status_code=404, detail="Поле не знайдено")

    name = " ".join(payload.name.split()).strip()
    if not name:
        raise HTTPException(status_code=400, detail="Вкажіть назву поля")

    field.name = name
    field.note = payload.note
    field.updated_at = datetime.utcnow()
    session.add(field)
    session.commit()
    session.refresh(field)
    return field


@router.delete("/{field_id}", response_model=AgriFieldResponse)
async def delete_field(
    field_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_super_admin)
):
    """Видалення поля"""
    field = session.get(AgriField, field_id)
    if not field:
        raise HTTPException(status_code=404, detail="Поле не знайдено")

    if field.lease_contract_id:
        raise HTTPException(
            status_code=400,
            detail="Неможливо видалити поле, прив'язане до контракту оренди"
        )

    session.delete(field)
    session.commit()
    return field


@router.get("/export")
async def export_fields(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Експорт списку полів у Excel"""
    fields = session.exec(select(AgriField).order_by(AgriField.name)).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Поля"

    headers = ["Назва поля", "Власник", "Контракт", "Примітка", "Дата створення"]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB")
    )
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    for f in fields:
        contract_info = f"#{f.lease_contract_id}" if f.lease_contract_id else "-"
        created = f.created_at.strftime("%Y-%m-%d %H:%M") if f.created_at else "-"
        sheet.append([f.name, f.owner_name, contract_info, f.note or "", created])

    for row in range(2, sheet.max_row + 1):
        for col in range(1, len(headers) + 1):
            sheet.cell(row=row, column=col).border = thin_border

    sheet.freeze_panes = "A2"
    if sheet.max_row > 1:
        sheet.auto_filter.ref = f"A1:E{sheet.max_row}"
    column_widths = [28, 24, 12, 24, 18]
    for idx, w in enumerate(column_widths, start=1):
        sheet.column_dimensions[chr(64 + idx)].width = w

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"fields_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
