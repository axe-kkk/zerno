from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select
from sqlalchemy import func, or_
from typing import Optional
from datetime import datetime, date, time, timedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from backend.database import get_session
from backend.models import (
    GrainCulture,
    VehicleType,
    Driver,
    GrainOwner,
    GrainStock,
    GrainIntake,
    GrainShipment,
    DriverStat,
    AgriField,
    TransactionType,
    StockAdjustmentLog,
    StockAdjustmentType,
    FarmerGrainDeduction,
    FarmerGrainMovement,
    User
)
from backend.schemas import (
    GrainCultureResponse,
    GrainCultureCreate,
    GrainCulturePriceUpdate,
    VehicleTypeResponse,
    DriverCreate,
    DriverResponse,
    DriverUpdate,
    GrainOwnerCreate,
    GrainOwnerResponse,
    GrainStockResponse,
    FarmerBalanceItem,
    FarmerGrainDeductRequest,
    FarmerGrainTransferRequest,
    FarmerGrainMovementResponse,
    GrainIntakeCreate,
    GrainIntakeResponse,
    GrainReserveRequest,
    GrainQualityUpdateRequest,
    GrainIntakeUpdateRequest,
    GrainShipmentCreate,
    GrainShipmentUpdate,
    GrainShipmentResponse,
    DriverStatResponse,
    StockAdjustRequest,
    StockAdjustmentResponse
)
from backend.auth import get_current_user, get_current_super_admin, get_current_admin_or_manager

router = APIRouter()


def _intake_on_stock(intake: GrainIntake) -> bool:
    """Картка вже враховується на складі (немає очікування ні тари, ні % втрат)."""
    return not intake.pending_quality and not intake.pending_tare


def _get_or_create_stock(session: Session, culture_id: int) -> GrainStock:
    stock = session.exec(
        select(GrainStock).where(GrainStock.culture_id == culture_id)
    ).first()
    if not stock:
        stock = GrainStock(
            culture_id=culture_id,
            quantity_kg=0.0,
            own_quantity_kg=0.0,
            farmer_quantity_kg=0.0,
            reserved_kg=0.0
        )
        session.add(stock)
        session.commit()
        session.refresh(stock)
    return stock


def _get_or_create_driver_stat(
    session: Session,
    driver_id: int,
    vehicle_type_id: int,
    culture_id: int,
    has_trailer: bool
) -> DriverStat:
    stat = session.exec(
        select(DriverStat).where(
            DriverStat.driver_id == driver_id,
            DriverStat.vehicle_type_id == vehicle_type_id,
            DriverStat.culture_id == culture_id,
            DriverStat.has_trailer == has_trailer
        )
    ).first()
    if not stat:
        stat = DriverStat(
            driver_id=driver_id,
            vehicle_type_id=vehicle_type_id,
            culture_id=culture_id,
            has_trailer=has_trailer,
            trips=0,
            total_net_weight_kg=0.0,
            total_accepted_weight_kg=0.0
        )
        session.add(stat)
        session.commit()
        session.refresh(stat)
    return stat


def _apply_stock_delta(
    session: Session,
    culture_id: int,
    delta_kg: float,
    is_own_grain: bool = False
) -> None:
    """Обновление склада с учетом разделения на наше и фермерское зерно"""
    stock = _get_or_create_stock(session, culture_id)
    stock.quantity_kg += delta_kg
    if is_own_grain:
        stock.own_quantity_kg += delta_kg
    else:
        stock.farmer_quantity_kg += delta_kg
    session.add(stock)
    session.commit()


def _apply_driver_stat_delta(
    session: Session,
    driver_id: int,
    vehicle_type_id: int,
    culture_id: int,
    has_trailer: bool,
    delta_trips: int,
    delta_net_kg: float,
    delta_accepted_kg: float
) -> None:
    stat = _get_or_create_driver_stat(
        session,
        driver_id=driver_id,
        vehicle_type_id=vehicle_type_id,
        culture_id=culture_id,
        has_trailer=has_trailer
    )
    stat.trips += delta_trips
    stat.total_net_weight_kg += delta_net_kg
    stat.total_accepted_weight_kg += delta_accepted_kg
    session.add(stat)
    session.commit()


@router.get("/cultures", response_model=list[GrainCultureResponse])
async def list_cultures(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Список культур"""
    return session.exec(select(GrainCulture).order_by(GrainCulture.name)).all()


@router.post("/cultures", response_model=GrainCultureResponse)
async def create_culture(
    payload: GrainCultureCreate,
    session: Session = Depends(get_session),
    current_admin: User = Depends(get_current_super_admin)
):
    """Створення культури (тільки супер адмін)"""
    existing = session.exec(
        select(GrainCulture).where(GrainCulture.name == payload.name)
    ).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Така культура вже існує"
        )
    culture = GrainCulture(name=payload.name, price_per_kg=max(1.0, payload.price_per_kg))
    session.add(culture)
    session.commit()
    session.refresh(culture)
    return culture


@router.patch("/cultures/{culture_id}/price", response_model=GrainCultureResponse)
async def update_culture_price(
    culture_id: int,
    payload: GrainCulturePriceUpdate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_admin_or_manager)
):
    """Оновлення ціни культури (super_admin або manager)"""
    culture = session.get(GrainCulture, culture_id)
    if not culture:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Культуру не знайдено"
        )
    culture.price_per_kg = max(1.0, payload.price_per_kg)
    session.add(culture)
    session.commit()
    session.refresh(culture)
    return culture


@router.get("/vehicle-types", response_model=list[VehicleTypeResponse])
async def list_vehicle_types(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Список типів транспорту"""
    return session.exec(select(VehicleType).order_by(VehicleType.name)).all()


@router.get("/drivers", response_model=list[DriverResponse])
async def list_drivers(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Список водіїв підприємства"""
    return session.exec(select(Driver).order_by(Driver.full_name)).all()


@router.get("/drivers/export")
async def export_drivers(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Експорт списку водіїв у Excel"""
    drivers = session.exec(select(Driver).order_by(Driver.full_name)).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Водії"

    headers = ["№", "ПІБ", "Телефон", "Статус"]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    alt_fill = PatternFill("solid", fgColor="F8FAFC")
    for idx, driver in enumerate(drivers, start=1):
        sheet.append([
            idx,
            driver.full_name,
            driver.phone or "—",
            "Активний" if driver.is_active else "Неактивний",
        ])

    for row in range(2, sheet.max_row + 1):
        row_fill = alt_fill if row % 2 == 0 else None
        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=row, column=col)
            if row_fill:
                cell.fill = row_fill
            cell.border = thin_border

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:D{sheet.max_row}"
    sheet.column_dimensions["A"].width = 8
    sheet.column_dimensions["B"].width = 30
    sheet.column_dimensions["C"].width = 20
    sheet.column_dimensions["D"].width = 16

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"drivers_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/drivers", response_model=DriverResponse)
async def create_driver(
    payload: DriverCreate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Створення водія"""
    driver = Driver(full_name=payload.full_name, phone=payload.phone)
    session.add(driver)
    session.commit()
    session.refresh(driver)
    return driver


@router.patch("/drivers/{driver_id}", response_model=DriverResponse)
async def update_driver(
    driver_id: int,
    payload: DriverUpdate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Оновлення водія"""
    driver = session.get(Driver, driver_id)
    if not driver:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Водія не знайдено"
        )
    if payload.full_name is not None:
        cleaned_name = " ".join(payload.full_name.split()).strip()
        if not cleaned_name:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Вкажіть ПІБ водія"
            )
        driver.full_name = cleaned_name
    if payload.phone is not None:
        driver.phone = payload.phone.strip() or None
    if payload.is_active is not None:
        driver.is_active = payload.is_active

    session.add(driver)
    session.commit()
    session.refresh(driver)
    return driver


@router.delete("/drivers/{driver_id}", response_model=DriverResponse)
async def delete_driver(
    driver_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_super_admin)
):
    """Деактивація водія"""
    driver = session.get(Driver, driver_id)
    if not driver:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Водія не знайдено"
        )
    driver.is_active = False
    session.add(driver)
    session.commit()
    session.refresh(driver)
    return driver


@router.get("/owners", response_model=list[GrainOwnerResponse])
async def list_owners(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    q: Optional[str] = Query(None, description="Пошук за ПІБ")
):
    """Список власників зерна (пошук для автодоповнення)"""
    query = select(GrainOwner)
    if q:
        # Используем ilike для поиска без учета регистра
        query = query.where(GrainOwner.full_name.ilike(f"%{q}%"))
    return session.exec(query.order_by(GrainOwner.full_name)).all()


@router.get("/owners/{owner_id}/balance", response_model=list[FarmerBalanceItem])
async def get_owner_balance(
    owner_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Баланс фермера по культурах (не викуплене зерно)"""
    owner = session.get(GrainOwner, owner_id)
    if not owner:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Фермера не знайдено"
        )

    rows = session.exec(
        select(
            GrainCulture.id,
            GrainCulture.name,
            func.sum(GrainIntake.accepted_weight_kg)
        )
        .join(GrainIntake, GrainIntake.culture_id == GrainCulture.id)
        .where(
            GrainIntake.owner_id == owner_id,
            GrainIntake.is_own_grain == False,
            GrainIntake.pending_quality == False,
            GrainIntake.pending_tare == False,
        )
        .group_by(GrainCulture.id, GrainCulture.name)
        .having(func.sum(GrainIntake.accepted_weight_kg) > 0)
        .order_by(GrainCulture.name)
    ).all()

    deductions = session.exec(
        select(
            FarmerGrainDeduction.culture_id,
            func.sum(FarmerGrainDeduction.quantity_kg)
        )
        .where(FarmerGrainDeduction.owner_id == owner_id)
        .group_by(FarmerGrainDeduction.culture_id)
    ).all()
    deduction_map = {row[0]: float(row[1] or 0) for row in deductions}

    result = []
    for row in rows:
        culture_id = row[0]
        total = float(row[2] or 0)
        total -= deduction_map.get(culture_id, 0.0)
        if total <= 0:
            continue
        result.append(
            FarmerBalanceItem(
                culture_id=culture_id,
                culture_name=row[1],
                quantity_kg=total
            )
        )
    return result


@router.get("/owners/{owner_id}/balance/export")
async def export_owner_balance(
    owner_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Експорт балансу фермера по культурах у Excel"""
    owner = session.get(GrainOwner, owner_id)
    if not owner:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Фермера не знайдено"
        )

    rows = session.exec(
        select(
            GrainCulture.id,
            GrainCulture.name,
            func.sum(GrainIntake.accepted_weight_kg)
        )
        .join(GrainIntake, GrainIntake.culture_id == GrainCulture.id)
        .where(
            GrainIntake.owner_id == owner_id,
            GrainIntake.is_own_grain == False,
            GrainIntake.pending_quality == False,
            GrainIntake.pending_tare == False,
        )
        .group_by(GrainCulture.id, GrainCulture.name)
        .having(func.sum(GrainIntake.accepted_weight_kg) > 0)
        .order_by(GrainCulture.name)
    ).all()

    deductions = session.exec(
        select(
            FarmerGrainDeduction.culture_id,
            func.sum(FarmerGrainDeduction.quantity_kg)
        )
        .where(FarmerGrainDeduction.owner_id == owner_id)
        .group_by(FarmerGrainDeduction.culture_id)
    ).all()
    deduction_map = {row[0]: float(row[1] or 0) for row in deductions}

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Баланс фермера"

    headers = ["Фермер", "Культура", "Не викуплено, кг"]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB")
    )

    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    alt_fill = PatternFill("solid", fgColor="F8FAFC")

    for culture_id, culture_name, quantity in rows:
        quantity = float(quantity or 0) - deduction_map.get(culture_id, 0.0)
        if quantity <= 0:
            continue
        sheet.append([
            owner.full_name,
            culture_name,
            quantity
        ])

    for row in range(2, sheet.max_row + 1):
        row_fill = alt_fill if row % 2 == 0 else None
        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=row, column=col)
            if row_fill:
                cell.fill = row_fill
            cell.border = thin_border
            if col == 3:
                cell.number_format = "#,##0.00"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:C{sheet.max_row}"
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 20
    sheet.column_dimensions["C"].width = 18

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename = f"farmer_balance_{owner_id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/farmer-movements", response_model=list[FarmerGrainMovementResponse])
async def list_farmer_grain_movements(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Список переміщень зерна фермерів"""
    return session.exec(
        select(FarmerGrainMovement).order_by(FarmerGrainMovement.created_at.desc())
    ).all()


@router.get("/farmer-movements/export")
async def export_farmer_grain_movements(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    start_date: str | None = None,
    end_date: str | None = None,
    movement_type: str | None = None,
    owner_id: int | None = None,
    culture_id: int | None = None,
):
    """Експорт переміщень зерна фермерів у Excel"""
    query = select(FarmerGrainMovement).order_by(FarmerGrainMovement.created_at.desc())

    if start_date:
        try:
            start_dt = datetime.combine(date.fromisoformat(start_date), time.min)
        except ValueError:
            raise HTTPException(status_code=400, detail="Некоректний формат дати початку")
        query = query.where(FarmerGrainMovement.created_at >= start_dt)
    if end_date:
        try:
            end_dt = datetime.combine(date.fromisoformat(end_date), time.max)
        except ValueError:
            raise HTTPException(status_code=400, detail="Некоректний формат дати завершення")
        query = query.where(FarmerGrainMovement.created_at <= end_dt)
    if movement_type:
        query = query.where(FarmerGrainMovement.movement_type == movement_type)
    if owner_id:
        query = query.where(
            (FarmerGrainMovement.from_owner_id == owner_id) |
            (FarmerGrainMovement.to_owner_id == owner_id)
        )
    if culture_id:
        query = query.where(FarmerGrainMovement.culture_id == culture_id)

    movements = session.exec(query).all()

    owner_map = {o.id: o.full_name for o in session.exec(select(GrainOwner)).all()}
    culture_map = {c.id: c.name for c in session.exec(select(GrainCulture)).all()}

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Переміщення зерна"

    headers = ["Дата", "Тип", "Від фермера", "До фермера", "Культура", "Кількість, кг", "Примітка"]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    alt_fill = PatternFill("solid", fgColor="F8FAFC")
    deduct_fill = PatternFill("solid", fgColor="FFEDD5")
    transfer_fill = PatternFill("solid", fgColor="DBEAFE")

    for m in movements:
        type_label = "Переміщення" if m.movement_type == "transfer" else "Списання"
        sheet.append([
            m.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            type_label,
            owner_map.get(m.from_owner_id, "—"),
            owner_map.get(m.to_owner_id, "—") if m.to_owner_id else "—",
            culture_map.get(m.culture_id, "—"),
            m.quantity_kg,
            m.note or "",
        ])

    for row in range(2, sheet.max_row + 1):
        row_fill = alt_fill if row % 2 == 0 else None
        type_val = sheet.cell(row=row, column=2).value
        type_fill = transfer_fill if type_val == "Переміщення" else deduct_fill
        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=row, column=col)
            if col == 2:
                cell.fill = type_fill
                cell.alignment = Alignment(horizontal="center")
            elif row_fill:
                cell.fill = row_fill
            cell.border = thin_border
            if col == 6:
                cell.number_format = "#,##0.00"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:G{sheet.max_row}"
    column_widths = [20, 16, 28, 28, 20, 16, 30]
    for idx, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[chr(64 + idx)].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"farmer_movements_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/farmer-movements/deduct", response_model=FarmerGrainMovementResponse)
async def deduct_farmer_grain(
    payload: FarmerGrainDeductRequest,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_super_admin)
):
    """Списання зерна з балансу фермера"""
    owner = session.get(GrainOwner, payload.owner_id)
    if not owner:
        raise HTTPException(status_code=404, detail="Фермера не знайдено")

    culture = session.get(GrainCulture, payload.culture_id)
    if not culture:
        raise HTTPException(status_code=404, detail="Культуру не знайдено")

    balance = await get_owner_balance(payload.owner_id, session, current_user)
    culture_balance = next((b for b in balance if b.culture_id == payload.culture_id), None)
    available = culture_balance.quantity_kg if culture_balance else 0.0

    if payload.quantity_kg > available:
        raise HTTPException(
            status_code=400,
            detail=f"Недостатньо зерна. Доступно: {available:.2f} кг"
        )

    deduction = FarmerGrainDeduction(
        owner_id=payload.owner_id,
        culture_id=payload.culture_id,
        quantity_kg=payload.quantity_kg,
        payment_id=None
    )
    session.add(deduction)

    movement = FarmerGrainMovement(
        movement_type="deduct",
        from_owner_id=payload.owner_id,
        to_owner_id=None,
        culture_id=payload.culture_id,
        quantity_kg=payload.quantity_kg,
        note=payload.note,
        created_by_user_id=current_user.id
    )
    session.add(movement)
    session.commit()
    session.refresh(movement)
    return movement


@router.post("/farmer-movements/transfer", response_model=FarmerGrainMovementResponse)
async def transfer_farmer_grain(
    payload: FarmerGrainTransferRequest,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_super_admin)
):
    """Переміщення зерна від фермера до іншого фермера або людини.

    Фермер → фермер: зерно лишається на складі, переходить на баланс іншого фермера
    (створюється синтетична картка приходу у отримувача, помічена `is_farmer_transfer`).
    Фермер → людина: зерно фізично залишає склад (`farmer_quantity_kg -= qty`,
    `quantity_kg -= qty`), у людини балансу не накопичується.
    """
    has_owner = bool(payload.to_owner_id)
    has_person = bool(payload.to_person_id)
    if has_owner == has_person:
        raise HTTPException(
            status_code=400,
            detail="Вкажіть отримувача — фермера або людину (рівно одне з полів)"
        )
    if has_owner and payload.from_owner_id == payload.to_owner_id:
        raise HTTPException(status_code=400, detail="Не можна переміщувати зерно самому собі")

    from_owner = session.get(GrainOwner, payload.from_owner_id)
    if not from_owner:
        raise HTTPException(status_code=404, detail="Фермера-відправника не знайдено")

    to_owner = None
    to_person = None
    receiver_label = ""
    if has_owner:
        to_owner = session.get(GrainOwner, payload.to_owner_id)
        if not to_owner:
            raise HTTPException(status_code=404, detail="Фермера-отримувача не знайдено")
        receiver_label = to_owner.full_name
    else:
        from backend.models import Person
        to_person = session.get(Person, payload.to_person_id)
        if not to_person:
            raise HTTPException(status_code=404, detail="Людину не знайдено")
        receiver_label = to_person.full_name

    culture = session.get(GrainCulture, payload.culture_id)
    if not culture:
        raise HTTPException(status_code=404, detail="Культуру не знайдено")

    balance = await get_owner_balance(payload.from_owner_id, session, current_user)
    culture_balance = next((b for b in balance if b.culture_id == payload.culture_id), None)
    available = culture_balance.quantity_kg if culture_balance else 0.0

    if payload.quantity_kg > available:
        raise HTTPException(
            status_code=400,
            detail=f"Недостатньо зерна у відправника. Доступно: {available:.2f} кг"
        )

    deduction = FarmerGrainDeduction(
        owner_id=payload.from_owner_id,
        culture_id=payload.culture_id,
        quantity_kg=payload.quantity_kg,
        payment_id=None
    )
    session.add(deduction)

    if to_owner:
        # Фермер → фермер: зерно фізично лишається на складі, отримувач накопичує баланс.
        to_intake = GrainIntake(
            culture_id=payload.culture_id,
            vehicle_type_id=1,
            owner_id=payload.to_owner_id,
            is_own_grain=False,
            owner_full_name=to_owner.full_name,
            owner_phone=to_owner.phone,
            is_internal_driver=False,
            gross_weight_kg=payload.quantity_kg,
            tare_weight_kg=0,
            net_weight_kg=payload.quantity_kg,
            accepted_weight_kg=payload.quantity_kg,
            pending_quality=False,
            pending_tare=False,
            impurity_percent=0,
            has_trailer=False,
            note=f"Трансфер від {from_owner.full_name}",
            is_farmer_transfer=True,
            created_by_user_id=current_user.id
        )
        session.add(to_intake)
    else:
        # Фермер → людина: людина фізично забирає зерно зі складу.
        stock = session.exec(
            select(GrainStock).where(GrainStock.culture_id == payload.culture_id)
        ).first()
        if stock:
            stock.farmer_quantity_kg = max(0.0, stock.farmer_quantity_kg - payload.quantity_kg)
            stock.quantity_kg = max(0.0, stock.quantity_kg - payload.quantity_kg)
            session.add(stock)
            session.add(StockAdjustmentLog(
                stock_type=StockAdjustmentType.GRAIN,
                culture_id=payload.culture_id,
                item_name=culture.name,
                transaction_type=TransactionType.SUBTRACT,
                amount=payload.quantity_kg,
                quantity_before=stock.quantity_kg + payload.quantity_kg,
                quantity_after=stock.quantity_kg,
                user_id=current_user.id,
                user_full_name=current_user.full_name,
                source="farmer_transfer_to_person",
                destination=f"Людина: {receiver_label}"
            ))

    movement = FarmerGrainMovement(
        movement_type="transfer",
        from_owner_id=payload.from_owner_id,
        to_owner_id=payload.to_owner_id,
        to_person_id=payload.to_person_id,
        culture_id=payload.culture_id,
        quantity_kg=payload.quantity_kg,
        note=payload.note,
        created_by_user_id=current_user.id
    )
    session.add(movement)
    session.commit()
    session.refresh(movement)
    return movement


@router.post("/owners", response_model=GrainOwnerResponse)
async def create_owner(
    payload: GrainOwnerCreate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Створення власника"""
    owner = GrainOwner(full_name=payload.full_name, phone=payload.phone)
    session.add(owner)
    session.commit()
    session.refresh(owner)
    return owner


@router.patch("/owners/{owner_id}", response_model=GrainOwnerResponse)
async def update_owner(
    owner_id: int,
    payload: GrainOwnerCreate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Оновлення фермера"""
    owner = session.get(GrainOwner, owner_id)
    if not owner:
        raise HTTPException(status_code=404, detail="Фермера не знайдено")
    owner.full_name = payload.full_name
    owner.phone = payload.phone
    session.add(owner)

    intakes = session.exec(
        select(GrainIntake).where(GrainIntake.owner_id == owner_id)
    ).all()
    for intake in intakes:
        intake.owner_full_name = payload.full_name
        intake.owner_phone = payload.phone
        session.add(intake)

    session.commit()
    session.refresh(owner)
    return owner


@router.get("/owners/export")
async def export_owners(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    search: Optional[str] = Query(None, description="Пошук за ПІБ")
):
    """Експорт списку фермерів у Excel"""
    query = select(GrainOwner).order_by(GrainOwner.full_name)
    if search:
        query = query.where(GrainOwner.full_name.ilike(f"%{search}%"))

    owners = session.exec(query).all()

    # Підрахувати кількість приходів та загальну вагу для кожного фермера
    intakes = session.exec(select(GrainIntake).where(GrainIntake.is_own_grain == False)).all()
    culture_map = {c.id: c.name for c in session.exec(select(GrainCulture)).all()}

    owner_stats: dict[int, dict] = {}
    for intake in intakes:
        if intake.owner_id not in owner_stats:
            owner_stats[intake.owner_id] = {"count": 0, "total_kg": 0.0}
        owner_stats[intake.owner_id]["count"] += 1
        if _intake_on_stock(intake):
            owner_stats[intake.owner_id]["total_kg"] += intake.accepted_weight_kg or 0.0

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Фермери"

    headers = ["ID", "ПІБ", "Телефон", "Кількість приходів", "Загальна вага, кг"]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB")
    )

    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    alt_fill = PatternFill("solid", fgColor="F8FAFC")

    for owner in owners:
        stats = owner_stats.get(owner.id, {"count": 0, "total_kg": 0.0})
        sheet.append([
            owner.id,
            owner.full_name,
            owner.phone or "-",
            stats["count"],
            round(stats["total_kg"], 2)
        ])

    for row in range(2, sheet.max_row + 1):
        row_fill = alt_fill if row % 2 == 0 else None
        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=row, column=col)
            if row_fill:
                cell.fill = row_fill
            cell.border = thin_border
            if col == 5:
                cell.number_format = "#,##0.00"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:E{sheet.max_row}"

    column_widths = [8, 30, 18, 20, 20]
    for idx, width in enumerate(column_widths, start=1):
        col_letter = chr(64 + idx)
        sheet.column_dimensions[col_letter].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename = f"farmers_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/owners/intakes/export")
async def export_owner_intakes(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    owner_id: Optional[int] = Query(None, description="Фільтр за фермером"),
    culture_id: Optional[int] = Query(None, description="Фільтр за культурою"),
    start_date: Optional[str] = Query(None, description="Дата початку (ISO)"),
    end_date: Optional[str] = Query(None, description="Дата завершення (ISO)"),
    period: Optional[str] = Query(None, description="Період: today|week|month")
):
    """Експорт приходів фермерів у Excel"""
    query = select(GrainIntake).where(
        GrainIntake.is_own_grain == False,
        GrainIntake.is_farmer_transfer == False
    ).order_by(GrainIntake.created_at.desc())

    if owner_id:
        query = query.where(GrainIntake.owner_id == owner_id)
    if culture_id:
        query = query.where(GrainIntake.culture_id == culture_id)

    if start_date:
        try:
            start_dt = datetime.combine(date.fromisoformat(start_date), time.min)
            query = query.where(GrainIntake.created_at >= start_dt)
        except ValueError:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Некоректний формат дати початку")

    if end_date:
        try:
            end_dt = datetime.combine(date.fromisoformat(end_date), time.max)
            query = query.where(GrainIntake.created_at <= end_dt)
        except ValueError:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Некоректний формат дати завершення")

    if period:
        now = datetime.utcnow()
        if period == "today":
            start_dt = datetime.combine(now.date(), time.min)
            query = query.where(GrainIntake.created_at >= start_dt)
        elif period == "week":
            start_dt = datetime.combine(now.date(), time.min) - timedelta(days=7)
            query = query.where(GrainIntake.created_at >= start_dt)
        elif period == "month":
            start_dt = datetime.combine(now.date(), time.min) - timedelta(days=30)
            query = query.where(GrainIntake.created_at >= start_dt)

    intakes = session.exec(query).all()

    culture_map = {c.id: c.name for c in session.exec(select(GrainCulture)).all()}

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Приходи фермерів"

    headers = ["Дата", "Фермер", "Телефон", "Культура", "Брутто, кг", "Тара, кг", "Нетто, кг", "Втрати, %", "Прийнято, кг", "Примітка"]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB")
    )

    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    alt_fill = PatternFill("solid", fgColor="F8FAFC")

    for intake in intakes:
        culture_name = culture_map.get(intake.culture_id, "-")
        sheet.append([
            intake.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            intake.owner_full_name or "-",
            intake.owner_phone or "-",
            culture_name,
            intake.gross_weight_kg,
            intake.tare_weight_kg,
            intake.net_weight_kg,
            intake.impurity_percent,
            intake.accepted_weight_kg if _intake_on_stock(intake) else "",
            intake.note or ""
        ])

    for row in range(2, sheet.max_row + 1):
        row_fill = alt_fill if row % 2 == 0 else None
        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=row, column=col)
            if row_fill:
                cell.fill = row_fill
            cell.border = thin_border
            if col in (5, 6, 7, 9):
                cell.number_format = "#,##0.00"
            if col == 8:
                cell.number_format = "0.00"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:J{sheet.max_row}"

    column_widths = [20, 28, 16, 16, 14, 14, 14, 12, 14, 30]
    for idx, width in enumerate(column_widths, start=1):
        col_letter = chr(64 + idx)
        sheet.column_dimensions[col_letter].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename = f"farmer_intakes_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/stock", response_model=list[GrainStockResponse])
async def get_stock(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Поточні залишки по культурах"""
    cultures = session.exec(select(GrainCulture)).all()
    stock_items = session.exec(select(GrainStock)).all()
    stock_map = {item.culture_id: item for item in stock_items}

    response: list[GrainStockResponse] = []
    for culture in cultures:
        stock = stock_map.get(culture.id)
        quantity = stock.quantity_kg if stock else 0.0
        own_quantity = stock.own_quantity_kg if stock else 0.0
        farmer_quantity = stock.farmer_quantity_kg if stock else 0.0
        reserved_quantity = stock.reserved_kg if stock else 0.0
        response.append(
            GrainStockResponse(
                culture_id=culture.id,
                culture_name=culture.name,
                quantity_kg=quantity,
                own_quantity_kg=own_quantity,
                farmer_quantity_kg=farmer_quantity,
                reserved_kg=reserved_quantity,
                price_per_kg=culture.price_per_kg
            )
        )
    return response


@router.patch("/stock/{culture_id}/adjust", response_model=GrainStockResponse)
async def adjust_stock(
    culture_id: int,
    payload: StockAdjustRequest,
    session: Session = Depends(get_session),
    current_admin: User = Depends(get_current_super_admin)
):
    """Ручне коригування складу зерна"""
    culture = session.get(GrainCulture, culture_id)
    if not culture:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Культуру не знайдено"
        )

    stock = session.exec(
        select(GrainStock).where(GrainStock.culture_id == culture_id)
    ).first()
    if not stock:
        stock = GrainStock(
            culture_id=culture_id,
            quantity_kg=0.0,
            own_quantity_kg=0.0,
            farmer_quantity_kg=0.0
        )
        session.add(stock)
        session.flush()

    delta = payload.amount if payload.transaction_type == TransactionType.ADD else -payload.amount
    quantity_before = stock.quantity_kg
    new_quantity = quantity_before + delta
    if new_quantity < 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Недостатньо залишку для списання"
        )

    # Важливо: `farmer_quantity_kg` — це "невикуплене у фермерів" (борг/лічильник),
    # а фізичне зерно на складі = own + farmer.
    #
    # Тому при ручному списанні з фізичного складу спочатку зменшуємо "наше" зерно,
    # і лише коли воно закінчилось — зачіпаємо `farmer_quantity_kg`.
    # Це узгоджується з логікою відправок/виплат, де `farmer_quantity_kg` не повинен
    # "стискатися пропорційно" при зменшенні фізичної кількості.
    if payload.transaction_type == TransactionType.SUBTRACT:
        remaining = float(payload.amount)
        take_own = min(float(stock.own_quantity_kg or 0.0), remaining)
        remaining -= take_own
        take_farmer = remaining
        stock.own_quantity_kg = max(0.0, float(stock.own_quantity_kg or 0.0) - take_own)
        stock.farmer_quantity_kg = max(0.0, float(stock.farmer_quantity_kg or 0.0) - take_farmer)
        stock.quantity_kg = max(0.0, float(stock.own_quantity_kg or 0.0) + float(stock.farmer_quantity_kg or 0.0))
    else:
        # ADD: за замовчуванням додаємо в "наше" зерно
        stock.own_quantity_kg = max(0.0, float(stock.own_quantity_kg or 0.0) + float(payload.amount))
        stock.quantity_kg = max(0.0, float(stock.own_quantity_kg or 0.0) + float(stock.farmer_quantity_kg or 0.0))
    session.add(stock)
    session.add(
        StockAdjustmentLog(
            stock_type=StockAdjustmentType.GRAIN,
            culture_id=culture.id,
            purchase_stock_id=None,
            category=None,
            item_name=culture.name,
            transaction_type=payload.transaction_type,
            amount=payload.amount,
            quantity_before=quantity_before,
            quantity_after=new_quantity,
            user_id=current_admin.id,
            user_full_name=current_admin.full_name,
            source="manual"
        )
    )
    session.commit()
    session.refresh(stock)

    return GrainStockResponse(
        culture_id=culture.id,
        culture_name=culture.name,
        quantity_kg=stock.quantity_kg,
        own_quantity_kg=stock.own_quantity_kg,
        farmer_quantity_kg=stock.farmer_quantity_kg,
        reserved_kg=stock.reserved_kg,
        price_per_kg=culture.price_per_kg
    )


@router.post("/stock/{culture_id}/reserve", response_model=GrainStockResponse)
async def reserve_stock(
    culture_id: int,
    payload: GrainReserveRequest,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_super_admin)
):
    """Бронювання зерна на складі"""
    culture = session.get(GrainCulture, culture_id)
    if not culture:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Культуру не знайдено")
    stock = _get_or_create_stock(session, culture_id)
    available = stock.own_quantity_kg - stock.reserved_kg
    if payload.quantity_kg > available:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Недостатньо доступного зерна. Доступно: {available}"
        )
    stock.reserved_kg += payload.quantity_kg
    session.add(stock)
    session.commit()
    session.refresh(stock)
    return GrainStockResponse(
        culture_id=culture.id,
        culture_name=culture.name,
        quantity_kg=stock.quantity_kg,
        own_quantity_kg=stock.own_quantity_kg,
        farmer_quantity_kg=stock.farmer_quantity_kg,
        reserved_kg=stock.reserved_kg,
        price_per_kg=culture.price_per_kg
    )


@router.post("/stock/{culture_id}/reserve/release", response_model=GrainStockResponse)
async def release_stock_reserve(
    culture_id: int,
    payload: GrainReserveRequest,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_super_admin)
):
    """Зняття бронювання зерна на складі"""
    culture = session.get(GrainCulture, culture_id)
    if not culture:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Культуру не знайдено")
    stock = _get_or_create_stock(session, culture_id)
    if payload.quantity_kg > stock.reserved_kg:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Недостатньо заброньованого зерна. Заброньовано: {stock.reserved_kg}"
        )
    stock.reserved_kg -= payload.quantity_kg
    session.add(stock)
    session.commit()
    session.refresh(stock)
    return GrainStockResponse(
        culture_id=culture.id,
        culture_name=culture.name,
        quantity_kg=stock.quantity_kg,
        own_quantity_kg=stock.own_quantity_kg,
        farmer_quantity_kg=stock.farmer_quantity_kg,
        reserved_kg=stock.reserved_kg,
        price_per_kg=culture.price_per_kg
    )


@router.get("/stock/adjustments", response_model=list[StockAdjustmentResponse])
async def list_grain_stock_adjustments(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Журнал ручних коригувань складу зерна"""
    return session.exec(
        select(StockAdjustmentLog)
        .where(StockAdjustmentLog.stock_type == StockAdjustmentType.GRAIN)
        .order_by(StockAdjustmentLog.created_at.desc())
    ).all()


@router.get("/stock/adjustments/export")
async def export_stock_adjustments(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    start_date: str | None = None,
    end_date: str | None = None,
    stock_type: str | None = None,
    culture_id: int | None = None,
    source: str | None = None
):
    """Експорт журналу ручних змін складу у Excel"""
    stock_types = []
    if stock_type:
        if stock_type == 'grain':
            stock_types = [StockAdjustmentType.GRAIN]
        elif stock_type == 'purchase':
            stock_types = [StockAdjustmentType.PURCHASE]
    else:
        stock_types = [StockAdjustmentType.GRAIN, StockAdjustmentType.PURCHASE]
    
    query = select(StockAdjustmentLog).where(
        StockAdjustmentLog.stock_type.in_(stock_types)
    ).order_by(StockAdjustmentLog.created_at.desc())

    if culture_id:
        query = query.where(StockAdjustmentLog.culture_id == culture_id)

    if source:
        if source == 'manual':
            query = query.where(StockAdjustmentLog.source.is_(None))
        elif source == 'shipment':
            query = query.where(StockAdjustmentLog.source == 'shipment')
        elif source == 'intake':
            query = query.where(StockAdjustmentLog.source == 'intake')

    if start_date:
        try:
            start_dt = datetime.combine(date.fromisoformat(start_date), time.min)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Некоректний формат дати початку"
            )
        query = query.where(StockAdjustmentLog.created_at >= start_dt)

    if end_date:
        try:
            end_dt = datetime.combine(date.fromisoformat(end_date), time.max)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Некоректний формат дати завершення"
            )
        query = query.where(StockAdjustmentLog.created_at <= end_dt)

    logs = session.exec(query).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Журнал змін"

    headers = [
        "Дата",
        "Тип",
        "Позиція",
        "Зміна, кг",
        "Стало, кг",
        "Користувач",
        "Примітка"
    ]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB")
    )

    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    alt_fill = PatternFill("solid", fgColor="F8FAFC")
    category_map = {
        "fertilizer": "Добрива",
        "seed": "Посівне"
    }

    for log in logs:
        if log.stock_type == StockAdjustmentType.GRAIN:
            type_label = "Зерно"
        else:
            type_label = category_map.get(log.category, "Закупівлі")
        delta_value = log.amount if log.transaction_type == TransactionType.ADD else -log.amount
        note = ''
        if log.source == 'shipment':
            note = f"Відправка: {log.destination or '-'}"
        elif log.source == 'intake':
            note = f"Прийом: {log.destination or '-'}"
        else:
            note = 'Ручне'
        sheet.append([
            log.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            type_label,
            log.item_name,
            delta_value,
            log.quantity_after,
            log.user_full_name,
            note
        ])

    for row in range(2, sheet.max_row + 1):
        row_fill = alt_fill if row % 2 == 0 else None
        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=row, column=col)
            if row_fill:
                cell.fill = row_fill
            cell.border = thin_border
            if col == 4:
                cell.number_format = "+0.00;-0.00"
            if col == 5:
                cell.number_format = "#,##0.00"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:G{sheet.max_row}"
    column_widths = [20, 14, 28, 14, 14, 22, 30]
    for idx, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[chr(64 + idx)].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename = f"stock_adjustments_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.post("/intakes", response_model=GrainIntakeResponse)
async def create_intake(
    payload: GrainIntakeCreate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_super_admin)
):
    """Створення картки приходу"""
    culture = session.get(GrainCulture, payload.culture_id)
    if not culture:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Культуру не знайдено"
        )

    vehicle_type = session.get(VehicleType, payload.vehicle_type_id)
    if not vehicle_type:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Тип транспорту не знайдено"
        )

    if payload.is_own_grain:
        if payload.owner_id or payload.owner_full_name or payload.owner_phone:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Для зерна підприємства власник не вказується"
            )
        if not payload.field_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Оберіть поле, з якого привезли зерно"
            )
        field = session.get(AgriField, payload.field_id)
        if not field:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Поле не знайдено"
            )
        owner_id = None
        owner_full_name = None
        owner_phone = None
    else:
        owner_id = payload.owner_id
        owner_full_name = payload.owner_full_name
        owner_phone = payload.owner_phone

        if not owner_id and not owner_full_name:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Потрібно вказати власника"
            )

        if owner_id:
            owner = session.get(GrainOwner, owner_id)
            if not owner:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Власника не знайдено"
                )
        else:
            owner = session.exec(
                select(GrainOwner).where(
                    GrainOwner.full_name == owner_full_name,
                    GrainOwner.phone == owner_phone
                )
            ).first()
            if not owner:
                owner = GrainOwner(full_name=owner_full_name, phone=owner_phone)
                session.add(owner)
                session.commit()
                session.refresh(owner)
            owner_id = owner.id

        owner_full_name = owner.full_name
        owner_phone = owner.phone

    if payload.is_internal_driver:
        if not payload.driver_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Потрібно вибрати водія підприємства"
            )
        driver = session.get(Driver, payload.driver_id)
        if not driver:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Водія не знайдено"
            )
        driver_id = driver.id
        external_driver_name = None
    else:
        if payload.driver_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Для стороннього водія ID не використовується"
            )
        driver_id = None
        external_driver_name = payload.external_driver_name or "Інший водій"

    if payload.pending_tare:
        net_weight = 0.0
        tare_stored = 0.0
        accepted_weight = 0.0
    else:
        net_weight = payload.gross_weight_kg - payload.tare_weight_kg
        if net_weight <= 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Нетто має бути більше 0"
            )
        tare_stored = payload.tare_weight_kg
        if payload.pending_quality:
            accepted_weight = 0.0
        else:
            accepted_weight = net_weight * (1 - payload.impurity_percent / 100)

    field_id = payload.field_id if payload.is_own_grain else None

    intake = GrainIntake(
        culture_id=payload.culture_id,
        vehicle_type_id=payload.vehicle_type_id,
        has_trailer=payload.has_trailer,
        is_own_combine=payload.is_own_combine,
        is_own_grain=payload.is_own_grain,
        field_id=field_id,
        owner_id=owner_id,
        owner_full_name=owner_full_name,
        owner_phone=owner_phone,
        is_internal_driver=payload.is_internal_driver,
        driver_id=driver_id,
        external_driver_name=external_driver_name,
        gross_weight_kg=payload.gross_weight_kg,
        tare_weight_kg=tare_stored,
        net_weight_kg=net_weight,
        impurity_percent=payload.impurity_percent,
        pending_quality=payload.pending_quality,
        pending_tare=payload.pending_tare,
        accepted_weight_kg=accepted_weight,
        note=payload.note,
        created_by_user_id=current_user.id
    )
    session.add(intake)
    session.commit()
    session.refresh(intake)

    if not payload.pending_quality and not payload.pending_tare:
        stock = _get_or_create_stock(session, payload.culture_id)
        quantity_before = stock.quantity_kg
        _apply_stock_delta(session, payload.culture_id, accepted_weight, payload.is_own_grain)
        session.refresh(stock)
        owner_label = "Підприємство" if payload.is_own_grain else (owner_full_name or "Фермер")
        session.add(StockAdjustmentLog(
            stock_type=StockAdjustmentType.GRAIN,
            culture_id=payload.culture_id,
            purchase_stock_id=None,
            category=None,
            item_name=culture.name,
            transaction_type=TransactionType.ADD,
            amount=accepted_weight,
            quantity_before=quantity_before,
            quantity_after=stock.quantity_kg,
            user_id=current_user.id,
            user_full_name=current_user.full_name,
            source="intake",
            destination=owner_label,
        ))
        session.commit()
        if payload.is_internal_driver and driver_id:
            _apply_driver_stat_delta(
                session,
                driver_id=driver_id,
                vehicle_type_id=payload.vehicle_type_id,
                culture_id=payload.culture_id,
                has_trailer=payload.has_trailer,
                delta_trips=1,
                delta_net_kg=net_weight,
                delta_accepted_kg=accepted_weight
            )

    return intake


@router.get("/intakes", response_model=list[GrainIntakeResponse])
async def list_intakes(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    pending_only: bool = Query(False, description="Очікують % втрат або тару"),
    is_own_combine: Optional[bool] = Query(None, description="Наш комбайн: true/false, None — всі"),
    field_id: Optional[int] = Query(None, description="Фільтр по полю (прийоми зерна підприємства з цього поля)"),
    include_transfers: bool = Query(False, description="Показати синтетичні картки від трансферів між фермерами")
):
    """Список карток приходу"""
    query = select(GrainIntake).order_by(GrainIntake.created_at.desc())
    if not include_transfers:
        query = query.where(GrainIntake.is_farmer_transfer == False)
    if pending_only:
        query = query.where(
            or_(GrainIntake.pending_quality == True, GrainIntake.pending_tare == True)
        )
    if is_own_combine is not None:
        query = query.where(GrainIntake.is_own_combine == is_own_combine)
    if field_id is not None:
        query = query.where(GrainIntake.field_id == field_id)
    return session.exec(query).all()


@router.get("/intakes/summary-export")
async def export_intakes_summary(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    start_date: str | None = None,
    end_date: str | None = None,
):
    """Спец-звіт: зведена таблиця приходів по культурах з розбивкою фермери / підприємство"""
    query = (
        select(GrainIntake)
        .where(
            GrainIntake.pending_quality == False,
            GrainIntake.pending_tare == False,
            GrainIntake.is_farmer_transfer == False,
        )
        .order_by(GrainIntake.created_at.desc())
    )

    if start_date:
        try:
            start_dt = datetime.combine(date.fromisoformat(start_date), time.min)
        except ValueError:
            raise HTTPException(status_code=400, detail="Некоректний формат дати початку")
        query = query.where(GrainIntake.created_at >= start_dt)

    if end_date:
        try:
            end_dt = datetime.combine(date.fromisoformat(end_date), time.max)
        except ValueError:
            raise HTTPException(status_code=400, detail="Некоректний формат дати завершення")
        query = query.where(GrainIntake.created_at <= end_dt)

    intakes = session.exec(query).all()

    cultures = session.exec(select(GrainCulture).order_by(GrainCulture.name)).all()
    culture_map = {c.id: c.name for c in cultures}
    owners = session.exec(select(GrainOwner)).all()
    owner_map = {o.id: o.full_name for o in owners}
    drivers = session.exec(select(Driver)).all()
    driver_map = {d.id: d.full_name for d in drivers}
    vehicles = session.exec(select(VehicleType)).all()
    vehicle_map = {v.id: v.name for v in vehicles}

    data_by_culture = {}
    for c in cultures:
        data_by_culture[c.id] = {
            "name": c.name,
            "farmer_intakes": [],
            "own_intakes": [],
            "farmer_total": 0.0,
            "own_total": 0.0,
        }

    for intake in intakes:
        cid = intake.culture_id
        if cid not in data_by_culture:
            data_by_culture[cid] = {
                "name": culture_map.get(cid, f"ID {cid}"),
                "farmer_intakes": [],
                "own_intakes": [],
                "farmer_total": 0.0,
                "own_total": 0.0,
            }
        entry = data_by_culture[cid]
        dt = intake.created_at.strftime("%d.%m.%Y %H:%M") if intake.created_at else ""
        driver_name = ""
        if intake.driver_id:
            driver_name = driver_map.get(intake.driver_id, "")
        elif intake.external_driver_name:
            driver_name = intake.external_driver_name
        vehicle_name = vehicle_map.get(intake.vehicle_type_id, "")
        row = {
            "date": dt,
            "source": "",
            "driver": driver_name,
            "vehicle": vehicle_name,
            "gross": intake.gross_weight_kg,
            "tare": intake.tare_weight_kg,
            "net": intake.net_weight_kg,
            "impurity": intake.impurity_percent,
            "accepted": intake.accepted_weight_kg,
            "note": intake.note or "",
        }
        if intake.is_own_grain:
            row["source"] = "Підприємство"
            entry["own_intakes"].append(row)
            entry["own_total"] += intake.accepted_weight_kg
        else:
            owner_name = owner_map.get(intake.owner_id, intake.owner_full_name or "—")
            row["source"] = owner_name
            entry["farmer_intakes"].append(row)
            entry["farmer_total"] += intake.accepted_weight_kg

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Спец. звіт"

    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )
    culture_fill = PatternFill("solid", fgColor="1F2937")
    culture_font = Font(color="FFFFFF", bold=True, size=13)
    col_header_fill = PatternFill("solid", fgColor="374151")
    col_header_font = Font(color="FFFFFF", bold=True)
    col_header_alignment = Alignment(horizontal="center", vertical="center")
    subtitle_fill = PatternFill("solid", fgColor="DBEAFE")
    subtitle_font = Font(bold=True, color="1E40AF")
    total_fill = PatternFill("solid", fgColor="E2E8F0")
    total_font = Font(bold=True)
    grand_total_fill = PatternFill("solid", fgColor="BBF7D0")
    grand_total_font = Font(bold=True, color="166534", size=11)

    detail_headers = [
        "Дата", "Власник / Джерело", "Водій", "Транспорт",
        "Брутто, кг", "Тара, кг", "Нетто, кг", "Втрати, %", "Прийнято, кг", "Примітка"
    ]
    ncols = len(detail_headers)
    weight_cols = {5, 6, 7, 9}

    def style_data_row(row_num):
        for col in range(1, ncols + 1):
            cell = sheet.cell(row=row_num, column=col)
            cell.border = thin_border
            if col in weight_cols:
                cell.number_format = "#,##0.00"
            elif col == 8:
                cell.number_format = "0.00"

    def write_culture_header(name):
        sheet.append([name] + [""] * (ncols - 1))
        r = sheet.max_row
        sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        for col in range(1, ncols + 1):
            cell = sheet.cell(row=r, column=col)
            cell.fill = culture_fill
            cell.font = culture_font
            cell.border = thin_border

    def write_col_headers():
        sheet.append(detail_headers)
        r = sheet.max_row
        for col in range(1, ncols + 1):
            cell = sheet.cell(row=r, column=col)
            cell.fill = col_header_fill
            cell.font = col_header_font
            cell.alignment = col_header_alignment
            cell.border = thin_border

    def write_section_label(label):
        sheet.append([label] + [""] * (ncols - 1))
        r = sheet.max_row
        sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        for col in range(1, ncols + 1):
            cell = sheet.cell(row=r, column=col)
            cell.fill = subtitle_fill
            cell.font = subtitle_font
            cell.border = thin_border

    def write_data_rows(rows):
        for rd in rows:
            sheet.append([
                rd["date"], rd["source"], rd["driver"], rd["vehicle"],
                rd["gross"], rd["tare"], rd["net"], rd["impurity"],
                rd["accepted"], rd["note"]
            ])
            style_data_row(sheet.max_row)

    def write_subtotal(label, total):
        row_vals = ["", label, "", "", "", "", "", "", total, ""]
        sheet.append(row_vals)
        r = sheet.max_row
        for col in range(1, ncols + 1):
            cell = sheet.cell(row=r, column=col)
            cell.fill = total_fill
            cell.font = total_font
            cell.border = thin_border
            if col == 9:
                cell.number_format = "#,##0.00"

    def write_culture_total(name, total):
        row_vals = ["", f"ВСЬОГО {name}:", "", "", "", "", "", "", total, ""]
        sheet.append(row_vals)
        r = sheet.max_row
        for col in range(1, ncols + 1):
            cell = sheet.cell(row=r, column=col)
            cell.fill = grand_total_fill
            cell.font = grand_total_font
            cell.border = thin_border
            if col == 9:
                cell.number_format = "#,##0.00"

    grand_farmer = 0
    grand_own = 0

    for cid in [c.id for c in cultures]:
        entry = data_by_culture.get(cid)
        if not entry:
            continue
        if not entry["farmer_intakes"] and not entry["own_intakes"]:
            continue

        if sheet.max_row > 1:
            sheet.append([""] * ncols)

        write_culture_header(entry["name"])
        write_col_headers()

        if entry["farmer_intakes"]:
            write_section_label("Від фермерів")
            write_data_rows(entry["farmer_intakes"])
            write_subtotal("Разом від фермерів:", entry["farmer_total"])

        if entry["own_intakes"]:
            write_section_label("Від підприємства")
            write_data_rows(entry["own_intakes"])
            write_subtotal("Разом від підприємства:", entry["own_total"])

        culture_total = entry["farmer_total"] + entry["own_total"]
        write_culture_total(entry["name"], culture_total)
        grand_farmer += entry["farmer_total"]
        grand_own += entry["own_total"]

    sheet.append([""] * ncols)
    sheet.append(["ЗАГАЛЬНИЙ ПІДСУМОК", "", "", "", "", "", "", "", "", ""])
    r = sheet.max_row
    sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    for col in range(1, ncols + 1):
        cell = sheet.cell(row=r, column=col)
        cell.fill = culture_fill
        cell.font = culture_font
        cell.border = thin_border

    sheet.append(["", "Всього від фермерів:", "", "", "", "", "", "", grand_farmer, ""])
    r = sheet.max_row
    for col in range(1, ncols + 1):
        cell = sheet.cell(row=r, column=col)
        cell.fill = total_fill
        cell.font = total_font
        cell.border = thin_border
        if col == 9:
            cell.number_format = "#,##0.00"

    sheet.append(["", "Всього від підприємства:", "", "", "", "", "", "", grand_own, ""])
    r = sheet.max_row
    for col in range(1, ncols + 1):
        cell = sheet.cell(row=r, column=col)
        cell.fill = total_fill
        cell.font = total_font
        cell.border = thin_border
        if col == 9:
            cell.number_format = "#,##0.00"

    sheet.append(["", "ВСЬОГО:", "", "", "", "", "", "", grand_farmer + grand_own, ""])
    r = sheet.max_row
    for col in range(1, ncols + 1):
        cell = sheet.cell(row=r, column=col)
        cell.fill = grand_total_fill
        cell.font = Font(bold=True, color="166534", size=12)
        cell.border = thin_border
        if col == 9:
            cell.number_format = "#,##0.00"

    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 28
    sheet.column_dimensions["C"].width = 22
    sheet.column_dimensions["D"].width = 16
    sheet.column_dimensions["E"].width = 14
    sheet.column_dimensions["F"].width = 14
    sheet.column_dimensions["G"].width = 14
    sheet.column_dimensions["H"].width = 12
    sheet.column_dimensions["I"].width = 16
    sheet.column_dimensions["J"].width = 22

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"intakes_summary_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/intakes/export")
async def export_intakes(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    start_date: str | None = None,
    end_date: str | None = None,
    culture_id: int | None = None,
    status_filter: str = Query("all", description="all|pending|confirmed"),
    driver_id: int | None = None,
    vehicle_type_id: int | None = None,
    internal_only: bool = Query(False, description="Тільки доставки наших водіїв"),
    is_own_combine: Optional[bool] = Query(None, description="Наш комбайн: true/false, None — всі"),
    only_field_intakes: bool = Query(False, description="Тільки прийоми з полів (зерно підприємства з поля)"),
    field_id: Optional[int] = Query(None, description="Фільтр по полю (для прийомів з полів)")
):
    """Експорт карток приходу у Excel"""
    query = (
        select(GrainIntake)
        .where(GrainIntake.is_farmer_transfer == False)
        .order_by(GrainIntake.created_at.desc())
    )

    if only_field_intakes:
        query = query.where(GrainIntake.field_id.isnot(None))
    if field_id is not None:
        query = query.where(GrainIntake.field_id == field_id)

    if start_date:
        try:
            start_dt = datetime.combine(date.fromisoformat(start_date), time.min)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Некоректний формат дати початку"
            )
        query = query.where(GrainIntake.created_at >= start_dt)

    if end_date:
        try:
            end_dt = datetime.combine(date.fromisoformat(end_date), time.max)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Некоректний формат дати завершення"
            )
        query = query.where(GrainIntake.created_at <= end_dt)

    if culture_id:
        query = query.where(GrainIntake.culture_id == culture_id)

    if internal_only:
        query = query.where(GrainIntake.is_internal_driver == True)

    if driver_id:
        query = query.where(GrainIntake.driver_id == driver_id)

    if vehicle_type_id:
        query = query.where(GrainIntake.vehicle_type_id == vehicle_type_id)

    if status_filter == "pending":
        query = query.where(
            or_(GrainIntake.pending_quality == True, GrainIntake.pending_tare == True)
        )
    elif status_filter == "confirmed":
        query = query.where(
            GrainIntake.pending_quality == False,
            GrainIntake.pending_tare == False,
        )
    elif status_filter != "all":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Некоректний статус фільтра"
        )

    if is_own_combine is not None:
        query = query.where(GrainIntake.is_own_combine == is_own_combine)

    intakes = session.exec(query).all()

    culture_map = {
        item.id: item.name for item in session.exec(select(GrainCulture)).all()
    }
    vehicle_map = {
        item.id: item.name for item in session.exec(select(VehicleType)).all()
    }
    driver_items = session.exec(select(Driver)).all()
    driver_map = {item.id: item.full_name for item in driver_items}
    driver_phone_map = {item.id: item.phone for item in driver_items}
    fields = session.exec(select(AgriField)).all()
    field_map = {f.id: f.name for f in fields}

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Прихід зерна"

    headers = [
        "Дата",
        "Культура",
        "Транспорт",
        "Є причіп",
        "Наш комбайн",
        "Зерно підприємства",
        "Поле",
        "Власник",
        "Телефон власника",
        "Тип водія",
        "Водій",
        "Телефон водія",
        "Брутто, кг",
        "Тара, кг",
        "Нетто, кг",
        "Втрати, %",
        "Статус",
        "Прийнято, кг",
        "Примітка"
    ]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB")
    )

    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    alt_fill = PatternFill("solid", fgColor="F8FAFC")
    pending_fill = PatternFill("solid", fgColor="FEF3C7")
    confirmed_fill = PatternFill("solid", fgColor="BBF7D0")

    for intake in intakes:
        culture_name = culture_map.get(intake.culture_id, "-")
        vehicle_name = vehicle_map.get(intake.vehicle_type_id, "-")
        owner_name = "Підприємство" if intake.is_own_grain else (intake.owner_full_name or "-")
        owner_phone = "-" if intake.is_own_grain else (intake.owner_phone or "-")
        field_name = field_map.get(intake.field_id, "-") if intake.field_id else "-"
        driver_type = "Наш водій" if intake.is_internal_driver else "Інший водій"
        if intake.is_internal_driver:
            driver_name = driver_map.get(intake.driver_id, "-")
            driver_phone = driver_phone_map.get(intake.driver_id) or "-"
        else:
            driver_name = intake.external_driver_name or "Інший водій"
            driver_phone = "-"
        if intake.pending_tare and intake.pending_quality:
            status_label = "Очікує тару та %"
        elif intake.pending_tare:
            status_label = "Очікує тару"
        elif intake.pending_quality:
            status_label = "Очікує %"
        else:
            status_label = "Підтверджено"

        sheet.append([
            intake.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            culture_name,
            vehicle_name,
            "Так" if intake.has_trailer else "Ні",
            "Так" if intake.is_own_combine else "Ні",
            "Так" if intake.is_own_grain else "Ні",
            field_name,
            owner_name,
            owner_phone,
            driver_type,
            driver_name,
            driver_phone,
            intake.gross_weight_kg,
            intake.tare_weight_kg,
            intake.net_weight_kg,
            intake.impurity_percent,
            status_label,
            "" if (intake.pending_quality or intake.pending_tare) else intake.accepted_weight_kg,
            intake.note or ""
        ])

    status_col = 17  # "Очікує %"
    for row in range(2, sheet.max_row + 1):
        row_fill = alt_fill if row % 2 == 0 else None
        status_value = sheet.cell(row=row, column=status_col).value
        status_fill = pending_fill if status_value != "Підтверджено" else confirmed_fill

        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=row, column=col)
            if col == status_col:
                cell.fill = status_fill
                cell.alignment = Alignment(horizontal="center")
            elif row_fill:
                cell.fill = row_fill
            cell.border = thin_border
            if col in (13, 14, 15, 18):
                cell.number_format = "#,##0.00"
            if col == 16:
                cell.number_format = "0.00"

    sheet.freeze_panes = "A2"
    last_col_letter = chr(64 + len(headers)) if len(headers) <= 26 else "S"
    sheet.auto_filter.ref = f"A1:{last_col_letter}{sheet.max_row}"

    column_widths = [20, 16, 16, 10, 12, 18, 20, 24, 16, 14, 22, 16, 14, 14, 14, 12, 14, 14, 30]
    for idx, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[chr(64 + idx)].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename = f"intake_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/shipments", response_model=list[GrainShipmentResponse])
async def list_shipments(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Список відправок зерна"""
    return session.exec(
        select(GrainShipment).order_by(GrainShipment.created_at.desc())
    ).all()


@router.post("/shipments", response_model=GrainShipmentResponse)
async def create_shipment(
    payload: GrainShipmentCreate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_super_admin)
):
    """Створення відправки зерна"""
    destination = " ".join(payload.destination.split()).strip()
    if not destination:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Вкажіть куди відправляємо"
        )
    culture = session.get(GrainCulture, payload.culture_id)
    if not culture:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Культуру не знайдено"
        )
    stock = _get_or_create_stock(session, payload.culture_id)
    if stock.quantity_kg < payload.quantity_kg:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Недостатньо залишку для відправки"
        )
    quantity_before = stock.quantity_kg
    # При відправці списуємо спочатку фермерське зерно, потім наше
    # farmer_quantity_kg залишається незмінним - це просто лічильник боргу перед фермерами
    remaining = payload.quantity_kg
    if stock.farmer_quantity_kg > 0:
        farmer_deduct = min(remaining, stock.farmer_quantity_kg)
        # Не змінюємо farmer_quantity_kg - це лічильник боргу
        remaining -= farmer_deduct
    if remaining > 0:
        stock.own_quantity_kg -= remaining
    stock.quantity_kg -= payload.quantity_kg
    session.add(stock)

    if payload.payment_format not in ("none", "cash", "cashless"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Невірний формат оплати"
        )
    if payload.driver_id is not None:
        driver = session.get(Driver, payload.driver_id)
        if not driver:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Водія не знайдено"
            )
    if payload.vehicle_type_id is not None:
        vehicle_type = session.get(VehicleType, payload.vehicle_type_id)
        if not vehicle_type:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Тип транспорту не знайдено"
            )

    shipment = GrainShipment(
        culture_id=payload.culture_id,
        destination=destination,
        quantity_kg=payload.quantity_kg,
        payment_format=payload.payment_format,
        driver_id=payload.driver_id,
        vehicle_type_id=payload.vehicle_type_id,
        created_by_user_id=current_user.id
    )
    session.add(shipment)
    session.add(
        StockAdjustmentLog(
            stock_type=StockAdjustmentType.GRAIN,
            culture_id=payload.culture_id,
            purchase_stock_id=None,
            category=None,
            item_name=culture.name,
            transaction_type=TransactionType.SUBTRACT,
            amount=payload.quantity_kg,
            quantity_before=quantity_before,
            quantity_after=stock.quantity_kg,
            user_id=current_user.id,
            user_full_name=current_user.full_name,
            source="shipment",
            destination=destination
        )
    )
    session.commit()
    session.refresh(shipment)
    return shipment


@router.patch("/shipments/{shipment_id}", response_model=GrainShipmentResponse)
async def update_shipment(
    shipment_id: int,
    payload: GrainShipmentUpdate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_super_admin)
):
    """Оновлення відправки зерна"""
    shipment = session.get(GrainShipment, shipment_id)
    if not shipment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Відправку не знайдено"
        )

    new_destination = payload.destination
    if new_destination is not None:
        new_destination = " ".join(new_destination.split()).strip()
        if not new_destination:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Вкажіть куди відправляємо"
            )

    new_culture_id = payload.culture_id or shipment.culture_id
    new_quantity = payload.quantity_kg or shipment.quantity_kg

    if payload.culture_id is not None:
        culture = session.get(GrainCulture, new_culture_id)
        if not culture:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Культуру не знайдено"
            )

    if new_culture_id == shipment.culture_id:
        delta = new_quantity - shipment.quantity_kg
        if delta != 0:
            stock = _get_or_create_stock(session, new_culture_id)
            quantity_before = stock.quantity_kg
            if delta > 0:
                # Увеличиваем отправку - списываем сначала фермерское, потом наше
                # farmer_quantity_kg остается неизменным
                remaining = delta
                if stock.farmer_quantity_kg > 0:
                    farmer_deduct = min(remaining, stock.farmer_quantity_kg)
                    remaining -= farmer_deduct
                if remaining > 0:
                    stock.own_quantity_kg -= remaining
                stock.quantity_kg -= delta
                transaction_type = TransactionType.SUBTRACT
                amount = delta
            else:
                # Уменьшаем отправку - возвращаем только в наше зерно
                stock.own_quantity_kg += abs(delta)
                stock.quantity_kg += abs(delta)
                transaction_type = TransactionType.ADD
                amount = abs(delta)
            session.add(stock)
            session.add(
                StockAdjustmentLog(
                    stock_type=StockAdjustmentType.GRAIN,
                    culture_id=new_culture_id,
                    purchase_stock_id=None,
                    category=None,
                    item_name=session.get(GrainCulture, new_culture_id).name,
                    transaction_type=transaction_type,
                    amount=amount,
                    quantity_before=quantity_before,
                    quantity_after=stock.quantity_kg,
                    user_id=current_user.id,
                    user_full_name=current_user.full_name,
                    source="shipment",
                    destination=new_destination or shipment.destination
                )
            )
    else:
        # Изменение культуры - возвращаем в старую культуру, списываем из новой
        old_stock = _get_or_create_stock(session, shipment.culture_id)
        old_before = old_stock.quantity_kg
        # Возвращаем только в наше зерно (так как при отправке списывалось из нашего)
        old_stock.own_quantity_kg += shipment.quantity_kg
        old_stock.quantity_kg += shipment.quantity_kg
        session.add(old_stock)

        new_stock = _get_or_create_stock(session, new_culture_id)
        new_before = new_stock.quantity_kg
        # Списываем сначала фермерское, потом наше (farmer_quantity_kg остается неизменным)
        remaining = new_quantity
        if new_stock.farmer_quantity_kg > 0:
            farmer_deduct = min(remaining, new_stock.farmer_quantity_kg)
            remaining -= farmer_deduct
        if remaining > 0:
            new_stock.own_quantity_kg -= remaining
        new_stock.quantity_kg -= new_quantity
        session.add(new_stock)
        session.add(
            StockAdjustmentLog(
                stock_type=StockAdjustmentType.GRAIN,
                culture_id=shipment.culture_id,
                purchase_stock_id=None,
                category=None,
                item_name=session.get(GrainCulture, shipment.culture_id).name,
                transaction_type=TransactionType.ADD,
                amount=shipment.quantity_kg,
                quantity_before=old_before,
                quantity_after=old_stock.quantity_kg,
                user_id=current_user.id,
                user_full_name=current_user.full_name,
                source="shipment",
                destination=shipment.destination
            )
        )
        session.add(
            StockAdjustmentLog(
                stock_type=StockAdjustmentType.GRAIN,
                culture_id=new_culture_id,
                purchase_stock_id=None,
                category=None,
                item_name=session.get(GrainCulture, new_culture_id).name,
                transaction_type=TransactionType.SUBTRACT,
                amount=new_quantity,
                quantity_before=new_before,
                quantity_after=new_stock.quantity_kg,
                user_id=current_user.id,
                user_full_name=current_user.full_name,
                source="shipment",
                destination=new_destination or shipment.destination
            )
        )

    if new_destination is not None:
        shipment.destination = new_destination
    shipment.culture_id = new_culture_id
    shipment.quantity_kg = new_quantity

    if payload.payment_format is not None:
        if payload.payment_format not in ("none", "cash", "cashless"):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Невірний формат оплати"
            )
        shipment.payment_format = payload.payment_format
    if payload.driver_id is not None:
        driver = session.get(Driver, payload.driver_id)
        if not driver:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Водія не знайдено")
        shipment.driver_id = payload.driver_id
    if payload.vehicle_type_id is not None:
        vt = session.get(VehicleType, payload.vehicle_type_id)
        if not vt:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Тип транспорту не знайдено")
        shipment.vehicle_type_id = payload.vehicle_type_id

    session.add(shipment)
    session.commit()
    session.refresh(shipment)
    return shipment


@router.get("/shipments/export")
async def export_shipments(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    start_date: str | None = None,
    end_date: str | None = None
):
    """Експорт відправок зерна у Excel"""
    query = select(GrainShipment).order_by(GrainShipment.created_at.desc())

    if start_date:
        try:
            start_dt = datetime.combine(date.fromisoformat(start_date), time.min)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Некоректний формат дати початку"
            )
        query = query.where(GrainShipment.created_at >= start_dt)

    if end_date:
        try:
            end_dt = datetime.combine(date.fromisoformat(end_date), time.max)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Некоректний формат дати завершення"
            )
        query = query.where(GrainShipment.created_at <= end_dt)

    shipments = session.exec(query).all()
    culture_map = {
        item.id: item.name for item in session.exec(select(GrainCulture)).all()
    }
    user_map = {
        item.id: item.full_name for item in session.exec(select(User)).all()
    }

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Відправки"

    headers = ["Дата", "Куди", "Культура", "Кількість, кг", "Користувач"]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB")
    )

    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    alt_fill = PatternFill("solid", fgColor="F8FAFC")

    for shipment in shipments:
        sheet.append([
            shipment.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            shipment.destination,
            culture_map.get(shipment.culture_id, "-"),
            shipment.quantity_kg,
            user_map.get(shipment.created_by_user_id, "-")
        ])

    for row in range(2, sheet.max_row + 1):
        row_fill = alt_fill if row % 2 == 0 else None
        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=row, column=col)
            if row_fill:
                cell.fill = row_fill
            cell.border = thin_border
            if col == 4:
                cell.number_format = "#,##0.00"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:E{sheet.max_row}"

    column_widths = [20, 28, 18, 16, 22]
    for idx, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[chr(64 + idx)].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename = f"shipments_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/driver-deliveries/export")
async def export_driver_deliveries(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    start_date: str | None = None,
    end_date: str | None = None,
    driver_id: int | None = None,
    culture_id: int | None = None,
    vehicle_type_id: int | None = None,
):
    """Експорт рейсів водіїв (прийом + відправки) у Excel"""
    start_dt = None
    end_dt = None
    if start_date:
        try:
            start_dt = datetime.combine(date.fromisoformat(start_date), time.min)
        except ValueError:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Некоректний формат дати початку")
    if end_date:
        try:
            end_dt = datetime.combine(date.fromisoformat(end_date), time.max)
        except ValueError:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Некоректний формат дати завершення")

    intake_query = select(GrainIntake).where(GrainIntake.is_internal_driver == True).order_by(GrainIntake.created_at.desc())
    if start_dt:
        intake_query = intake_query.where(GrainIntake.created_at >= start_dt)
    if end_dt:
        intake_query = intake_query.where(GrainIntake.created_at <= end_dt)
    if driver_id:
        intake_query = intake_query.where(GrainIntake.driver_id == driver_id)
    if culture_id:
        intake_query = intake_query.where(GrainIntake.culture_id == culture_id)
    if vehicle_type_id:
        intake_query = intake_query.where(GrainIntake.vehicle_type_id == vehicle_type_id)
    intakes = session.exec(intake_query).all()

    ship_query = select(GrainShipment).where(GrainShipment.driver_id.isnot(None)).order_by(GrainShipment.created_at.desc())
    if start_dt:
        ship_query = ship_query.where(GrainShipment.created_at >= start_dt)
    if end_dt:
        ship_query = ship_query.where(GrainShipment.created_at <= end_dt)
    if driver_id:
        ship_query = ship_query.where(GrainShipment.driver_id == driver_id)
    if culture_id:
        ship_query = ship_query.where(GrainShipment.culture_id == culture_id)
    if vehicle_type_id:
        ship_query = ship_query.where(GrainShipment.vehicle_type_id == vehicle_type_id)
    shipments = session.exec(ship_query).all()

    culture_map = {c.id: c.name for c in session.exec(select(GrainCulture)).all()}
    vehicle_map = {v.id: v.name for v in session.exec(select(VehicleType)).all()}
    driver_items = session.exec(select(Driver)).all()
    driver_map = {d.id: d.full_name for d in driver_items}
    driver_phone_map = {d.id: d.phone for d in driver_items}

    rows = []
    for intake in intakes:
        rows.append({
            "date": intake.created_at,
            "type": "Прийом",
            "driver": driver_map.get(intake.driver_id, "-"),
            "phone": driver_phone_map.get(intake.driver_id) or "-",
            "vehicle": vehicle_map.get(intake.vehicle_type_id, "-"),
            "trailer": "Так" if intake.has_trailer else "Ні",
            "culture": culture_map.get(intake.culture_id, "-"),
            "destination": "-",
            "quantity": intake.net_weight_kg if not intake.pending_tare else "",
            "accepted": "" if not _intake_on_stock(intake) else intake.accepted_weight_kg,
        })
    for ship in shipments:
        rows.append({
            "date": ship.created_at,
            "type": "Відправка",
            "driver": driver_map.get(ship.driver_id, "-"),
            "phone": driver_phone_map.get(ship.driver_id) or "-",
            "vehicle": vehicle_map.get(ship.vehicle_type_id, "-") if ship.vehicle_type_id else "-",
            "trailer": "-",
            "culture": culture_map.get(ship.culture_id, "-"),
            "destination": ship.destination,
            "quantity": ship.quantity_kg,
            "accepted": ship.quantity_kg,
        })
    rows.sort(key=lambda r: r["date"], reverse=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Рейси водіїв"

    headers = ["Дата", "Тип", "Водій", "Телефон", "Транспорт", "Причіп", "Культура", "Куди", "Кількість, кг", "Прийнято, кг"]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    alt_fill = PatternFill("solid", fgColor="F8FAFC")
    intake_fill = PatternFill("solid", fgColor="DBEAFE")
    shipment_fill = PatternFill("solid", fgColor="FEF3C7")

    for r in rows:
        sheet.append([
            r["date"].strftime("%Y-%m-%d %H:%M:%S"),
            r["type"],
            r["driver"],
            r["phone"],
            r["vehicle"],
            r["trailer"],
            r["culture"],
            r["destination"],
            r["quantity"],
            r["accepted"],
        ])

    for row in range(2, sheet.max_row + 1):
        row_fill = alt_fill if row % 2 == 0 else None
        type_val = sheet.cell(row=row, column=2).value
        type_fill = intake_fill if type_val == "Прийом" else shipment_fill
        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=row, column=col)
            if col == 2:
                cell.fill = type_fill
                cell.alignment = Alignment(horizontal="center")
            elif row_fill:
                cell.fill = row_fill
            cell.border = thin_border
            if col in (9, 10):
                cell.number_format = "#,##0.00"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{chr(64 + len(headers))}{sheet.max_row}"
    column_widths = [20, 14, 22, 16, 16, 10, 16, 24, 16, 16]
    for idx, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[chr(64 + idx)].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"driver_deliveries_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/stock/summary-export")
async def export_stock_summary(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Спец-звіт: зведена таблиця по культурах (склад, фермерське, власне)"""
    cultures = session.exec(select(GrainCulture).order_by(GrainCulture.name)).all()
    stocks = session.exec(select(GrainStock)).all()
    stock_map = {s.culture_id: s for s in stocks}

    all_owners = session.exec(select(GrainOwner)).all()
    owner_map = {o.id: o for o in all_owners}

    intake_farmer_totals = session.exec(
        select(
            GrainIntake.culture_id,
            GrainIntake.owner_id,
            func.sum(GrainIntake.accepted_weight_kg)
        )
        .where(
            GrainIntake.is_own_grain == False,
            GrainIntake.pending_quality == False,
            GrainIntake.pending_tare == False,
        )
        .group_by(GrainIntake.culture_id, GrainIntake.owner_id)
    ).all()

    deductions = session.exec(
        select(
            FarmerGrainDeduction.culture_id,
            FarmerGrainDeduction.owner_id,
            func.sum(FarmerGrainDeduction.quantity_kg)
        )
        .group_by(FarmerGrainDeduction.culture_id, FarmerGrainDeduction.owner_id)
    ).all()
    deduction_map = {}
    for cid, oid, qty in deductions:
        deduction_map[(cid, oid)] = float(qty or 0)

    farmer_detail_by_culture = {}
    for cid, oid, total_qty in intake_farmer_totals:
        total = float(total_qty or 0) - deduction_map.get((cid, oid), 0)
        if total <= 0:
            continue
        if cid not in farmer_detail_by_culture:
            farmer_detail_by_culture[cid] = []
        owner = owner_map.get(oid)
        farmer_detail_by_culture[cid].append({
            "owner_name": owner.full_name if owner else f"ID {oid}",
            "quantity_kg": total
        })

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Зведений звіт"

    headers = ["Культура", "На складі, кг", "Власне, кг", "Не викуплено у фермерів, кг", "Забронировано, кг", "Ціна, грн/кг"]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    alt_fill = PatternFill("solid", fgColor="F8FAFC")
    total_stock = 0
    total_own = 0
    total_farmer = 0
    total_reserved = 0

    for culture in cultures:
        s = stock_map.get(culture.id)
        qty = s.quantity_kg if s else 0
        own = s.own_quantity_kg if s else 0
        farmer = s.farmer_quantity_kg if s else 0
        reserved = s.reserved_kg if s else 0
        total_stock += qty
        total_own += own
        total_farmer += farmer
        total_reserved += reserved
        sheet.append([
            culture.name,
            qty,
            own,
            farmer,
            reserved,
            culture.price_per_kg
        ])

    total_row = ["РАЗОМ", total_stock, total_own, total_farmer, total_reserved, ""]
    sheet.append(total_row)
    total_row_idx = sheet.max_row
    total_fill = PatternFill("solid", fgColor="E2E8F0")
    total_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=total_row_idx, column=col)
        cell.fill = total_fill
        cell.font = total_font
        cell.border = thin_border
        if col in (2, 3, 4, 5):
            cell.number_format = "#,##0.00"

    for row in range(2, total_row_idx):
        row_fill = alt_fill if row % 2 == 0 else None
        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=row, column=col)
            if row_fill:
                cell.fill = row_fill
            cell.border = thin_border
            if col in (2, 3, 4, 5, 6):
                cell.number_format = "#,##0.00"

    detail_sheet = workbook.create_sheet("Деталі по фермерах")
    detail_headers = ["Культура", "Фермер", "Не викуплено, кг"]
    detail_sheet.append(detail_headers)
    for col in range(1, len(detail_headers) + 1):
        cell = detail_sheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    for culture in cultures:
        details = farmer_detail_by_culture.get(culture.id, [])
        for d in sorted(details, key=lambda x: x["quantity_kg"], reverse=True):
            detail_sheet.append([culture.name, d["owner_name"], d["quantity_kg"]])

    for row in range(2, detail_sheet.max_row + 1):
        row_fill = alt_fill if row % 2 == 0 else None
        for col in range(1, len(detail_headers) + 1):
            cell = detail_sheet.cell(row=row, column=col)
            if row_fill:
                cell.fill = row_fill
            cell.border = thin_border
            if col == 3:
                cell.number_format = "#,##0.00"

    detail_sheet.column_dimensions["A"].width = 20
    detail_sheet.column_dimensions["B"].width = 30
    detail_sheet.column_dimensions["C"].width = 20
    detail_sheet.freeze_panes = "A2"
    detail_sheet.auto_filter.ref = f"A1:C{detail_sheet.max_row}"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:F{sheet.max_row}"
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["C"].width = 18
    sheet.column_dimensions["D"].width = 28
    sheet.column_dimensions["E"].width = 20
    sheet.column_dimensions["F"].width = 16

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"stock_summary_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/stock/export")
async def export_grain_stock(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Експорт залишків зерна на складі у Excel"""
    cultures = session.exec(select(GrainCulture)).all()
    stock_items = session.exec(select(GrainStock)).all()
    stock_map = {item.culture_id: item for item in stock_items}

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Склад зерна"

    headers = ["Культура", "Всього, кг", "Не викуплене у фермерів, кг", "Заброньовано, кг", "Ціна, грн/кг"]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB")
    )

    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    for culture in cultures:
        stock = stock_map.get(culture.id)
        quantity = stock.quantity_kg if stock else 0.0
        farmer_qty = stock.farmer_quantity_kg if stock else 0.0
        reserved_qty = stock.reserved_kg if stock else 0.0
        sheet.append([culture.name, quantity, farmer_qty, reserved_qty, culture.price_per_kg])

    alt_fill = PatternFill("solid", fgColor="F8FAFC")
    for row in range(2, sheet.max_row + 1):
        row_fill = alt_fill if row % 2 == 0 else None
        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=row, column=col)
            if row_fill:
                cell.fill = row_fill
            cell.border = thin_border
            if col in (2, 3, 4, 5):
                cell.number_format = "#,##0.00"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:E{sheet.max_row}"
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 16
    sheet.column_dimensions["C"].width = 20
    sheet.column_dimensions["D"].width = 18
    sheet.column_dimensions["E"].width = 16

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename = f"grain_stock_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.patch("/intakes/{intake_id}/quality", response_model=GrainIntakeResponse)
async def update_intake_quality(
    intake_id: int,
    payload: GrainQualityUpdateRequest,
    session: Session = Depends(get_session),
    current_admin: User = Depends(get_current_super_admin)
):
    """Оновлення відсотку втрат і включення в статистику"""
    intake = session.get(GrainIntake, intake_id)
    if not intake:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Картку не знайдено"
        )

    if intake.pending_tare:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Спочатку вкажіть тару в картці приходу",
        )

    new_accepted = intake.net_weight_kg * (1 - payload.impurity_percent / 100)
    old_accepted = intake.accepted_weight_kg
    was_pending = intake.pending_quality

    intake.impurity_percent = payload.impurity_percent
    intake.pending_quality = False
    intake.accepted_weight_kg = new_accepted
    session.add(intake)
    session.commit()
    session.refresh(intake)

    culture = session.get(GrainCulture, intake.culture_id)
    culture_name = culture.name if culture else "-"
    owner_label = "Підприємство" if intake.is_own_grain else (intake.owner_full_name or "Фермер")

    if was_pending:
        stock = _get_or_create_stock(session, intake.culture_id)
        quantity_before = stock.quantity_kg
        _apply_stock_delta(session, intake.culture_id, new_accepted, intake.is_own_grain)
        session.refresh(stock)
        session.add(StockAdjustmentLog(
            stock_type=StockAdjustmentType.GRAIN,
            culture_id=intake.culture_id,
            purchase_stock_id=None,
            category=None,
            item_name=culture_name,
            transaction_type=TransactionType.ADD,
            amount=new_accepted,
            quantity_before=quantity_before,
            quantity_after=stock.quantity_kg,
            user_id=current_admin.id,
            user_full_name=current_admin.full_name,
            source="intake",
            destination=owner_label,
        ))
        session.commit()
        if intake.is_internal_driver and intake.driver_id:
            _apply_driver_stat_delta(
                session,
                driver_id=intake.driver_id,
                vehicle_type_id=intake.vehicle_type_id,
                culture_id=intake.culture_id,
                has_trailer=intake.has_trailer,
                delta_trips=1,
                delta_net_kg=intake.net_weight_kg,
                delta_accepted_kg=new_accepted
            )
    else:
        delta_accepted = new_accepted - old_accepted
        if abs(delta_accepted) > 0:
            stock = _get_or_create_stock(session, intake.culture_id)
            quantity_before = stock.quantity_kg
            _apply_stock_delta(session, intake.culture_id, delta_accepted, intake.is_own_grain)
            session.refresh(stock)
            is_add = delta_accepted > 0
            session.add(StockAdjustmentLog(
                stock_type=StockAdjustmentType.GRAIN,
                culture_id=intake.culture_id,
                purchase_stock_id=None,
                category=None,
                item_name=culture_name,
                transaction_type=TransactionType.ADD if is_add else TransactionType.SUBTRACT,
                amount=abs(delta_accepted),
                quantity_before=quantity_before,
                quantity_after=stock.quantity_kg,
                user_id=current_admin.id,
                user_full_name=current_admin.full_name,
                source="intake",
                destination=owner_label,
            ))
            session.commit()
            if intake.is_internal_driver and intake.driver_id:
                _apply_driver_stat_delta(
                    session,
                    driver_id=intake.driver_id,
                    vehicle_type_id=intake.vehicle_type_id,
                    culture_id=intake.culture_id,
                    has_trailer=intake.has_trailer,
                    delta_trips=0,
                    delta_net_kg=0.0,
                    delta_accepted_kg=delta_accepted
                )

    return intake


@router.patch("/intakes/{intake_id}", response_model=GrainIntakeResponse)
async def update_intake(
    intake_id: int,
    payload: GrainIntakeUpdateRequest,
    session: Session = Depends(get_session),
    current_admin: User = Depends(get_current_super_admin)
):
    """Оновлення картки приходу (тільки супер адмін)"""
    intake = session.get(GrainIntake, intake_id)
    if not intake:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Картку не знайдено"
        )

    update_data = payload.model_dump(exclude_unset=True)

    # Зберігаємо старі значення для перерахунку
    old_culture_id = intake.culture_id
    old_vehicle_type_id = intake.vehicle_type_id
    old_has_trailer = intake.has_trailer
    old_is_internal_driver = intake.is_internal_driver
    old_driver_id = intake.driver_id
    old_on_stock = _intake_on_stock(intake)
    old_net = intake.net_weight_kg
    old_accepted = intake.accepted_weight_kg
    old_is_own_grain = intake.is_own_grain

    # Власник та поле
    if update_data.get("is_own_grain") is True:
        update_data["owner_id"] = None
        update_data["owner_full_name"] = None
        update_data["owner_phone"] = None
        field_id = update_data.get("field_id", intake.field_id)
        if not field_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Оберіть поле, з якого привезли зерно"
            )
        field = session.get(AgriField, field_id)
        if not field:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Поле не знайдено")
    elif update_data.get("is_own_grain") is False:
        update_data["field_id"] = None
        owner_id = update_data.get("owner_id", intake.owner_id)
        owner_full_name = update_data.get("owner_full_name", intake.owner_full_name)
        owner_phone = update_data.get("owner_phone", intake.owner_phone)
        if not owner_id and not owner_full_name:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Потрібно вказати власника"
            )
        if owner_id:
            owner = session.get(GrainOwner, owner_id)
            if not owner:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Власника не знайдено"
                )
            update_data["owner_full_name"] = owner.full_name
            update_data["owner_phone"] = owner.phone
        else:
            owner = session.exec(
                select(GrainOwner).where(
                    GrainOwner.full_name == owner_full_name,
                    GrainOwner.phone == owner_phone
                )
            ).first()
            if not owner:
                owner = GrainOwner(full_name=owner_full_name, phone=owner_phone)
                session.add(owner)
                session.commit()
                session.refresh(owner)
            update_data["owner_id"] = owner.id
            update_data["owner_full_name"] = owner.full_name
            update_data["owner_phone"] = owner.phone

    # Водій
    if update_data.get("is_internal_driver") is True:
        driver_id = update_data.get("driver_id", intake.driver_id)
        if not driver_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Потрібно вибрати водія підприємства"
            )
        driver = session.get(Driver, driver_id)
        if not driver:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Водія не знайдено"
            )
        update_data["external_driver_name"] = None
    elif update_data.get("is_internal_driver") is False:
        external_driver_name = update_data.get("external_driver_name", intake.external_driver_name)
        update_data["external_driver_name"] = external_driver_name or "Інший водій"
        update_data["driver_id"] = None

    # Перерахунок ваги
    gross = update_data.get("gross_weight_kg", intake.gross_weight_kg)
    tare = update_data.get("tare_weight_kg", intake.tare_weight_kg)
    pending = update_data.get("pending_quality", intake.pending_quality)
    pending_tare = update_data.get("pending_tare", intake.pending_tare)
    impurity = update_data.get("impurity_percent", intake.impurity_percent)
    if pending_tare:
        update_data["net_weight_kg"] = 0.0
        update_data["accepted_weight_kg"] = 0.0
        update_data["tare_weight_kg"] = 0.0
        update_data["pending_tare"] = True
    elif gross is not None and tare is not None:
        net = float(gross) - float(tare)
        if net <= 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Нетто має бути більше 0"
            )
        update_data["net_weight_kg"] = net
        update_data["pending_tare"] = False
        if pending:
            update_data["accepted_weight_kg"] = 0.0
        else:
            update_data["accepted_weight_kg"] = net * (1 - float(impurity) / 100)

    # Оновлюємо модель
    for field, value in update_data.items():
        setattr(intake, field, value)

    session.add(intake)
    session.commit()
    session.refresh(intake)

    # Коригуємо склад і статистику
    if old_on_stock:
        _apply_stock_delta(session, old_culture_id, -old_accepted, old_is_own_grain)
        if old_is_internal_driver and old_driver_id:
            _apply_driver_stat_delta(
                session,
                driver_id=old_driver_id,
                vehicle_type_id=old_vehicle_type_id,
                culture_id=old_culture_id,
                has_trailer=old_has_trailer,
                delta_trips=-1,
                delta_net_kg=-old_net,
                delta_accepted_kg=-old_accepted
            )

    if _intake_on_stock(intake):
        _apply_stock_delta(session, intake.culture_id, intake.accepted_weight_kg, intake.is_own_grain)
        if intake.is_internal_driver and intake.driver_id:
            _apply_driver_stat_delta(
                session,
                driver_id=intake.driver_id,
                vehicle_type_id=intake.vehicle_type_id,
                culture_id=intake.culture_id,
                has_trailer=intake.has_trailer,
                delta_trips=1,
                delta_net_kg=intake.net_weight_kg,
                delta_accepted_kg=intake.accepted_weight_kg
            )

    return intake


@router.get("/drivers/{driver_id}/stats", response_model=list[DriverStatResponse])
async def get_driver_stats(
    driver_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Статистика по водію"""
    driver = session.get(Driver, driver_id)
    if not driver:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Водія не знайдено"
        )

    stats = session.exec(
        select(DriverStat).where(DriverStat.driver_id == driver_id)
    ).all()

    vehicle_map = {
        v.id: v for v in session.exec(select(VehicleType)).all()
    }
    culture_map = {
        c.id: c for c in session.exec(select(GrainCulture)).all()
    }

    response: list[DriverStatResponse] = []
    for stat in stats:
        response.append(
            DriverStatResponse(
                driver_id=stat.driver_id,
                driver_name=driver.full_name,
                vehicle_type_id=stat.vehicle_type_id,
                vehicle_type_name=vehicle_map.get(stat.vehicle_type_id).name
                if vehicle_map.get(stat.vehicle_type_id)
                else "",
                culture_id=stat.culture_id,
                culture_name=culture_map.get(stat.culture_id).name
                if culture_map.get(stat.culture_id)
                else "",
                has_trailer=stat.has_trailer,
                trips=stat.trips,
                total_net_weight_kg=stat.total_net_weight_kg,
                total_accepted_weight_kg=stat.total_accepted_weight_kg
            )
        )

    return response

