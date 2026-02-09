from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select
from backend.database import get_session
from backend.models import CashRegister, Transaction, Currency, TransactionType, User
from backend.schemas import (
    CashRegisterResponse,
    BalanceUpdateRequest,
    TransactionResponse
)
from backend.auth import get_current_super_admin, get_current_user
from datetime import datetime, date, time
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

router = APIRouter()


def get_or_create_cash_register(session: Session) -> CashRegister:
    """Получение или создание кассы (должна быть только одна)"""
    cash_register = session.exec(select(CashRegister)).first()
    
    if not cash_register:
        cash_register = CashRegister(
            uah_balance=0.0,
            usd_balance=0.0,
            eur_balance=0.0
        )
        session.add(cash_register)
        session.commit()
        session.refresh(cash_register)
    
    return cash_register


@router.get("/balance", response_model=CashRegisterResponse)
async def get_balance(
    session: Session = Depends(get_session)
):
    """Получение текущих балансов кассы (доступно всем авторизованным)"""
    cash_register = get_or_create_cash_register(session)
    return cash_register


@router.post("/update-balance", response_model=TransactionResponse)
async def update_balance(
    update_request: BalanceUpdateRequest,
    session: Session = Depends(get_session),
    current_admin: User = Depends(get_current_super_admin)
):
    """Изменение баланса кассы (только для супер админа)"""
    cash_register = get_or_create_cash_register(session)
    
    # Определяем, какое поле баланса нужно изменить
    balance_field_map = {
        Currency.UAH: "uah_balance",
        Currency.USD: "usd_balance",
        Currency.EUR: "eur_balance"
    }
    currency_label_map = {
        Currency.UAH: "UAH (гривня)",
        Currency.USD: "USD (долар США)",
        Currency.EUR: "EUR (євро)"
    }
    
    balance_field = balance_field_map[update_request.currency]
    current_balance = getattr(cash_register, balance_field)
    
    # Вычисляем новый баланс
    if update_request.transaction_type == TransactionType.ADD:
        new_balance = current_balance + update_request.amount
    else:  # SUBTRACT
        new_balance = current_balance - update_request.amount
        # Проверка на отрицательный баланс
        if new_balance < 0:
            currency_label = currency_label_map[update_request.currency]
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Недостатньо коштів. Поточний баланс {currency_label}: {current_balance}"
            )
    
    # Обновляем баланс
    setattr(cash_register, balance_field, new_balance)
    
    # Обновляем updated_at
    from datetime import datetime
    cash_register.updated_at = datetime.utcnow()
    
    # Сохраняем изменения
    session.add(cash_register)
    session.commit()
    session.refresh(cash_register)
    
    # Создаем запись о транзакции
    transaction = Transaction(
        currency=update_request.currency,
        amount=update_request.amount,
        transaction_type=update_request.transaction_type,
        user_id=current_admin.id,
        description=update_request.description,
        uah_balance_after=cash_register.uah_balance,
        usd_balance_after=cash_register.usd_balance,
        eur_balance_after=cash_register.eur_balance
    )
    
    session.add(transaction)
    session.commit()
    session.refresh(transaction)
    
    return TransactionResponse(
        id=transaction.id,
        currency=transaction.currency,
        amount=transaction.amount,
        transaction_type=transaction.transaction_type,
        user_id=transaction.user_id,
        user_full_name=current_admin.full_name,
        description=transaction.description,
        uah_balance_after=transaction.uah_balance_after,
        usd_balance_after=transaction.usd_balance_after,
        eur_balance_after=transaction.eur_balance_after,
        created_at=transaction.created_at
    )


@router.get("/transactions", response_model=list[TransactionResponse])
async def get_transactions(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    limit: int = 100,
    offset: int = 0
):
    """Получение истории транзакций (доступно всем авторизованным пользователям)"""
    transactions = session.exec(
        select(Transaction)
        .order_by(Transaction.created_at.desc())
        .limit(limit)
        .offset(offset)
    ).all()

    user_ids = {item.user_id for item in transactions if item.user_id}
    users = session.exec(select(User).where(User.id.in_(list(user_ids)))).all() if user_ids else []
    user_map = {user.id: user.full_name for user in users}

    return [
        TransactionResponse(
            id=item.id,
            currency=item.currency,
            amount=item.amount,
            transaction_type=item.transaction_type,
            user_id=item.user_id,
            user_full_name=user_map.get(item.user_id),
            description=item.description,
            uah_balance_after=item.uah_balance_after,
            usd_balance_after=item.usd_balance_after,
            eur_balance_after=item.eur_balance_after,
            created_at=item.created_at
        )
        for item in transactions
    ]


@router.get("/transactions/export")
async def export_transactions(
    session: Session = Depends(get_session),
    current_admin: User = Depends(get_current_super_admin),
    start_date: str | None = None,
    end_date: str | None = None
):
    """Експорт транзакцій у Excel (тільки для супер адміна)"""
    query = select(Transaction).order_by(Transaction.created_at.desc())

    if start_date:
        try:
            start_dt = datetime.combine(date.fromisoformat(start_date), time.min)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Некоректний формат дати початку"
            )
        query = query.where(Transaction.created_at >= start_dt)

    if end_date:
        try:
            end_dt = datetime.combine(date.fromisoformat(end_date), time.max)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Некоректний формат дати завершення"
            )
        query = query.where(Transaction.created_at <= end_dt)

    transactions = session.exec(query).all()
    user_ids = {item.user_id for item in transactions if item.user_id}
    users = session.exec(select(User).where(User.id.in_(list(user_ids)))).all() if user_ids else []
    user_map = {user.id: user.full_name for user in users}

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Операції"

    headers = [
        "Дата",
        "Користувач",
        "Валюта",
        "Сума",
        "Тип",
        "Опис",
        "Баланс UAH",
        "Баланс USD",
        "Баланс EUR"
    ]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB")
    )

    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    type_label_map = {
        TransactionType.ADD.value: "Додано",
        TransactionType.SUBTRACT.value: "Віднято"
    }

    for item in transactions:
        sheet.append([
            item.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            user_map.get(item.user_id) or "-",
            item.currency.value,
            item.amount,
            type_label_map.get(item.transaction_type.value, item.transaction_type.value),
            item.description or "",
            item.uah_balance_after,
            item.usd_balance_after,
            item.eur_balance_after
        ])

    alt_fill = PatternFill("solid", fgColor="F8FAFC")
    add_fill = PatternFill("solid", fgColor="BBF7D0")
    subtract_fill = PatternFill("solid", fgColor="FECACA")

    for row in range(2, sheet.max_row + 1):
        row_fill = alt_fill if row % 2 == 0 else None
        transaction_type = sheet.cell(row=row, column=5).value
        type_fill = add_fill if transaction_type == type_label_map[TransactionType.ADD.value] else subtract_fill

        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=row, column=col)
            if col == 5:
                cell.fill = type_fill
                cell.alignment = Alignment(horizontal="center")
            elif col == 4:
                cell.fill = type_fill
            elif row_fill:
                cell.fill = row_fill
            cell.border = thin_border
            if col in (4, 7, 8, 9):
                cell.number_format = "#,##0.00"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:I{sheet.max_row}"

    column_widths = [20, 22, 12, 14, 12, 30, 14, 14, 14]
    for idx, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[chr(64 + idx)].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename = f"cash_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/transactions/{transaction_id}", response_model=TransactionResponse)
async def get_transaction(
    transaction_id: int,
    session: Session = Depends(get_session),
    current_admin: User = Depends(get_current_super_admin)
):
    """Получение конкретной транзакции по ID (только для супер админа)"""
    transaction = session.get(Transaction, transaction_id)
    
    if not transaction:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Транзакцію не знайдено"
        )
    
    return transaction

