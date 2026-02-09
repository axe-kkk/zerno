from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select
from typing import Optional
from datetime import datetime, date, time
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from backend.database import get_session
from backend.models import (
    PurchaseStock,
    PurchaseRecord,
    CashRegister,
    Transaction,
    Currency,
    TransactionType,
    PurchaseCategory,
    StockAdjustmentLog,
    StockAdjustmentType,
    User
)
from backend.schemas import (
    PurchaseCreate,
    PurchaseResponse,
    PurchaseStockResponse,
    PurchaseStockPriceUpdate,
    StockAdjustRequest,
    StockAdjustmentResponse
)
from backend.auth import get_current_user, get_current_super_admin
from backend.api.cash import get_or_create_cash_register

router = APIRouter()


def normalize_name(value: str) -> str:
    return " ".join(value.split()).strip().lower()


@router.get("/stock", response_model=list[PurchaseStockResponse])
async def list_purchase_stock(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Склад закупівель"""
    items = session.exec(
        select(PurchaseStock)
        .where(PurchaseStock.quantity_kg > 0)
        .order_by(PurchaseStock.category, PurchaseStock.name)
    ).all()
    return items


@router.patch("/stock/{stock_id}/price", response_model=PurchaseStockResponse)
async def update_purchase_stock_price(
    stock_id: int,
    payload: PurchaseStockPriceUpdate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Оновлення ціни продажу для складу закупівель"""
    stock = session.get(PurchaseStock, stock_id)
    if not stock:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Позицію складу не знайдено"
        )
    stock.sale_price_per_kg = payload.sale_price_per_kg
    session.add(stock)
    session.commit()
    session.refresh(stock)
    return stock


@router.patch("/stock/{stock_id}/adjust", response_model=PurchaseStockResponse)
async def adjust_purchase_stock(
    stock_id: int,
    payload: StockAdjustRequest,
    session: Session = Depends(get_session),
    current_admin: User = Depends(get_current_super_admin)
):
    """Ручне коригування складу закупівель"""
    stock = session.get(PurchaseStock, stock_id)
    if not stock:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Позицію складу не знайдено"
        )

    delta = payload.amount if payload.transaction_type == TransactionType.ADD else -payload.amount
    quantity_before = stock.quantity_kg
    new_quantity = quantity_before + delta
    if new_quantity < 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Недостатньо залишку для списання"
        )

    stock.quantity_kg = new_quantity
    session.add(stock)
    session.add(
        StockAdjustmentLog(
            stock_type=StockAdjustmentType.PURCHASE,
            culture_id=None,
            purchase_stock_id=stock.id,
            category=stock.category,
            item_name=stock.name,
            transaction_type=payload.transaction_type,
            amount=payload.amount,
            quantity_before=quantity_before,
            quantity_after=new_quantity,
            user_id=current_admin.id,
            user_full_name=current_admin.full_name,
            source="manual"
        )
    )
    session.commit()
    session.refresh(stock)
    return stock


@router.get("/stock/adjustments", response_model=list[StockAdjustmentResponse])
async def list_purchase_stock_adjustments(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Журнал ручних коригувань складу закупівель"""
    return session.exec(
        select(StockAdjustmentLog)
        .where(StockAdjustmentLog.stock_type == StockAdjustmentType.PURCHASE)
        .order_by(StockAdjustmentLog.created_at.desc())
    ).all()


@router.get("", response_model=list[PurchaseResponse])
async def list_purchases(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Історія закупівель"""
    return session.exec(
        select(PurchaseRecord).order_by(PurchaseRecord.created_at.desc())
    ).all()


@router.get("/export")
async def export_purchases(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    start_date: str | None = None,
    end_date: str | None = None
):
    """Експорт історії закупівель у Excel"""
    query = select(PurchaseRecord).order_by(PurchaseRecord.created_at.desc())

    if start_date:
        try:
            start_dt = datetime.combine(date.fromisoformat(start_date), time.min)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Некоректний формат дати початку"
            )
        query = query.where(PurchaseRecord.created_at >= start_dt)

    if end_date:
        try:
            end_dt = datetime.combine(date.fromisoformat(end_date), time.max)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Некоректний формат дати завершення"
            )
        query = query.where(PurchaseRecord.created_at <= end_dt)

    records = session.exec(query).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Закупівлі"

    headers = [
        "Дата",
        "Назва",
        "Категорія",
        "Ціна/кг",
        "Валюта",
        "Кількість, кг",
        "Сума"
    ]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB")
    )

    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    alt_fill = PatternFill("solid", fgColor="F8FAFC")

    for record in records:
        category_label = "Добрива" if record.category.value == "fertilizer" else "Посівне зерно"
        sheet.append([
            record.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            record.item_name,
            category_label,
            record.price_per_kg,
            record.currency.value,
            record.quantity_kg,
            record.total_amount
        ])

    for row in range(2, sheet.max_row + 1):
        row_fill = alt_fill if row % 2 == 0 else None
        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=row, column=col)
            if row_fill:
                cell.fill = row_fill
            cell.border = thin_border
            if col in (4, 6, 7):
                cell.number_format = "#,##0.00"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:G{sheet.max_row}"

    column_widths = [20, 28, 18, 12, 10, 14, 14]
    for idx, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[chr(64 + idx)].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename = f"purchases_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/stock/export")
async def export_purchase_stock(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    category: str | None = Query(default=None)
):
    """Експорт залишків складу закупівель у Excel"""
    query = select(PurchaseStock).where(PurchaseStock.quantity_kg > 0)
    if category:
        try:
            category_enum = PurchaseCategory(category)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Некоректна категорія"
            )
        query = query.where(PurchaseStock.category == category_enum)

    items = session.exec(query.order_by(PurchaseStock.category, PurchaseStock.name)).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Склад закупівель"

    headers = ["Категорія", "Назва", "Всього, кг", "Забронировано, кг", "Вільне, кг", "Ціна, грн/кг"]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB")
    )

    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    category_label_map = {
        PurchaseCategory.FERTILIZER: "Добрива",
        PurchaseCategory.SEED: "Посівне зерно"
    }

    for item in items:
        reserved = item.reserved_kg or 0.0
        available = item.quantity_kg - reserved
        sheet.append([
            category_label_map.get(item.category, item.category.value),
            item.name,
            item.quantity_kg,
            reserved,
            available,
            item.sale_price_per_kg or 0
        ])

    alt_fill = PatternFill("solid", fgColor="F8FAFC")
    for row in range(2, sheet.max_row + 1):
        row_fill = alt_fill if row % 2 == 0 else None
        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=row, column=col)
            if row_fill:
                cell.fill = row_fill
            cell.border = thin_border
            if col in (3, 4, 5, 6):
                cell.number_format = "#,##0.00"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:F{sheet.max_row}"
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 28
    sheet.column_dimensions["C"].width = 16
    sheet.column_dimensions["D"].width = 18
    sheet.column_dimensions["E"].width = 16
    sheet.column_dimensions["F"].width = 16

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename_suffix = f"_{category}" if category else ""
    filename = f"purchase_stock{filename_suffix}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.post("", response_model=PurchaseResponse)
async def create_purchase(
    payload: PurchaseCreate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Створення закупівлі та списання з каси"""
    cleaned_name = " ".join(payload.item_name.split()).strip()
    if not cleaned_name:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Вкажіть назву позиції"
        )
    normalized = normalize_name(cleaned_name)

    stock = session.exec(
        select(PurchaseStock).where(
            PurchaseStock.normalized_name == normalized,
            PurchaseStock.category == payload.category
        )
    ).first()
    if not stock:
        stock = PurchaseStock(
            name=cleaned_name,
            normalized_name=normalized,
            category=payload.category,
            quantity_kg=0.0
        )
        session.add(stock)
        session.flush()

    total_amount = payload.price_per_kg * payload.quantity_kg

    cash_register: CashRegister = get_or_create_cash_register(session)
    balance_field_map = {
        Currency.UAH: "uah_balance",
        Currency.USD: "usd_balance",
        Currency.EUR: "eur_balance"
    }
    currency_label_map = {
        Currency.UAH: "UAH (гривня)",
        Currency.USD: "USD (долар США)",
        Currency.EUR: "EUR (євро)"
    }
    balance_field = balance_field_map[payload.currency]
    current_balance = getattr(cash_register, balance_field)
    new_balance = current_balance - total_amount
    if new_balance < 0:
        currency_label = currency_label_map[payload.currency]
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Недостатньо коштів. Поточний баланс {currency_label}: {current_balance}"
        )

    setattr(cash_register, balance_field, new_balance)
    session.add(cash_register)

    transaction = Transaction(
        currency=payload.currency,
        amount=total_amount,
        transaction_type=TransactionType.SUBTRACT,
        user_id=current_user.id,
        description=f"Закупівля: {stock.name}",
        uah_balance_after=cash_register.uah_balance,
        usd_balance_after=cash_register.usd_balance,
        eur_balance_after=cash_register.eur_balance
    )
    session.add(transaction)

    stock.quantity_kg += payload.quantity_kg
    if stock.quantity_kg <= 0:
        session.delete(stock)
    else:
        session.add(stock)

    record = PurchaseRecord(
        stock_id=stock.id,
        item_name=stock.name,
        normalized_name=normalized,
        category=payload.category,
        price_per_kg=payload.price_per_kg,
        currency=payload.currency,
        quantity_kg=payload.quantity_kg,
        total_amount=total_amount,
        created_by_user_id=current_user.id
    )
    session.add(record)
    session.commit()
    session.refresh(record)
    return record

