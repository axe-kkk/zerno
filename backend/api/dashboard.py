from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select, func
from sqlalchemy import and_, or_
from backend.database import get_session
from backend.models import (
    CashRegister, GrainStock, GrainCulture, GrainIntake, GrainShipment,
    FarmerContract, FarmerContractStatus, FarmerContractType, GrainOwner, Transaction,
    PurchaseStock, PurchaseCategory, FarmerContractItem, FarmerContractPayment,
    FarmerContractItemDirection, FarmerContractItemType, FarmerContractPaymentType,
    GrainVoucher, FarmerGrainDeduction, FarmerGrainMovement,
    LeasePayment, LeasePaymentGrainItem, Person
)
from backend.auth import get_current_user, User
from datetime import datetime, timedelta, date, time as dtime
from io import BytesIO
from typing import Optional
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

router = APIRouter()


class DashboardStatsResponse(BaseModel):
    """Статистика для дашборда"""
    # Каса
    cash_balances: dict[str, float]  # {"uah": 0.0, "usd": 0.0, "eur": 0.0}
    
    # Склад
    total_stock_kg: float  # Общее количество зерна на складе
    stock_by_culture: list[dict]  # [{"name": "Пшениця", "quantity_kg": 1000.0}, ...]
    top_cultures: list[dict]  # Топ-5 культур по количеству
    
    # Контракти фермерів
    contracts_total: int
    contracts_open: int
    contracts_closed: int
    contracts_total_value: float  # Общая сумма всех контрактов
    contracts_balance: float  # Общий остаток по контрактам
    
    # Фермери
    farmers_total: int
    farmers_active: int  # С активными контрактами
    
    # Операції
    intakes_today: int
    intakes_pending: int  # Ожидают подтверждения
    shipments_today: int
    
    # Последние транзакции кассы
    recent_transactions: list[dict]
    
    # Закупки
    purchases_stock_total: float  # Общее количество закупок на складе
    purchases_by_category: list[dict]  # По категориям
    
    # Зерно у фермерів
    grain_purchased_from_farmers_kg: float  # Выкуплено у фермеров
    grain_not_purchased_from_farmers_kg: float  # Не выкуплено у фермеров

    # Тоннаж за сегодня
    intakes_today_kg: float
    shipments_today_kg: float

    # Доля собственного/фермерского зерна
    own_stock_kg: float
    farmer_stock_kg: float

    # Талони (хлібний завод)
    vouchers_total_value_uah: float
    vouchers_remaining_uah: float
    vouchers_count: int
    vouchers_open_count: int


def get_or_create_cash_register(session: Session) -> CashRegister:
    """Получение или создание кассы"""
    cash_register = session.exec(select(CashRegister)).first()
    if not cash_register:
        cash_register = CashRegister(
            uah_balance=0.0,
            usd_balance=0.0,
            eur_balance=0.0
        )
        session.add(cash_register)
        session.commit()
        session.refresh(cash_register)
    return cash_register


@router.get("/stats")
async def get_dashboard_stats(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Получение статистики для дашборда"""
    
    # ── Каса ──
    cash_register = get_or_create_cash_register(session)
    cash_balances = {
        "uah": round(cash_register.uah_balance, 2),
        "usd": round(cash_register.usd_balance, 2),
        "eur": round(cash_register.eur_balance, 2)
    }
    
    # ── Склад ──
    stocks = session.exec(select(GrainStock)).all()
    total_stock_kg = sum(s.quantity_kg for s in stocks)
    
    # Создаем словарь для быстрого поиска по culture_id
    stock_dict = {s.culture_id: s for s in stocks}
    
    # Получаем все культуры
    all_cultures = session.exec(select(GrainCulture)).all()
    
    stock_by_culture = []
    for culture in all_cultures:
        stock = stock_dict.get(culture.id)
        if stock:
            stock_by_culture.append({
                "name": culture.name,
                "quantity_kg": round(stock.quantity_kg, 2),
                "own_quantity_kg": round(stock.own_quantity_kg, 2),
                "farmer_quantity_kg": round(stock.farmer_quantity_kg, 2)
            })
        else:
            # Если нет записи в GrainStock, добавляем с нулевыми значениями
            stock_by_culture.append({
                "name": culture.name,
                "quantity_kg": 0.0,
                "own_quantity_kg": 0.0,
                "farmer_quantity_kg": 0.0
            })
    
    # Сортируем по количеству и берем топ-5
    top_cultures = sorted(stock_by_culture, key=lambda x: x["quantity_kg"], reverse=True)[:5]
    
    # ── Контракти фермерів ──
    all_contracts = session.exec(select(FarmerContract)).all()
    contracts_total = len(all_contracts)
    contracts_open = sum(1 for c in all_contracts if c.status == FarmerContractStatus.OPEN)
    contracts_closed = sum(1 for c in all_contracts if c.status == FarmerContractStatus.CLOSED)
    contracts_total_value = sum(c.total_value_uah for c in all_contracts)
    contracts_balance = sum(c.balance_uah for c in all_contracts)
    
    # ── Фермери ──
    all_farmers = session.exec(select(GrainOwner)).all()
    farmers_total = len(all_farmers)
    
    # Фермеры с активными (открытыми) контрактами
    active_farmer_ids = {c.owner_id for c in all_contracts if c.status == FarmerContractStatus.OPEN}
    farmers_active = len(active_farmer_ids)
    
    # ── Операції (сегодня) ──
    today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = today_start + timedelta(days=1)
    
    intakes_today = session.exec(
        select(func.count(GrainIntake.id))
        .where(and_(
            GrainIntake.created_at >= today_start,
            GrainIntake.created_at < today_end,
            GrainIntake.is_farmer_transfer == False,
        ))
    ).one() or 0
    
    intakes_pending = session.exec(
        select(func.count(GrainIntake.id))
        .where(
            or_(
                GrainIntake.pending_quality == True,
                GrainIntake.pending_tare == True,
            )
        )
    ).one() or 0
    
    shipments_today = session.exec(
        select(func.count(GrainShipment.id))
        .where(and_(
            GrainShipment.created_at >= today_start,
            GrainShipment.created_at < today_end
        ))
    ).one() or 0
    
    # ── Последние транзакции кассы ──
    recent_trans = session.exec(
        select(Transaction)
        .order_by(Transaction.created_at.desc())
        .limit(5)
    ).all()
    
    recent_transactions = []
    for t in recent_trans:
        recent_transactions.append({
            "id": t.id,
            "currency": t.currency.value,
            "amount": round(t.amount, 2),
            "type": t.transaction_type.value,
            "description": t.description or "",
            "created_at": t.created_at.isoformat() if t.created_at else None
        })
    
    # ── Закупки ──
    purchases_stock = session.exec(select(PurchaseStock)).all()
    purchases_stock_total = sum(p.quantity_kg for p in purchases_stock)
    
    # По категориям
    purchases_by_category = {}
    for p in purchases_stock:
        cat_name = p.category.value if p.category else "Інше"
        if cat_name not in purchases_by_category:
            purchases_by_category[cat_name] = 0.0
        purchases_by_category[cat_name] += p.quantity_kg
    
    purchases_by_category_list = [
        {"category": k, "quantity_kg": round(v, 2)}
        for k, v in sorted(purchases_by_category.items(), key=lambda x: x[1], reverse=True)
    ]
    
    # ── Зерно у фермерів ──
    # Выкупленное зерно: сумма quantity_kg из платежей типа goods_receive (не отмененных)
    purchased_payments = session.exec(
        select(FarmerContractPayment)
        .where(and_(
            FarmerContractPayment.payment_type == FarmerContractPaymentType.GOODS_RECEIVE.value,
            FarmerContractPayment.is_cancelled == False,
            FarmerContractPayment.quantity_kg.isnot(None)
        ))
    ).all()
    grain_purchased_from_farmers_kg = sum(p.quantity_kg for p in purchased_payments if p.quantity_kg)
    
    # Невыкупленное зерно: сумма (quantity_kg - delivered_kg) из позиций контрактов типа FROM_FARMER
    farmer_items = session.exec(
        select(FarmerContractItem)
        .where(FarmerContractItem.direction == FarmerContractItemDirection.FROM_FARMER.value)
    ).all()
    grain_not_purchased_from_farmers_kg = sum(
        max(0.0, item.quantity_kg - item.delivered_kg)
        for item in farmer_items
    )
    
    # ── Тоннаж за сегодня ──
    today_intakes = session.exec(
        select(GrainIntake)
        .where(and_(
            GrainIntake.created_at >= today_start,
            GrainIntake.created_at < today_end,
            GrainIntake.is_farmer_transfer == False,
        ))
    ).all()
    intakes_today_kg = sum(
        (i.accepted_weight_kg or 0.0)
        for i in today_intakes
        if not i.pending_quality and not i.pending_tare
    )
    
    today_shipments = session.exec(
        select(GrainShipment)
        .where(and_(
            GrainShipment.created_at >= today_start,
            GrainShipment.created_at < today_end
        ))
    ).all()
    shipments_today_kg = sum(s.quantity_kg for s in today_shipments if s.quantity_kg)
    
    # ── Доля собственного/фермерского ──
    own_stock_kg = sum(s.own_quantity_kg for s in stocks)
    farmer_stock_kg = sum(s.farmer_quantity_kg for s in stocks)

    # ── Талони (хлібний завод) ──
    vouchers = session.exec(select(GrainVoucher)).all()
    vouchers_total_value_uah = sum(v.total_value_uah for v in vouchers)
    vouchers_remaining_uah = sum(v.remaining_value_uah for v in vouchers)
    vouchers_count = len(vouchers)
    vouchers_open_count = sum(1 for v in vouchers if not v.is_closed)
    
    return DashboardStatsResponse(
        cash_balances=cash_balances,
        total_stock_kg=round(total_stock_kg, 2),
        stock_by_culture=stock_by_culture,
        top_cultures=top_cultures,
        contracts_total=contracts_total,
        contracts_open=contracts_open,
        contracts_closed=contracts_closed,
        contracts_total_value=round(contracts_total_value, 2),
        contracts_balance=round(contracts_balance, 2),
        farmers_total=farmers_total,
        farmers_active=farmers_active,
        intakes_today=intakes_today,
        intakes_pending=intakes_pending,
        shipments_today=shipments_today,
        recent_transactions=recent_transactions,
        purchases_stock_total=round(purchases_stock_total, 2),
        purchases_by_category=purchases_by_category_list,
        grain_purchased_from_farmers_kg=round(grain_purchased_from_farmers_kg, 2),
        grain_not_purchased_from_farmers_kg=round(grain_not_purchased_from_farmers_kg, 2),
        intakes_today_kg=round(intakes_today_kg, 2),
        shipments_today_kg=round(shipments_today_kg, 2),
        own_stock_kg=round(own_stock_kg, 2),
        farmer_stock_kg=round(farmer_stock_kg, 2),
        vouchers_total_value_uah=round(vouchers_total_value_uah, 2),
        vouchers_remaining_uah=round(vouchers_remaining_uah, 2),
        vouchers_count=vouchers_count,
        vouchers_open_count=vouchers_open_count
    )


# ============================================================
#  Period Report — звіт-залишки на дату/період
#  • Зведена картина по культурах: надійшло, видано, залишок
#  • Розрахунки з фермерами по культурах: що отримали, що віддали (по типах)
#  • Список боргів — відкриті контракти з невиплаченим балансом
# ============================================================

def _parse_date(value: Optional[str], end: bool = False) -> Optional[datetime]:
    if not value:
        return None
    try:
        d = date.fromisoformat(value)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Некоректна дата: {value}")
    return datetime.combine(d, dtime.max if end else dtime.min)


@router.get("/period-report")
async def period_report(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    start_date: Optional[str] = Query(default=None, description="ISO date"),
    end_date: Optional[str] = Query(default=None, description="ISO date"),
):
    """Сводный отчет за период:

    1) `movements` — рух зерна по кожній культурі: всі фізичні переміщення (прихід,
       втрати, відвантаження компаніям, видача через контракти, виплати оренди,
       перекази людям). Залишок — snapshot на момент запиту.
    2) `farmer_settlements` — розрахунки з фермерами: що надійшло, скільки викуплено,
       перекази між фермерами/людьми, списання, поточний залишок на балансі фермерів.
    3) `debts` — відкриті контракти з ненульовим балансом + сума, що вже сплачена.
    """
    start_dt = _parse_date(start_date, end=False)
    end_dt = _parse_date(end_date, end=True)

    cultures = session.exec(
        select(GrainCulture).where(GrainCulture.is_active == True).order_by(GrainCulture.name)
    ).all()

    stocks = session.exec(select(GrainStock)).all()
    stock_by_culture = {s.culture_id: s for s in stocks}

    cash = get_or_create_cash_register(session)

    def in_period(ts):
        if start_dt and (ts is None or ts < start_dt):
            return False
        if end_dt and (ts is None or ts > end_dt):
            return False
        return True

    # ─── Прихід (intake) ───
    intakes = session.exec(
        select(GrainIntake).where(
            GrainIntake.pending_quality == False,
            GrainIntake.pending_tare == False,
        )
    ).all()

    received_from_farmers: dict[int, float] = {}
    received_from_own: dict[int, float] = {}
    losses_by_culture: dict[int, float] = {}
    for i in intakes:
        if not in_period(i.created_at):
            continue
        if i.is_farmer_transfer:
            # Синтетичний intake від трансферу між фермерами — не реальний прихід.
            continue
        kg = i.accepted_weight_kg or 0.0
        if i.is_own_grain:
            received_from_own[i.culture_id] = received_from_own.get(i.culture_id, 0.0) + kg
        else:
            received_from_farmers[i.culture_id] = received_from_farmers.get(i.culture_id, 0.0) + kg
        loss = (i.net_weight_kg or 0.0) - kg
        if loss > 0:
            losses_by_culture[i.culture_id] = losses_by_culture.get(i.culture_id, 0.0) + loss

    # ─── Відвантаження компаніям ───
    shipments = session.exec(select(GrainShipment)).all()
    shipped_total: dict[int, float] = {}
    shipped_cash: dict[int, float] = {}
    shipped_cashless: dict[int, float] = {}
    for s in shipments:
        if not in_period(s.created_at):
            continue
        cid = s.culture_id
        kg = s.quantity_kg or 0.0
        shipped_total[cid] = shipped_total.get(cid, 0.0) + kg
        fmt = (s.payment_format or "none").lower()
        if fmt == "cash":
            shipped_cash[cid] = shipped_cash.get(cid, 0.0) + kg
        elif fmt == "cashless":
            shipped_cashless[cid] = shipped_cashless.get(cid, 0.0) + kg

    # ─── Платежі по контрактах ───
    fc_payments = session.exec(
        select(FarmerContractPayment).where(FarmerContractPayment.is_cancelled == False)
    ).all()

    issued_via_contracts: dict[int, float] = {}            # всі GOODS_ISSUE GRAIN — зерно фізично пішло клієнту
    bought_back_via_payments: dict[int, float] = {}        # GRAIN + GOODS_RECEIVE — викуп зерна з балансу фермера
    land_service_uah_total: float = 0.0
    for p in fc_payments:
        if not in_period(p.payment_date or p.created_at):
            continue
        if p.payment_type == FarmerContractPaymentType.GOODS_ISSUE.value:
            if p.culture_id and p.quantity_kg:
                issued_via_contracts[p.culture_id] = issued_via_contracts.get(p.culture_id, 0.0) + p.quantity_kg
            if (p.item_name or "").startswith("Обробка землі"):
                land_service_uah_total += p.amount_uah or 0.0
        elif p.payment_type in (
            FarmerContractPaymentType.GRAIN.value,
            FarmerContractPaymentType.GOODS_RECEIVE.value,
        ):
            if p.culture_id and p.quantity_kg:
                bought_back_via_payments[p.culture_id] = bought_back_via_payments.get(p.culture_id, 0.0) + p.quantity_kg

    # ─── Викуп через PAYMENT-контракти (миттєвий, без окремого payment) ───
    bought_back_via_payment_contracts: dict[int, float] = {}
    payment_contracts = session.exec(
        select(FarmerContract).where(
            FarmerContract.contract_type == FarmerContractType.PAYMENT.value
        )
    ).all()
    payment_contract_ids = [c.id for c in payment_contracts if in_period(c.created_at)]
    if payment_contract_ids:
        farmer_grain_items = session.exec(
            select(FarmerContractItem).where(
                FarmerContractItem.contract_id.in_(payment_contract_ids),
                FarmerContractItem.direction == FarmerContractItemDirection.FROM_FARMER.value,
                FarmerContractItem.item_type == FarmerContractItemType.GRAIN.value,
            )
        ).all()
        for it in farmer_grain_items:
            if it.culture_id and it.quantity_kg:
                bought_back_via_payment_contracts[it.culture_id] = bought_back_via_payment_contracts.get(it.culture_id, 0.0) + it.quantity_kg

    # ─── Виплати оренди зерном ───
    lease_grain_items = session.exec(select(LeasePaymentGrainItem)).all()
    lease_payments_kg: dict[int, float] = {}
    for li in lease_grain_items:
        parent = session.get(LeasePayment, li.payment_id)
        if not parent or parent.is_cancelled:
            continue
        if not in_period(parent.payment_date):
            continue
        lease_payments_kg[li.culture_id] = lease_payments_kg.get(li.culture_id, 0.0) + (li.quantity_kg or 0.0)

    # ─── Переміщення фермерського зерна ───
    movements = session.exec(select(FarmerGrainMovement)).all()
    transfer_between_farmers: dict[int, float] = {}    # to_owner_id IS NOT NULL — лишається на складі
    transfer_to_people: dict[int, float] = {}          # to_person_id IS NOT NULL — фізично йде
    deducted: dict[int, float] = {}                    # списання
    for m in movements:
        if not in_period(m.created_at):
            continue
        kg = m.quantity_kg or 0.0
        if m.to_person_id:
            transfer_to_people[m.culture_id] = transfer_to_people.get(m.culture_id, 0.0) + kg
        elif m.to_owner_id:
            transfer_between_farmers[m.culture_id] = transfer_between_farmers.get(m.culture_id, 0.0) + kg
        else:
            deducted[m.culture_id] = deducted.get(m.culture_id, 0.0) + kg

    # ─── Борги (відкриті контракти) ───
    debts_rows: list[dict] = []
    open_contracts = session.exec(
        select(FarmerContract).where(FarmerContract.balance_uah > 0.01)
    ).all()
    for c in open_contracts:
        if c.owner_id:
            owner = session.get(GrainOwner, c.owner_id)
            name = owner.full_name if owner else f"#{c.owner_id}"
            is_person = False
        elif c.person_id:
            person = session.get(Person, c.person_id)
            name = person.full_name if person else f"#{c.person_id}"
            is_person = True
        else:
            continue
        total = c.total_value_uah or 0.0
        balance = c.balance_uah or 0.0
        debts_rows.append({
            "contract_id": c.id,
            "name": name,
            "is_person": is_person,
            "type": c.contract_type,
            "created_at": c.created_at.isoformat() if c.created_at else None,
            "total_uah": round(total, 2),
            "paid_uah": round(max(0.0, total - balance), 2),
            "balance_uah": round(balance, 2),
            "note": c.note or "",
        })
    debts_rows.sort(key=lambda d: d["balance_uah"], reverse=True)

    # ─── Build per-culture rows ───
    movements_table = []
    farmer_settlements = []
    for c in cultures:
        cid = c.id
        rec_farmers = round(received_from_farmers.get(cid, 0.0), 2)
        rec_own = round(received_from_own.get(cid, 0.0), 2)
        rec_total = round(rec_farmers + rec_own, 2)

        losses = round(losses_by_culture.get(cid, 0.0), 2)
        gross = rec_total + losses
        loss_pct = round((losses / gross) * 100, 2) if gross > 0 else 0.0

        ship_total = round(shipped_total.get(cid, 0.0), 2)
        ship_cash = round(shipped_cash.get(cid, 0.0), 2)
        ship_cashless = round(shipped_cashless.get(cid, 0.0), 2)

        issued = round(issued_via_contracts.get(cid, 0.0), 2)
        lease = round(lease_payments_kg.get(cid, 0.0), 2)
        to_people = round(transfer_to_people.get(cid, 0.0), 2)

        st = stock_by_culture.get(cid)
        balance = round(st.quantity_kg if st else 0.0, 2)
        own_balance = round(st.own_quantity_kg if st else 0.0, 2)
        farmer_balance = round(st.farmer_quantity_kg if st else 0.0, 2)

        movements_table.append({
            "culture_id": cid,
            "culture_name": c.name,
            "received_from_farmers_kg": rec_farmers,
            "received_from_own_kg": rec_own,
            "received_total_kg": rec_total,
            "losses_kg": losses,
            "loss_percent": loss_pct,
            "shipped_cash_kg": ship_cash,
            "shipped_cashless_kg": ship_cashless,
            "shipped_total_kg": ship_total,
            "issued_via_contracts_kg": issued,
            "lease_payments_kg": lease,
            "transfer_to_people_kg": to_people,
            "balance_kg": balance,
            "own_balance_kg": own_balance,
            "farmer_balance_kg": farmer_balance,
        })

        bought = round(
            bought_back_via_payments.get(cid, 0.0)
            + bought_back_via_payment_contracts.get(cid, 0.0),
            2,
        )
        between = round(transfer_between_farmers.get(cid, 0.0), 2)
        deduct = round(deducted.get(cid, 0.0), 2)

        farmer_settlements.append({
            "culture_id": cid,
            "culture_name": c.name,
            "received_from_farmers_kg": rec_farmers,
            "bought_back_kg": bought,
            "transfer_between_farmers_kg": between,
            "transfer_to_people_kg": to_people,
            "deduct_kg": deduct,
            "farmer_balance_kg": farmer_balance,
        })

    return {
        "period": {
            "start_date": start_date or None,
            "end_date": end_date or None,
        },
        "cash_balances": {
            "uah": round(cash.uah_balance, 2),
            "usd": round(cash.usd_balance, 2),
            "eur": round(cash.eur_balance, 2),
        },
        "movements": movements_table,
        "farmer_settlements": farmer_settlements,
        "debts": debts_rows,
        "land_service_total_uah": round(land_service_uah_total, 2),
    }


# ============================================================
#  Excel export — those 3 tables + cash snapshot, one workbook
# ============================================================

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)
_ALT_FILL = PatternFill("solid", fgColor="F8FAFC")
_BALANCE_FILL = PatternFill("solid", fgColor="FEF3C7")


def _style_header_row(sheet, row_index: int, ncols: int):
    for col in range(1, ncols + 1):
        cell = sheet.cell(row=row_index, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


def _style_body(sheet, ncols: int, header_rows: int = 1, number_columns=(), balance_column=None):
    for row in range(header_rows + 1, sheet.max_row + 1):
        row_fill = _ALT_FILL if (row - header_rows) % 2 == 0 else None
        for col in range(1, ncols + 1):
            cell = sheet.cell(row=row, column=col)
            cell.border = _THIN_BORDER
            if col == balance_column:
                cell.fill = _BALANCE_FILL
                cell.font = Font(bold=True)
            elif row_fill:
                cell.fill = row_fill
            if col in number_columns:
                cell.number_format = "#,##0.00"


_CONTRACT_TYPE_LABELS = {
    "debt": "Борговий",
    "payment": "Виплата",
    "reserve": "Резерв",
    "exchange": "Обмін",
}


@router.get("/period-report/export")
async def export_period_report(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    start_date: Optional[str] = Query(default=None, description="ISO date"),
    end_date: Optional[str] = Query(default=None, description="ISO date"),
):
    """Excel-звіт по дашборду: каса + 3 таблиці (рух, розрахунки, борги)."""
    data = await period_report(
        session=session,
        current_user=current_user,
        start_date=start_date,
        end_date=end_date,
    )

    wb = Workbook()

    # ─── Sheet 0: meta + cash ───
    s0 = wb.active
    s0.title = "Каса і період"
    period = data["period"]
    cash = data["cash_balances"]
    s0.append(["Звіт сформовано", datetime.utcnow().strftime("%Y-%m-%d %H:%M")])
    s0.append(["Період з", period.get("start_date") or "—"])
    s0.append(["Період по", period.get("end_date") or "—"])
    s0.append([])
    s0.append(["Каса", "Сума"])
    _style_header_row(s0, row_index=5, ncols=2)
    s0.append(["UAH", round(cash.get("uah", 0.0), 2)])
    s0.append(["USD", round(cash.get("usd", 0.0), 2)])
    s0.append(["EUR", round(cash.get("eur", 0.0), 2)])
    if data.get("land_service_total_uah"):
        s0.append([])
        s0.append(["Обробка землі за період, ₴", round(data["land_service_total_uah"], 2)])
    for col_idx, width in enumerate([28, 22], start=1):
        s0.column_dimensions[chr(64 + col_idx)].width = width

    # ─── Sheet 1: Рух зерна по складу ───
    s1 = wb.create_sheet("Рух зерна по складу")
    s1.append([
        "Культура",
        "Прихід: Від фермерів, кг",
        "Прихід: На підприємство, кг",
        "Прихід: Всього, кг",
        "Втрати, кг",
        "% втрат",
        "Відвантажено компаніям: Готівка, кг",
        "Відвантажено компаніям: Безготівка, кг",
        "Відвантажено компаніям: Всього, кг",
        "Видано контрактами, кг",
        "Виплати оренди, кг",
        "Перекази людям, кг",
        "Залишок, кг",
    ])
    _style_header_row(s1, row_index=1, ncols=13)
    for r in data["movements"]:
        s1.append([
            r["culture_name"],
            r["received_from_farmers_kg"],
            r["received_from_own_kg"],
            r["received_total_kg"],
            r["losses_kg"],
            r["loss_percent"],
            r["shipped_cash_kg"],
            r["shipped_cashless_kg"],
            r["shipped_total_kg"],
            r["issued_via_contracts_kg"],
            r["lease_payments_kg"],
            r["transfer_to_people_kg"],
            r["balance_kg"],
        ])
    _style_body(
        s1,
        ncols=13,
        number_columns=tuple(range(2, 14)),
        balance_column=13,
    )
    s1.freeze_panes = "B2"
    if s1.max_row > 1:
        s1.auto_filter.ref = f"A1:M{s1.max_row}"
    for col_idx, width in enumerate(
        [16, 18, 20, 16, 14, 12, 22, 22, 18, 20, 18, 18, 16],
        start=1,
    ):
        s1.column_dimensions[chr(64 + col_idx)].width = width

    # ─── Sheet 2: Розрахунки з фермерами ───
    s2 = wb.create_sheet("Розрахунки з фермерами")
    s2.append([
        "Культура",
        "Надійшло, кг",
        "Викуплено, кг",
        "Перекази між фермерами, кг",
        "Перекази людям, кг",
        "Списання, кг",
        "Не викуплено (на балансі), кг",
    ])
    _style_header_row(s2, row_index=1, ncols=7)
    for r in data["farmer_settlements"]:
        s2.append([
            r["culture_name"],
            r["received_from_farmers_kg"],
            r["bought_back_kg"],
            r["transfer_between_farmers_kg"],
            r["transfer_to_people_kg"],
            r["deduct_kg"],
            r["farmer_balance_kg"],
        ])
    _style_body(
        s2,
        ncols=7,
        number_columns=(2, 3, 4, 5, 6, 7),
        balance_column=7,
    )
    s2.freeze_panes = "B2"
    if s2.max_row > 1:
        s2.auto_filter.ref = f"A1:G{s2.max_row}"
    for col_idx, width in enumerate(
        [16, 16, 16, 22, 20, 16, 24],
        start=1,
    ):
        s2.column_dimensions[chr(64 + col_idx)].width = width

    # ─── Sheet 3: Борги ───
    s3 = wb.create_sheet("Борги")
    s3.append([
        "#",
        "Контрагент",
        "Тип",
        "Дата",
        "Сума, ₴",
        "Сплачено, ₴",
        "Залишок, ₴",
        "Примітка",
    ])
    _style_header_row(s3, row_index=1, ncols=8)
    for d in data["debts"]:
        name = d["name"] + (" (людина)" if d.get("is_person") else "")
        date_str = "—"
        if d.get("created_at"):
            try:
                date_str = datetime.fromisoformat(d["created_at"]).strftime("%Y-%m-%d")
            except (TypeError, ValueError):
                date_str = "—"
        s3.append([
            d["contract_id"],
            name,
            _CONTRACT_TYPE_LABELS.get(d["type"], d["type"]),
            date_str,
            d["total_uah"],
            d["paid_uah"],
            d["balance_uah"],
            d["note"] or "",
        ])
    _style_body(
        s3,
        ncols=8,
        number_columns=(5, 6, 7),
        balance_column=7,
    )
    s3.freeze_panes = "A2"
    if s3.max_row > 1:
        s3.auto_filter.ref = f"A1:H{s3.max_row}"
    for col_idx, width in enumerate(
        [8, 28, 14, 14, 14, 14, 14, 30],
        start=1,
    ):
        s3.column_dimensions[chr(64 + col_idx)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Filename: dashboard_2026-05-01_2026-05-09.xlsx або dashboard_all_time.xlsx
    if start_date and end_date:
        period_label = f"{start_date}_{end_date}"
    elif start_date:
        period_label = f"from_{start_date}"
    elif end_date:
        period_label = f"to_{end_date}"
    else:
        period_label = "all_time"
    filename = f"dashboard_{period_label}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
