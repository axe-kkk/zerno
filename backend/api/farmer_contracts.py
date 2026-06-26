from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select
from sqlalchemy import func
from typing import Optional
from datetime import datetime, date, time as dtime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from backend.database import get_session
from backend.models import (
    GrainOwner,
    GrainCulture,
    GrainIntake,
    GrainStock,
    PurchaseStock,
    CashRegister,
    Transaction,
    TransactionType,
    Currency,
    FarmerContract,
    FarmerContractItem,
    FarmerContractPayment,
    FarmerContractItemType,
    FarmerContractItemDirection,
    FarmerContractPaymentType,
    FarmerContractType,
    FarmerContractStatus,
    FarmerGrainDeduction,
    GrainVoucher,
    Person,
    User
)
from backend.schemas import (
    FarmerContractCreate,
    FarmerContractResponse,
    FarmerContractDetailResponse,
    FarmerContractItemResponse,
    FarmerContractPaymentCreate,
    FarmerContractPaymentResponse,
    ReserveActivateRequest,
)
from backend.auth import get_current_user, get_current_super_admin

router = APIRouter()

# Типи операцій, що зменшують залишок боргу (реальна оплата боргу фермером).
# Видача товару/талону фермеру НЕ є оплатою і не повинна зменшувати борг.
_BALANCE_REDUCING_PAYMENT_TYPES = (
    FarmerContractPaymentType.CASH.value,
    FarmerContractPaymentType.GRAIN.value,
    # SETTLEMENT — повне погашення боргу одним кроком (PAYMENT-контракт).
    # Без цього balance закритого PAYMENT-контракту обчислюється як total
    # замість 0, бо settlement-payment не враховувався як «оплата боргу».
    FarmerContractPaymentType.SETTLEMENT.value,
)


def _compute_contract_balance_uah(session: Session, contract_id: int, total_value_uah: float) -> float:
    """Залишок боргу = сума контракту мінус сума оплат боргу (гроші, зерно).

    Видача товарів/талонів фермеру є виконанням зобов'язання компанії, а не оплатою боргу,
    тому не повинна зменшувати balance_uah.
    """
    payments = session.exec(
        select(FarmerContractPayment).where(
            FarmerContractPayment.contract_id == contract_id,
            FarmerContractPayment.payment_type.in_(_BALANCE_REDUCING_PAYMENT_TYPES),
            FarmerContractPayment.is_cancelled == False,
        )
    ).all()
    paid = sum(p.amount_uah or 0 for p in payments)
    return max(0.0, (total_value_uah or 0) - paid)


def _get_or_create_grain_stock(session: Session, culture_id: int) -> GrainStock:
    stock = session.exec(select(GrainStock).where(GrainStock.culture_id == culture_id)).first()
    if not stock:
        stock = GrainStock(
            culture_id=culture_id,
            quantity_kg=0.0,
            own_quantity_kg=0.0,
            farmer_quantity_kg=0.0,
            reserved_kg=0.0
        )
        session.add(stock)
        session.flush()
    return stock


def _get_cash_register(session: Session) -> CashRegister:
    register = session.exec(select(CashRegister)).first()
    if not register:
        register = CashRegister(uah_balance=0.0, usd_balance=0.0, eur_balance=0.0)
        session.add(register)
        session.flush()
    return register


def _get_farmer_balance(session: Session, owner_id: int, culture_id: int) -> float:
    total = session.exec(
        select(func.sum(GrainIntake.accepted_weight_kg)).where(
            GrainIntake.owner_id == owner_id,
            GrainIntake.culture_id == culture_id,
            GrainIntake.is_own_grain == False,
            GrainIntake.pending_quality == False,
            GrainIntake.pending_tare == False,
        )
    ).first() or 0.0
    deductions = session.exec(
        select(func.sum(FarmerGrainDeduction.quantity_kg)).where(
            FarmerGrainDeduction.owner_id == owner_id,
            FarmerGrainDeduction.culture_id == culture_id
        )
    ).first() or 0.0
    return float(total) - float(deductions)


def _get_person_balance(session: Session, person_id: int, culture_id: int) -> float:
    """Баланс зерна людини по культурі.
    Прибутки: переказів to_person_id за цю культуру.
    Витрати:
      1) активні GRAIN-оплати по DEBT-контрактах цієї людини за цю культуру;
      2) FROM_FARMER GRAIN-позиції у PAYMENT-контрактах цієї людини (викуп —
         зерно одразу йде у нас, незалежно від типу оплати).
    Скасовані контракти/оплати не враховуються.
    """
    from backend.models import FarmerGrainMovement, FarmerContract, FarmerContractItem
    incoming = session.exec(
        select(func.sum(FarmerGrainMovement.quantity_kg)).where(
            FarmerGrainMovement.to_person_id == person_id,
            FarmerGrainMovement.culture_id == culture_id,
            FarmerGrainMovement.movement_type == "transfer",
        )
    ).first() or 0.0

    # OUT-перекази від цієї людини (наприклад, людина → підприємство або людина → фермер)
    outgoing_transfer = session.exec(
        select(func.sum(FarmerGrainMovement.quantity_kg)).where(
            FarmerGrainMovement.from_person_id == person_id,
            FarmerGrainMovement.culture_id == culture_id,
            FarmerGrainMovement.movement_type == "transfer",
        )
    ).first() or 0.0

    contract_ids_query = select(FarmerContract.id).where(
        FarmerContract.person_id == person_id,
        FarmerContract.status != FarmerContractStatus.CANCELLED.value,
    )
    grain_spent = session.exec(
        select(func.sum(FarmerContractPayment.quantity_kg)).where(
            FarmerContractPayment.contract_id.in_(contract_ids_query),
            FarmerContractPayment.payment_type == FarmerContractPaymentType.GRAIN.value,
            FarmerContractPayment.culture_id == culture_id,
            FarmerContractPayment.is_cancelled == False,
        )
    ).first() or 0.0

    # Викуп (PAYMENT-контракт): зерно йде у нас одразу при створенні контракту.
    payment_spent = session.exec(
        select(func.sum(FarmerContractItem.delivered_kg)).where(
            FarmerContractItem.contract_id.in_(contract_ids_query),
            FarmerContractItem.direction == FarmerContractItemDirection.FROM_FARMER.value,
            FarmerContractItem.item_type == FarmerContractItemType.GRAIN.value,
            FarmerContractItem.culture_id == culture_id,
        )
    ).first() or 0.0

    return float(incoming) - float(outgoing_transfer) - float(grain_spent) - float(payment_spent)


def _normalize_name_for_match(name: str) -> str:
    """Канонічна назва для зіставлення: однакове написання кирилицею/латиницею (С/С, А/а тощо)."""
    import re
    if not name:
        return ""
    s = re.sub(r"\s+", " ", name.strip()).lower()
    # Кириличні літери-близнюки замінюємо на латинські, щоб "Сульфат Амонію" = "Cульфат амонію"
    cyr_to_lat = {
        "а": "a", "в": "v", "с": "c", "е": "e", "о": "o", "р": "p", "у": "y", "х": "x",
        "і": "i", "ї": "i", "є": "e", "и": "y", "н": "n", "т": "t", "м": "m", "л": "l",
        "ф": "f", "ю": "u", "я": "a", "б": "b", "г": "g", "д": "d", "ж": "zh", "з": "z",
        "к": "k", "п": "p", "ч": "ch", "ш": "sh", "щ": "shch", "ь": "", "’": "",
    }
    result = []
    for c in s:
        result.append(cyr_to_lat.get(c, c))
    return "".join(result)


def _find_purchase_stock_for_activation(
    session: Session, item: FarmerContractItem
) -> Optional[PurchaseStock]:
    """Для активації резерву: знайти складську позицію по item (за id або за назвою з урахуванням різного написання)."""
    # 1) Якщо є привʼязка і на тому складі вистачає — повертаємо її
    if item.purchase_stock_id:
        pstock = session.get(PurchaseStock, item.purchase_stock_id)
        if pstock:
            need = item.quantity_kg
            available = pstock.quantity_kg - (pstock.reserved_kg - (item.quantity_kg if pstock.id == item.purchase_stock_id else 0))
            if available >= need - 0.01:
                return pstock
    # 2) Шукаємо будь-яку позицію з такою ж назвою (нормалізованою) і з достатньою кількістю
    item_name = (item.item_name or "").strip()
    if not item_name:
        return session.get(PurchaseStock, item.purchase_stock_id) if item.purchase_stock_id else None
    key = _normalize_name_for_match(item_name)
    all_stocks = session.exec(select(PurchaseStock)).all()
    for s in all_stocks:
        if _normalize_name_for_match(s.name) != key:
            continue
        available = s.quantity_kg - s.reserved_kg
        if s.id == item.purchase_stock_id:
            available += item.quantity_kg
        if available >= item.quantity_kg - 0.01:
            return s
    return session.get(PurchaseStock, item.purchase_stock_id) if item.purchase_stock_id else None


def _find_or_create_purchase_stock_by_name(session: Session, name: str, category: str = "fertilizer") -> PurchaseStock:
    """Для резерву: шукаємо за назвою, якщо немає — створюємо з 0 кількістю."""
    import re
    normalized = re.sub(r'\s+', ' ', name.strip().lower())
    stock = session.exec(
        select(PurchaseStock).where(PurchaseStock.normalized_name == normalized)
    ).first()
    if stock:
        return stock
    # Створюємо нову позицію з 0 на складі
    stock = PurchaseStock(
        name=name.strip(),
        normalized_name=normalized,
        category=category,
        quantity_kg=0.0,
        reserved_kg=0.0,
        sale_price_per_kg=0.0
    )
    session.add(stock)
    session.flush()
    return stock


@router.get("", response_model=list[FarmerContractResponse])
async def list_farmer_contracts(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    owner_id: Optional[int] = Query(None),
    person_id: Optional[int] = Query(None),
    status_filter: Optional[FarmerContractStatus] = Query(None)
):
    query = select(FarmerContract).order_by(FarmerContract.created_at.desc())
    if owner_id:
        query = query.where(FarmerContract.owner_id == owner_id)
    if person_id:
        query = query.where(FarmerContract.person_id == person_id)
    if status_filter:
        query = query.where(FarmerContract.status == status_filter.value)
    contracts = session.exec(query).all()
    if not contracts:
        return []
    # Залишок боргу рахуємо з оплат боргу (гроші/зерно)
    contract_ids = [c.id for c in contracts]
    payments = session.exec(
        select(FarmerContractPayment).where(
            FarmerContractPayment.contract_id.in_(contract_ids),
            FarmerContractPayment.payment_type.in_(_BALANCE_REDUCING_PAYMENT_TYPES),
            FarmerContractPayment.is_cancelled == False,
        )
    ).all()
    paid_by_id = {}
    for p in payments:
        paid_by_id[p.contract_id] = paid_by_id.get(p.contract_id, 0) + (p.amount_uah or 0)
    return [
        FarmerContractResponse(
            **{**(c.model_dump() if hasattr(c, "model_dump") else c.dict()), "balance_uah": max(0.0, (c.total_value_uah or 0) - paid_by_id.get(c.id, 0))}
        )
        for c in contracts
    ]


@router.get("/payments", response_model=list[FarmerContractPaymentResponse])
async def list_all_farmer_contract_payments(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    owner_id: Optional[int] = Query(None),
    contract_id: Optional[int] = Query(None)
):
    query = select(FarmerContractPayment)
    if contract_id:
        query = query.where(FarmerContractPayment.contract_id == contract_id)
    if owner_id:
        query = query.join(FarmerContract, FarmerContract.id == FarmerContractPayment.contract_id)\
            .where(FarmerContract.owner_id == owner_id)
    query = query.order_by(FarmerContractPayment.payment_date.desc())
    return session.exec(query).all()


def _excel_header_style():
    return {
        "fill": PatternFill("solid", fgColor="1F2937"),
        "font": Font(name="Calibri", bold=True, color="FFFFFF", size=11),
        "alignment": Alignment(horizontal="center", vertical="center"),
        "border": Border(
            left=Side(style="thin", color="374151"),
            right=Side(style="thin", color="374151"),
            top=Side(style="thin", color="374151"),
            bottom=Side(style="thin", color="374151")
        )
    }


def _fc_apply_header(sheet, headers):
    sheet.append(headers)
    style = _excel_header_style()
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = style["fill"]
        cell.font = style["font"]
        cell.alignment = style["alignment"]
        cell.border = style["border"]


def _fc_apply_body_style(sheet, num_cols):
    alt_fill = PatternFill("solid", fgColor="F8FAFC")
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB")
    )
    for row in range(2, sheet.max_row + 1):
        row_fill = alt_fill if row % 2 == 0 else None
        for col in range(1, num_cols + 1):
            cell = sheet.cell(row=row, column=col)
            if row_fill:
                cell.fill = row_fill
            cell.border = thin_border


def _fc_to_response(workbook, filename):
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/export")
async def export_farmer_contracts(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    owner_id: Optional[int] = Query(None),
    contract_type: Optional[str] = Query(None),
    status_filter: Optional[str] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None)
):
    """Експорт контрактів фермерів у Excel"""
    query = select(FarmerContract).order_by(FarmerContract.created_at.desc())
    if owner_id:
        query = query.where(FarmerContract.owner_id == owner_id)
    if contract_type:
        query = query.where(FarmerContract.contract_type == contract_type)
    if status_filter:
        query = query.where(FarmerContract.status == status_filter)
    if start_date:
        try:
            sd = datetime.combine(date.fromisoformat(start_date), dtime.min)
            query = query.where(FarmerContract.created_at >= sd)
        except ValueError:
            pass
    if end_date:
        try:
            ed = datetime.combine(date.fromisoformat(end_date), dtime(23, 59, 59))
            query = query.where(FarmerContract.created_at <= ed)
        except ValueError:
            pass

    contracts = session.exec(query).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Контракти фермерів"
    headers = ["№", "Фермер", "Тип", "Сума, грн", "Залишок, грн", "Статус", "Валюта", "Дата", "Примітка"]
    _fc_apply_header(ws, headers)

    type_labels = {"payment": "Виплата", "debt": "Контракт", "reserve": "Резерв", "exchange": "Обмін"}
    status_labels = {"open": "Відкритий", "pending": "Очікує", "closed": "Закритий", "cancelled": "Скасований"}
    dir_labels = {"from_company": "Від компанії", "from_farmer": "Від фермера"}
    open_fill = PatternFill("solid", fgColor="DBEAFE")
    closed_fill = PatternFill("solid", fgColor="BBF7D0")
    pending_fill = PatternFill("solid", fgColor="FEF3C7")
    cancelled_fill = PatternFill("solid", fgColor="FECACA")

    # Collect all items for second sheet
    all_items_data = []

    for c in contracts:
        owner = session.get(GrainOwner, c.owner_id)
        items = session.exec(
            select(FarmerContractItem).where(FarmerContractItem.contract_id == c.id)
        ).all()

        curr_text = ""
        if c.currency:
            curr_text = c.currency.upper()
            if c.exchange_rate:
                curr_text += f" (курс: {c.exchange_rate})"
        else:
            curr_text = "UAH"

        ws.append([
            c.id,
            owner.full_name if owner else f"#{c.owner_id}",
            type_labels.get(c.contract_type, c.contract_type),
            round(c.total_value_uah, 2),
            round(c.balance_uah, 2),
            status_labels.get(c.status, c.status),
            curr_text,
            c.created_at.strftime("%d.%m.%Y") if c.created_at else "-",
            c.note or "-"
        ])

        owner_name = owner.full_name if owner else f"#{c.owner_id}"
        for item in items:
            all_items_data.append((c.id, owner_name, item))

    for row in range(2, ws.max_row + 1):
        status_cell = ws.cell(row=row, column=6)
        val = status_cell.value
        if val == "Відкритий":
            status_cell.fill = open_fill
        elif val == "Закритий":
            status_cell.fill = closed_fill
        elif val == "Очікує":
            status_cell.fill = pending_fill
        elif val == "Скасований":
            status_cell.fill = cancelled_fill
        status_cell.alignment = Alignment(horizontal="center")

    _fc_apply_body_style(ws, len(headers))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{chr(64 + len(headers))}{ws.max_row}"
    widths = [8, 30, 14, 16, 16, 14, 18, 14, 30]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = w

    # ── Sheet 2: Позиції контрактів ──
    ws2 = wb.create_sheet("Позиції контрактів")
    item_headers = ["Контракт №", "Фермер", "Напрямок", "Позиція", "Кількість, кг", "Ціна, грн/кг", "Сума, грн", "Видано, кг"]
    _fc_apply_header(ws2, item_headers)
    for cid, oname, item in all_items_data:
        ws2.append([
            cid,
            oname,
            dir_labels.get(item.direction, item.direction),
            item.item_name or "-",
            round(item.quantity_kg, 2),
            round(item.price_per_kg, 2),
            round(item.total_value_uah, 2),
            round(item.delivered_kg, 2),
        ])
    _fc_apply_body_style(ws2, len(item_headers))
    ws2.freeze_panes = "A2"
    if ws2.max_row > 1:
        ws2.auto_filter.ref = f"A1:{chr(64 + len(item_headers))}{ws2.max_row}"
    item_widths = [12, 30, 18, 25, 16, 16, 16, 14]
    for idx, w in enumerate(item_widths, 1):
        ws2.column_dimensions[chr(64 + idx)].width = w

    return _fc_to_response(wb, f"farmer_contracts_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx")


@router.get("/payments/export")
async def export_farmer_contract_payments(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    owner_id: Optional[int] = Query(None),
    contract_id: Optional[int] = Query(None),
    payment_type: Optional[str] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    show_cancelled: bool = Query(False)
):
    """Експорт виплат по контрактах фермерів у Excel"""
    query = select(FarmerContractPayment)
    if contract_id:
        query = query.where(FarmerContractPayment.contract_id == contract_id)
    if owner_id:
        query = query.join(FarmerContract, FarmerContract.id == FarmerContractPayment.contract_id)\
            .where(FarmerContract.owner_id == owner_id)
    if payment_type:
        query = query.where(FarmerContractPayment.payment_type == payment_type)
    if not show_cancelled:
        query = query.where(FarmerContractPayment.is_cancelled == False)
    if start_date:
        try:
            sd = datetime.combine(date.fromisoformat(start_date), dtime.min)
            query = query.where(FarmerContractPayment.payment_date >= sd)
        except ValueError:
            pass
    if end_date:
        try:
            ed = datetime.combine(date.fromisoformat(end_date), dtime(23, 59, 59))
            query = query.where(FarmerContractPayment.payment_date <= ed)
        except ValueError:
            pass

    payments = session.exec(query.order_by(FarmerContractPayment.payment_date.desc())).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Виплати по контрактах"
    headers = ["Дата", "Фермер", "Контракт", "Тип", "Позиція", "Кількість, кг", "Сума", "Валюта", "Курс", "Сума, грн", "Статус"]
    _fc_apply_header(ws, headers)

    type_labels = {
        "goods_issue": "Видача",
        "goods_receive": "Прийом",
        "cash": "Гроші",
        "grain": "Зерно",
        "settlement": "Розрахунок"
    }
    active_fill = PatternFill("solid", fgColor="BBF7D0")
    cancelled_fill = PatternFill("solid", fgColor="FECACA")

    for p in payments:
        contract = session.get(FarmerContract, p.contract_id)
        owner_name = "-"
        if contract:
            owner = session.get(GrainOwner, contract.owner_id)
            owner_name = owner.full_name if owner else f"#{contract.owner_id}"

        ws.append([
            p.payment_date.strftime("%d.%m.%Y") if p.payment_date else "-",
            owner_name,
            f"#{p.contract_id}",
            type_labels.get(p.payment_type, p.payment_type),
            p.item_name or "-",
            round(p.quantity_kg, 2) if p.quantity_kg else "-",
            round(p.amount, 2) if p.amount else "-",
            (p.currency or "UAH").upper(),
            p.exchange_rate if p.exchange_rate else "-",
            round(p.amount_uah, 2) if p.amount_uah else "-",
            "Скасовано" if p.is_cancelled else "Активна"
        ])

    for row in range(2, ws.max_row + 1):
        status_cell = ws.cell(row=row, column=11)
        if status_cell.value == "Скасовано":
            status_cell.fill = cancelled_fill
        else:
            status_cell.fill = active_fill
        status_cell.alignment = Alignment(horizontal="center")

    _fc_apply_body_style(ws, len(headers))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{chr(64 + len(headers))}{ws.max_row}"
    widths = [14, 30, 12, 14, 25, 16, 14, 10, 10, 16, 14]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = w

    return _fc_to_response(wb, f"farmer_payments_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx")


@router.get("/{contract_id}/export")
async def export_single_farmer_contract(
    contract_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Експорт деталей одного контракту фермера у Excel"""
    contract = session.get(FarmerContract, contract_id)
    if not contract:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Контракт не знайдено")

    owner = session.get(GrainOwner, contract.owner_id)
    owner_name = owner.full_name if owner else f"#{contract.owner_id}"
    items = session.exec(
        select(FarmerContractItem).where(FarmerContractItem.contract_id == contract_id)
    ).all()
    payments = session.exec(
        select(FarmerContractPayment)
        .where(FarmerContractPayment.contract_id == contract_id)
        .order_by(FarmerContractPayment.payment_date.desc())
    ).all()

    type_labels = {"payment": "Виплата", "debt": "Контракт", "reserve": "Резерв", "exchange": "Обмін"}
    status_labels = {"open": "Відкритий", "pending": "Очікує", "closed": "Закритий", "cancelled": "Скасований"}
    dir_labels = {"from_company": "Від компанії", "from_farmer": "Від фермера"}
    pay_type_labels = {
        "goods_issue": "Видача", "goods_receive": "Прийом",
        "cash": "Гроші", "grain": "Зерно", "settlement": "Розрахунок"
    }

    wb = Workbook()

    # ── Sheet 1: Інформація ──
    ws_info = wb.active
    ws_info.title = "Контракт"

    title_font = Font(bold=True, size=16, color="1F2937")
    label_font = Font(bold=True, size=11, color="64748B")
    value_font = Font(size=12, color="1F2937")
    section_font = Font(bold=True, size=13, color="1F2937")
    section_fill = PatternFill("solid", fgColor="EFF6FF")

    ws_info.merge_cells("A1:D1")
    cell_title = ws_info.cell(row=1, column=1)
    cell_title.value = f"Контракт #{contract_id}"
    cell_title.font = title_font
    cell_title.alignment = Alignment(vertical="center")
    ws_info.row_dimensions[1].height = 30

    curr_text = ""
    if contract.currency:
        curr_text = contract.currency.upper()
        if contract.exchange_rate:
            curr_text += f" (курс: {contract.exchange_rate})"
    else:
        curr_text = "UAH"

    info_data = [
        ("Фермер", owner_name),
        ("Тип", type_labels.get(contract.contract_type, contract.contract_type)),
        ("Статус", status_labels.get(contract.status, contract.status)),
        ("Дата створення", contract.created_at.strftime("%d.%m.%Y %H:%M") if contract.created_at else "-"),
        ("Сума контракту, грн", round(contract.total_value_uah, 2)),
        ("Залишок, грн", round(contract.balance_uah, 2)),
        ("Валюта / Курс", curr_text),
        ("Примітка", contract.note or "—"),
    ]

    for i, (label, val) in enumerate(info_data, 3):
        cl = ws_info.cell(row=i, column=1)
        cl.value = label
        cl.font = label_font
        cv = ws_info.cell(row=i, column=2)
        cv.value = val
        cv.font = value_font

    ws_info.column_dimensions["A"].width = 25
    ws_info.column_dimensions["B"].width = 40

    # ── Sheet 2: Позиції ──
    ws_items = wb.create_sheet("Позиції")
    item_headers = ["Напрямок", "Позиція", "Кількість, кг", "Ціна, грн/кг", "Сума, грн", "Видано, кг"]
    _fc_apply_header(ws_items, item_headers)
    for item in items:
        ws_items.append([
            dir_labels.get(item.direction, item.direction),
            item.item_name or "-",
            round(item.quantity_kg, 2),
            round(item.price_per_kg, 2),
            round(item.total_value_uah, 2),
            round(item.delivered_kg, 2),
        ])
    _fc_apply_body_style(ws_items, len(item_headers))
    ws_items.freeze_panes = "A2"
    if ws_items.max_row > 1:
        ws_items.auto_filter.ref = f"A1:{chr(64 + len(item_headers))}{ws_items.max_row}"
    item_widths = [20, 28, 16, 16, 16, 14]
    for idx, w in enumerate(item_widths, 1):
        ws_items.column_dimensions[chr(64 + idx)].width = w

    # ── Sheet 3: Операції ──
    ws_pay = wb.create_sheet("Операції")
    pay_headers = ["Дата", "Тип", "Позиція", "Кількість, кг", "Сума", "Валюта", "Курс", "Сума, грн", "Статус"]
    _fc_apply_header(ws_pay, pay_headers)
    active_fill = PatternFill("solid", fgColor="BBF7D0")
    cancelled_fill = PatternFill("solid", fgColor="FECACA")
    for p in payments:
        ws_pay.append([
            p.payment_date.strftime("%d.%m.%Y") if p.payment_date else "-",
            pay_type_labels.get(p.payment_type, p.payment_type),
            p.item_name or "-",
            round(p.quantity_kg, 2) if p.quantity_kg else "-",
            round(p.amount, 2) if p.amount else "-",
            (p.currency or "UAH").upper(),
            p.exchange_rate if p.exchange_rate else "-",
            round(p.amount_uah, 2) if p.amount_uah else "-",
            "Скасовано" if p.is_cancelled else "Активна"
        ])
    for row in range(2, ws_pay.max_row + 1):
        status_cell = ws_pay.cell(row=row, column=9)
        if status_cell.value == "Скасовано":
            status_cell.fill = cancelled_fill
        else:
            status_cell.fill = active_fill
        status_cell.alignment = Alignment(horizontal="center")
    _fc_apply_body_style(ws_pay, len(pay_headers))
    ws_pay.freeze_panes = "A2"
    if ws_pay.max_row > 1:
        ws_pay.auto_filter.ref = f"A1:{chr(64 + len(pay_headers))}{ws_pay.max_row}"
    pay_widths = [14, 14, 25, 16, 14, 10, 10, 16, 14]
    for idx, w in enumerate(pay_widths, 1):
        ws_pay.column_dimensions[chr(64 + idx)].width = w

    return _fc_to_response(
        wb, f"contract_{contract_id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )


@router.get("/{contract_id}", response_model=FarmerContractDetailResponse)
async def get_farmer_contract(
    contract_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    contract = session.get(FarmerContract, contract_id)
    if not contract:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Контракт не знайдено")
    items = session.exec(
        select(FarmerContractItem).where(FarmerContractItem.contract_id == contract_id)
    ).all()
    # Залишок боргу рахуємо з оплат боргу (гроші/зерно)
    balance_uah = _compute_contract_balance_uah(session, contract_id, contract.total_value_uah)
    data = contract.model_dump() if hasattr(contract, "model_dump") else contract.dict()
    data["balance_uah"] = balance_uah
    return FarmerContractDetailResponse(
        **data,
        items=[FarmerContractItemResponse.from_orm(item) for item in items]
    )


@router.post("", response_model=FarmerContractDetailResponse)
async def create_farmer_contract(
    payload: FarmerContractCreate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_super_admin)
):
    # Контрагент: фермер або людина — рівно один.
    has_owner = bool(payload.owner_id)
    has_person = bool(payload.person_id)
    if has_owner == has_person:
        raise HTTPException(
            status_code=400,
            detail="Вкажіть або фермера, або людину (рівно одне з полів owner_id/person_id)"
        )

    owner = None
    person = None
    counterparty_label = ""
    if has_owner:
        owner = session.get(GrainOwner, payload.owner_id)
        if not owner:
            raise HTTPException(status_code=404, detail="Фермера не знайдено")
        counterparty_label = owner.full_name
    else:
        person = session.get(Person, payload.person_id)
        if not person:
            raise HTTPException(status_code=404, detail="Людину не знайдено")
        counterparty_label = person.full_name

    ctype = payload.contract_type

    # Обмеження для контрактів з людиною. Дозволені:
    #   • DEBT — «Контракт»: людина дає нам гроші/зерно за наші товари/гроші.
    #   • PAYMENT — «Викуп»: людина продає нам своє зерно (з person_quantity_kg на складі)
    #     за гроші — миттєвий розрахунок.
    # Резерв (RESERVE) і EXCHANGE — заборонені.
    if person:
        if ctype not in (FarmerContractType.DEBT, FarmerContractType.PAYMENT):
            raise HTTPException(
                status_code=400,
                detail="Для людини доступні типи «Контракт» або «Викуп/Виплата»"
            )
        # У DEBT-контракті від людини дозволені лише гроші. У PAYMENT — лише зерно (це викуп).
        if ctype == FarmerContractType.DEBT:
            for fi in payload.farmer_items:
                it = fi.item_type.value if hasattr(fi.item_type, "value") else fi.item_type
                if it != FarmerContractItemType.CASH.value:
                    raise HTTPException(
                        status_code=400,
                        detail="Людина в контракті може розраховуватись лише грошима"
                    )
        # Талони не виписуються на людину
        for ci in payload.company_items:
            it = ci.item_type.value if hasattr(ci.item_type, "value") else ci.item_type
            if it == FarmerContractItemType.VOUCHER.value:
                raise HTTPException(
                    status_code=400,
                    detail="Талон на зерно недоступний для контракту з людиною"
                )

    # ─────────────────────────────────────────────
    #  PAYMENT — контракт виплати/викупу (миттєвий розрахунок).
    #  Працює і для фермера, і для людини. Різниця лише у:
    #   • джерелі балансу для перевірки доступного зерна
    #   • бакеті складу, з якого списуємо (farmer vs person)
    #   • чи створюємо FarmerGrainDeduction (тільки для фермера)
    # ─────────────────────────────────────────────
    if ctype == FarmerContractType.PAYMENT:
        if not payload.farmer_items:
            raise HTTPException(status_code=400, detail="Вкажіть зерно для викупу")

        # Валідуємо та збираємо позиції
        farmer_total = 0.0
        items_to_create = []
        for fi in payload.farmer_items:
            if fi.item_type != FarmerContractItemType.GRAIN or not fi.culture_id:
                raise HTTPException(status_code=400, detail="Для контракту виплати потрібні лише культури зерна")
            culture = session.get(GrainCulture, fi.culture_id)
            if not culture:
                raise HTTPException(status_code=404, detail="Культуру не знайдено")
            price = fi.price_per_kg or culture.price_per_kg
            if person:
                available = _get_person_balance(session, payload.person_id, fi.culture_id)
            else:
                available = _get_farmer_balance(session, payload.owner_id, fi.culture_id)
            if fi.quantity_kg > available + 0.01:
                raise HTTPException(status_code=400, detail=f"Недостатньо {culture.name} на балансі. Доступно: {available:.2f}")
            total_uah = fi.quantity_kg * price
            farmer_total += total_uah
            items_to_create.append(FarmerContractItem(
                contract_id=0,
                direction=FarmerContractItemDirection.FROM_FARMER.value,
                item_type=FarmerContractItemType.GRAIN.value,
                culture_id=fi.culture_id,
                item_name=culture.name,
                quantity_kg=fi.quantity_kg,
                price_per_kg=price,
                total_value_uah=total_uah,
                delivered_kg=fi.quantity_kg  # одразу видано
            ))

        # Визначаємо суму у вибраній валюті
        currency_str = payload.currency or "UAH"
        currency_enum = Currency(currency_str)
        exchange_rate = payload.exchange_rate
        if currency_enum != Currency.UAH and (not exchange_rate or exchange_rate <= 0):
            raise HTTPException(status_code=400, detail="Вкажіть курс валюти")
        if currency_enum == Currency.UAH:
            payout_amount = farmer_total
        else:
            payout_amount = round(farmer_total / exchange_rate, 2)

        # Створюємо контракт (відразу закритий)
        contract = FarmerContract(
            owner_id=payload.owner_id,
            person_id=payload.person_id,
            contract_type=FarmerContractType.PAYMENT.value,
            status=FarmerContractStatus.CLOSED.value,
            total_value_uah=farmer_total,
            balance_uah=0.0,
            currency=currency_str,
            exchange_rate=exchange_rate,
            note=payload.note,
            created_by_user_id=current_user.id
        )
        session.add(contract)
        session.flush()

        for item in items_to_create:
            item.contract_id = contract.id
            session.add(item)
        session.flush()

        # Рухи на складі: зерно контрагента → наше.
        for item in items_to_create:
            stock = _get_or_create_grain_stock(session, item.culture_id)
            if person:
                stock.person_quantity_kg = max(0.0, (stock.person_quantity_kg or 0.0) - item.quantity_kg)
            else:
                stock.farmer_quantity_kg = max(0.0, stock.farmer_quantity_kg - item.quantity_kg)
            stock.own_quantity_kg += item.quantity_kg
            session.add(stock)
            # Списання з балансу фермера — лише для фермера. Балас людини
            # автоматично перерахується через GRAIN-payments + transfers.
            if not person:
                session.add(FarmerGrainDeduction(
                    owner_id=payload.owner_id,
                    culture_id=item.culture_id,
                    quantity_kg=item.quantity_kg
                ))

        # Списуємо гроші з каси
        cash_register = _get_cash_register(session)
        balance_field_map = {Currency.UAH: "uah_balance", Currency.USD: "usd_balance", Currency.EUR: "eur_balance"}
        field = balance_field_map[currency_enum]
        new_balance = getattr(cash_register, field) - payout_amount
        setattr(cash_register, field, new_balance)
        session.add(cash_register)
        session.add(Transaction(
            currency=currency_enum,
            amount=payout_amount,
            transaction_type=TransactionType.SUBTRACT,
            user_id=current_user.id,
            description=f"Виплата за контрактом #{contract.id} ({counterparty_label})",
            uah_balance_after=cash_register.uah_balance,
            usd_balance_after=cash_register.usd_balance,
            eur_balance_after=cash_register.eur_balance
        ))

        # Запис виплати (settlement)
        settlement = FarmerContractPayment(
            contract_id=contract.id,
            payment_type=FarmerContractPaymentType.SETTLEMENT.value,
            item_name=f"Виплата {currency_str}",
            amount=payout_amount,
            currency=currency_enum,
            exchange_rate=exchange_rate,
            amount_uah=farmer_total,
            created_by_user_id=current_user.id
        )
        session.add(settlement)
        session.commit()
        session.refresh(contract)

        return FarmerContractDetailResponse(
            **contract.dict(),
            items=[FarmerContractItemResponse.from_orm(i) for i in items_to_create]
        )

    # ─────────────────────────────────────────────
    #  RESERVE — контракт резерву (очікує наявності)
    # ─────────────────────────────────────────────
    if ctype == FarmerContractType.RESERVE:
        if not payload.company_items:
            raise HTTPException(status_code=400, detail="Вкажіть що фермер хоче зарезервувати")

        company_total = 0.0
        items_to_create = []
        for ci in payload.company_items:
            if ci.quantity_kg <= 0:
                raise HTTPException(status_code=400, detail="Кількість має бути більше 0")
            # Ціна при резерві не вказується — буде вказана при активації
            price = ci.price_per_kg if (ci.price_per_kg and ci.price_per_kg > 0) else 0.0
            purchase_stock_id = ci.purchase_stock_id
            culture_id = ci.culture_id

            if ci.item_type == FarmerContractItemType.GRAIN:
                # Зерно — шукаємо по culture_id
                if ci.culture_id:
                    culture = session.get(GrainCulture, ci.culture_id)
                    if not culture:
                        raise HTTPException(status_code=404, detail="Культуру не знайдено")
                    item_name = culture.name
                    stock = _get_or_create_grain_stock(session, ci.culture_id)
                    stock.reserved_kg += ci.quantity_kg
                    session.add(stock)
                else:
                    raise HTTPException(status_code=400, detail="Оберіть культуру")
            elif ci.item_type == FarmerContractItemType.PURCHASE:
                # Товар — шукаємо за id або створюємо за назвою
                if ci.purchase_stock_id:
                    stock = session.get(PurchaseStock, ci.purchase_stock_id)
                    if not stock:
                        raise HTTPException(status_code=404, detail="Позицію складу не знайдено")
                    item_name = stock.name
                elif ci.item_name:
                    # Нова позиція — створюємо на складі з 0 кількістю
                    stock = _find_or_create_purchase_stock_by_name(session, ci.item_name)
                    purchase_stock_id = stock.id
                    item_name = stock.name
                else:
                    raise HTTPException(status_code=400, detail="Вкажіть назву позиції")
                stock.reserved_kg += ci.quantity_kg
                session.add(stock)
            else:
                item_name = ci.item_name or "Інше"

            total_uah = ci.quantity_kg * price
            company_total += total_uah
            items_to_create.append(FarmerContractItem(
                contract_id=0,
                direction=FarmerContractItemDirection.FROM_COMPANY.value,
                item_type=ci.item_type.value if hasattr(ci.item_type, 'value') else ci.item_type,
                culture_id=culture_id,
                purchase_stock_id=purchase_stock_id,
                item_name=item_name,
                quantity_kg=ci.quantity_kg,
                price_per_kg=price,
                total_value_uah=total_uah
            ))

        contract = FarmerContract(
            owner_id=payload.owner_id,
            person_id=payload.person_id,
            contract_type=FarmerContractType.RESERVE.value,
            status=FarmerContractStatus.PENDING.value,
            total_value_uah=company_total,
            balance_uah=company_total,
            note=payload.note,
            created_by_user_id=current_user.id
        )
        session.add(contract)
        session.flush()

        for item in items_to_create:
            item.contract_id = contract.id
            session.add(item)

        session.commit()
        session.refresh(contract)

        return FarmerContractDetailResponse(
            **contract.dict(),
            items=[FarmerContractItemResponse.from_orm(i) for i in items_to_create]
        )

    # ─────────────────────────────────────────────
    #  DEBT / EXCHANGE — борговий або обмінний контракт
    # ─────────────────────────────────────────────
    if not payload.farmer_items and not payload.company_items:
        raise HTTPException(status_code=400, detail="Додайте позиції контракту")

    company_total = 0.0
    farmer_total = 0.0
    items_to_create: list[FarmerContractItem] = []

    def build_item(item, direction):
        if item.quantity_kg <= 0:
            raise HTTPException(status_code=400, detail="Кількість має бути більше 0")

        if item.item_type == FarmerContractItemType.GRAIN:
            if not item.culture_id:
                raise HTTPException(status_code=400, detail="Оберіть культуру")
            culture = session.get(GrainCulture, item.culture_id)
            if not culture:
                raise HTTPException(status_code=404, detail="Культуру не знайдено")
            price = item.price_per_kg or culture.price_per_kg
            item_name = culture.name
            if direction == FarmerContractItemDirection.FROM_COMPANY:
                stock = _get_or_create_grain_stock(session, item.culture_id)
                # Дозволяємо контракт на всю наявну кількість на складі (в т.ч. невикуплене зерно фермерів)
                available = stock.quantity_kg - stock.reserved_kg
                if item.quantity_kg > available + 0.01:
                    raise HTTPException(status_code=400, detail=f"Недостатньо зерна {culture.name} на складі. Доступно: {available:.2f}")
                stock.reserved_kg += item.quantity_kg
            else:
                available = _get_farmer_balance(session, payload.owner_id, item.culture_id)
                if item.quantity_kg > available + 0.01:
                    raise HTTPException(status_code=400, detail=f"Недостатньо {culture.name} на балансі фермера. Доступно: {available:.2f}")
        elif item.item_type == FarmerContractItemType.PURCHASE:
            if not item.purchase_stock_id:
                raise HTTPException(status_code=400, detail="Оберіть позицію складу")
            stock = session.get(PurchaseStock, item.purchase_stock_id)
            if not stock:
                raise HTTPException(status_code=404, detail="Позицію складу не знайдено")
            price = item.price_per_kg or stock.sale_price_per_kg
            if direction == FarmerContractItemDirection.FROM_COMPANY:
                available = stock.quantity_kg - stock.reserved_kg
                if item.quantity_kg > available + 0.01:
                    raise HTTPException(status_code=400, detail=f"Недостатньо {stock.name} на складі. Доступно: {available:.2f}")
                stock.reserved_kg += item.quantity_kg
            item_name = stock.name
        elif item.item_type == FarmerContractItemType.VOUCHER:
            if not item.culture_id:
                raise HTTPException(status_code=400, detail="Оберіть культуру для талону")
            culture = session.get(GrainCulture, item.culture_id)
            if not culture:
                raise HTTPException(status_code=404, detail="Культуру не знайдено")
            if culture.name != "Пшениця":
                raise HTTPException(status_code=400, detail="Талони можна виписувати тільки на Пшеницю")
            price = item.price_per_kg or culture.price_per_kg
            item_name = f"Талон: {culture.name}"
            # Талон не резервує фізичне зерно на складі
        elif item.item_type == FarmerContractItemType.LAND_SERVICE:
            # Обробка землі: quantity_kg фактично = гектари, price_per_kg = грн/га.
            # Послуга — складських залишків не торкаємось, тільки фіксуємо суму.
            if direction != FarmerContractItemDirection.FROM_COMPANY:
                raise HTTPException(status_code=400, detail="Обробка землі може бути тільки від компанії")
            if not item.price_per_kg or item.price_per_kg <= 0:
                raise HTTPException(status_code=400, detail="Вкажіть ціну за гектар")
            price = item.price_per_kg
            item_name = "Обробка землі"
        else:
            # Гроші: price_per_kg = курс (1 для UAH)
            currency_str = getattr(item, "currency", None) or "UAH"
            price = item.price_per_kg or 1.0
            if currency_str != "UAH" and (not price or price <= 0):
                raise HTTPException(status_code=400, detail="Вкажіть курс валюти до гривні")
            item_name = "Готівка"

        item_kw = dict(
            contract_id=0,
            direction=direction.value if hasattr(direction, 'value') else direction,
            item_type=item.item_type.value if hasattr(item.item_type, 'value') else item.item_type,
            culture_id=item.culture_id,
            purchase_stock_id=item.purchase_stock_id,
            item_name=item_name,
            quantity_kg=item.quantity_kg,
            price_per_kg=price,
            total_value_uah=item.quantity_kg * price,
        )
        if getattr(item, "item_type", None) == FarmerContractItemType.CASH or (
            hasattr(item.item_type, "value") and item.item_type.value == "cash"
        ):
            item_kw["currency"] = getattr(item, "currency", None) or "UAH"
        items_to_create.append(FarmerContractItem(**item_kw))
        return item.quantity_kg * price

    for item in payload.company_items:
        item_value = build_item(item, FarmerContractItemDirection.FROM_COMPANY)
        company_total += item_value

    for item in payload.farmer_items:
        farmer_total += build_item(item, FarmerContractItemDirection.FROM_FARMER)

    # Залишок боргу = сума від компанії (включно з талонами) мінус сума від фермера
    contract = FarmerContract(
        owner_id=payload.owner_id,
        person_id=payload.person_id,
        contract_type=payload.contract_type.value if hasattr(payload.contract_type, 'value') else payload.contract_type,
        status=FarmerContractStatus.OPEN.value,
        total_value_uah=company_total,
        balance_uah=max(company_total - farmer_total, 0.0),
        note=payload.note,
        created_by_user_id=current_user.id
    )
    session.add(contract)
    session.flush()

    for item in items_to_create:
        item.contract_id = contract.id
        session.add(item)

    session.commit()
    session.refresh(contract)

    return FarmerContractDetailResponse(
        **contract.dict(),
        items=[FarmerContractItemResponse.from_orm(i) for i in items_to_create]
    )


@router.post("/{contract_id}/settle", response_model=FarmerContractDetailResponse)
async def settle_payment_contract(
    contract_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_super_admin)
):
    """Виконати розрахунок по відкритому PAYMENT-контракту (Виплата).
    Один крок: списати зерно з балансу контрагента, додати у склад підприємства,
    видати гроші з каси, створити settlement-payment, закрити контракт.
    Використовує позиції контракту як є (qty + price з створення).
    """
    contract = session.get(FarmerContract, contract_id)
    if not contract:
        raise HTTPException(status_code=404, detail="Контракт не знайдено")
    if contract.contract_type != FarmerContractType.PAYMENT.value:
        raise HTTPException(status_code=400, detail="Розрахунок можливий лише для контрактів типу Виплата")
    if contract.status != FarmerContractStatus.OPEN.value:
        raise HTTPException(status_code=400, detail="Контракт не у статусі «Відкритий»")

    items = session.exec(
        select(FarmerContractItem).where(FarmerContractItem.contract_id == contract_id)
    ).all()
    if not items:
        raise HTTPException(status_code=400, detail="У контракті немає позицій")

    is_person = bool(contract.person_id)
    person = session.get(Person, contract.person_id) if is_person else None
    owner = session.get(GrainOwner, contract.owner_id) if contract.owner_id else None
    if is_person and not person:
        raise HTTPException(status_code=404, detail="Людину не знайдено")
    if not is_person and not owner:
        raise HTTPException(status_code=404, detail="Фермера не знайдено")

    # Перевіряємо доступний баланс по кожній GRAIN-позиції
    farmer_total = 0.0
    for item in items:
        if item.item_type != FarmerContractItemType.GRAIN.value or not item.culture_id:
            raise HTTPException(status_code=400, detail="У контракті виплати лише зернові позиції")
        qty = item.quantity_kg or 0.0
        if qty <= 0:
            continue
        if is_person:
            available = _get_person_balance(session, contract.person_id, item.culture_id)
        else:
            available = _get_farmer_balance(session, contract.owner_id, item.culture_id)
        if qty > available + 0.01:
            culture = session.get(GrainCulture, item.culture_id)
            cname = culture.name if culture else f"#{item.culture_id}"
            raise HTTPException(
                status_code=400,
                detail=f"Недостатньо {cname} на балансі. Доступно: {available:.2f}, потрібно: {qty:.2f}"
            )
        farmer_total += qty * (item.price_per_kg or 0.0)

    # Сума виплати у валюті контракту
    currency_str = contract.currency or "UAH"
    try:
        currency_enum = Currency(currency_str)
    except ValueError:
        currency_enum = Currency.UAH
    exchange_rate = contract.exchange_rate or 1.0
    if currency_enum == Currency.UAH:
        payout_amount = farmer_total
    else:
        if not exchange_rate or exchange_rate <= 0:
            raise HTTPException(status_code=400, detail="Невалідний курс валюти у контракті")
        payout_amount = round(farmer_total / exchange_rate, 2)

    # Рух зерна: контрагент → склад підприємства
    for item in items:
        qty = item.quantity_kg or 0.0
        if qty <= 0:
            continue
        stock = _get_or_create_grain_stock(session, item.culture_id)
        if is_person:
            stock.person_quantity_kg = max(0.0, (stock.person_quantity_kg or 0.0) - qty)
        else:
            stock.farmer_quantity_kg = max(0.0, stock.farmer_quantity_kg - qty)
        stock.own_quantity_kg += qty
        session.add(stock)
        item.delivered_kg = qty
        session.add(item)
        # Списання з балансу фермера (для людини deduction-ів немає)
        if not is_person:
            session.add(FarmerGrainDeduction(
                owner_id=contract.owner_id,
                culture_id=item.culture_id,
                quantity_kg=qty
            ))

    # Списуємо гроші з каси + Transaction
    cash_register = _get_cash_register(session)
    balance_field_map = {Currency.UAH: "uah_balance", Currency.USD: "usd_balance", Currency.EUR: "eur_balance"}
    field = balance_field_map[currency_enum]
    setattr(cash_register, field, getattr(cash_register, field) - payout_amount)
    session.add(cash_register)
    counterparty_label = (person.full_name if person else (owner.full_name if owner else "?"))
    session.add(Transaction(
        currency=currency_enum,
        amount=payout_amount,
        transaction_type=TransactionType.SUBTRACT,
        user_id=current_user.id,
        description=f"Виплата за контрактом #{contract.id} ({counterparty_label})",
        uah_balance_after=cash_register.uah_balance,
        usd_balance_after=cash_register.usd_balance,
        eur_balance_after=cash_register.eur_balance
    ))

    # Settlement payment
    session.add(FarmerContractPayment(
        contract_id=contract.id,
        payment_type=FarmerContractPaymentType.SETTLEMENT.value,
        item_name=f"Виплата {currency_str}",
        amount=payout_amount,
        currency=currency_enum,
        exchange_rate=exchange_rate,
        amount_uah=farmer_total,
        created_by_user_id=current_user.id
    ))

    # Закриваємо контракт
    contract.balance_uah = 0.0
    contract.status = FarmerContractStatus.CLOSED.value
    session.add(contract)

    try:
        session.commit()
    except Exception:
        session.rollback()
        raise
    session.refresh(contract)
    fresh_items = session.exec(
        select(FarmerContractItem).where(FarmerContractItem.contract_id == contract.id)
    ).all()
    return FarmerContractDetailResponse(
        **contract.dict(),
        items=[FarmerContractItemResponse.from_orm(i) for i in fresh_items]
    )


@router.post("/{contract_id}/activate", response_model=FarmerContractResponse)
async def activate_reserve_contract(
    contract_id: int,
    payload: Optional[ReserveActivateRequest] = None,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_super_admin)
):
    """Активувати резервний контракт → перетворити в борговий. Ціни позицій вказуються в payload (при створенні резерву ціна не вказувалась)."""
    contract = session.get(FarmerContract, contract_id)
    if not contract:
        raise HTTPException(status_code=404, detail="Контракт не знайдено")
    if contract.contract_type != FarmerContractType.RESERVE.value:
        raise HTTPException(status_code=400, detail="Тільки резервні контракти можна активувати")
    if contract.status != FarmerContractStatus.PENDING.value:
        raise HTTPException(status_code=400, detail="Контракт не в стані очікування")

    items = session.exec(
        select(FarmerContractItem).where(FarmerContractItem.contract_id == contract_id)
    ).all()
    if not items:
        raise HTTPException(status_code=400, detail="Контракт без позицій")

    # Перевіряємо наявність на складі (враховуємо всю кількість, в т.ч. невикуплене зерно)
    for item in items:
        if item.item_type == FarmerContractItemType.GRAIN.value and item.culture_id:
            stock = _get_or_create_grain_stock(session, item.culture_id)
            available = stock.quantity_kg - (stock.reserved_kg - item.quantity_kg)
            if item.quantity_kg > available + 0.01:
                raise HTTPException(
                    status_code=400,
                    detail=f"Недостатньо {item.item_name} на складі для активації. Потрібно: {item.quantity_kg:.2f}, доступно: {available:.2f}"
                )
        elif item.item_type == FarmerContractItemType.PURCHASE.value:
            pstock = _find_purchase_stock_for_activation(session, item)
            if not pstock:
                raise HTTPException(
                    status_code=400,
                    detail=f"Позицію «{item.item_name}» не знайдено на складі або недостатньо кількості для активації"
                )
            available = pstock.quantity_kg - pstock.reserved_kg
            if pstock.id == item.purchase_stock_id:
                available += item.quantity_kg
            if item.quantity_kg > available + 0.01:
                raise HTTPException(
                    status_code=400,
                    detail=f"Недостатньо {item.item_name} на складі для активації. Потрібно: {item.quantity_kg:.2f}, доступно: {available:.2f}"
                )
            # Привʼязуємо позицію до того складу, де є товар (назва може відрізнятися написанням)
            if pstock.id != item.purchase_stock_id:
                old_id = item.purchase_stock_id
                item.purchase_stock_id = pstock.id
                item.item_name = pstock.name
                session.add(item)
                if old_id:
                    old_stock = session.get(PurchaseStock, old_id)
                    if old_stock:
                        old_stock.reserved_kg = max(0.0, old_stock.reserved_kg - item.quantity_kg)
                        session.add(old_stock)
                pstock.reserved_kg += item.quantity_kg
                session.add(pstock)

    # Оновлюємо ціни позицій і рахуємо суму контракту
    price_by_id = {}
    if payload and payload.items:
        for p in payload.items:
            if p.price_per_kg < 0:
                raise HTTPException(status_code=400, detail=f"Ціна для позиції {p.contract_item_id} не може бути від'ємною")
            price_by_id[p.contract_item_id] = p.price_per_kg

    company_total = 0.0
    for item in items:
        if item.id in price_by_id:
            price = price_by_id[item.id]
        elif item.item_type == FarmerContractItemType.GRAIN.value and item.culture_id:
            culture = session.get(GrainCulture, item.culture_id)
            price = culture.price_per_kg if culture else 0.0
        elif item.item_type == FarmerContractItemType.PURCHASE.value and item.purchase_stock_id:
            pstock = session.get(PurchaseStock, item.purchase_stock_id)
            price = pstock.sale_price_per_kg if pstock else 0.0
        else:
            price = item.price_per_kg or 0.0
        item.price_per_kg = price
        item.total_value_uah = item.quantity_kg * price
        company_total += item.total_value_uah
        session.add(item)

    contract.contract_type = FarmerContractType.DEBT.value
    contract.status = FarmerContractStatus.OPEN.value
    contract.total_value_uah = company_total
    contract.balance_uah = company_total
    contract.was_reserve = True
    contract.note = (contract.note or "") + " [Активовано з резерву]"
    session.add(contract)
    session.commit()
    session.refresh(contract)
    return contract


@router.post("/{contract_id}/close", response_model=FarmerContractResponse)
async def close_farmer_contract(
    contract_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_super_admin)
):
    """Закрити контракт вручну"""
    contract = session.get(FarmerContract, contract_id)
    if not contract:
        raise HTTPException(status_code=404, detail="Контракт не знайдено")
    if contract.status not in (FarmerContractStatus.OPEN.value, FarmerContractStatus.PENDING.value):
        raise HTTPException(status_code=400, detail="Контракт вже закритий або скасований")

    # Знімаємо бронювання для невиданих позицій
    items = session.exec(
        select(FarmerContractItem).where(FarmerContractItem.contract_id == contract_id)
    ).all()
    for item in items:
        unreserve_qty = max(0.0, item.quantity_kg - item.delivered_kg)
        if unreserve_qty <= 0:
            continue
        if item.direction != FarmerContractItemDirection.FROM_COMPANY.value:
            continue
        if item.item_type == FarmerContractItemType.GRAIN.value and item.culture_id:
            stock = _get_or_create_grain_stock(session, item.culture_id)
            stock.reserved_kg = max(0.0, stock.reserved_kg - unreserve_qty)
            session.add(stock)
        elif item.item_type == FarmerContractItemType.PURCHASE.value and item.purchase_stock_id:
            pstock = session.get(PurchaseStock, item.purchase_stock_id)
            if pstock:
                pstock.reserved_kg = max(0.0, pstock.reserved_kg - unreserve_qty)
                session.add(pstock)

    contract.status = FarmerContractStatus.CLOSED.value
    session.add(contract)
    session.commit()
    session.refresh(contract)
    return contract


@router.get("/{contract_id}/payments", response_model=list[FarmerContractPaymentResponse])
async def list_farmer_contract_payments(
    contract_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    contract = session.get(FarmerContract, contract_id)
    if not contract:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Контракт не знайдено")
    payments = session.exec(
        select(FarmerContractPayment).where(FarmerContractPayment.contract_id == contract_id)
        .order_by(FarmerContractPayment.payment_date.desc())
    ).all()
    return payments


def _check_contract_completion(session: Session, contract: FarmerContract):
    """Перевірити чи контракт повністю виконаний.
    Контракт закривається коли:
    - balance_uah <= 0 (включно з талонами)
    - всі позиції (в т.ч. талонні) видані/прийняті за потреби
    """
    if contract.balance_uah > 0.01:
        return  # Ще є борг

    items = session.exec(
        select(FarmerContractItem).where(FarmerContractItem.contract_id == contract.id)
    ).all()

    if not items:
        contract.status = FarmerContractStatus.CLOSED.value
        return

    # Закриваємо тільки коли ВСІ позиції фактично виконані.
    # Для талонів це теж "виконання": delivered_kg має дорівнювати quantity_kg.
    all_delivered = all(item.delivered_kg >= item.quantity_kg - 0.01 for item in items)
    if all_delivered:
        contract.status = FarmerContractStatus.CLOSED.value


@router.post("/{contract_id}/payments", response_model=FarmerContractPaymentResponse)
async def create_farmer_contract_payment(
    contract_id: int,
    payload: FarmerContractPaymentCreate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_super_admin)
):
    contract = session.get(FarmerContract, contract_id)
    if not contract:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Контракт не знайдено")
    if contract.status != FarmerContractStatus.OPEN.value:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Контракт не активний")

    # Контрагент: фермер або людина
    owner = session.get(GrainOwner, contract.owner_id) if contract.owner_id else None
    person = session.get(Person, contract.person_id) if contract.person_id else None

    # Для контрактів з людиною дозволено: оплату готівкою, видачу товару, і — якщо у людини
    # є зерно на нашому складі (від переказу фермером) — оплату цим зерном.
    # Талон/прийом товару від людини — досі заборонені.
    if person:
        allowed_for_person = {
            FarmerContractPaymentType.CASH,
            FarmerContractPaymentType.GOODS_ISSUE,
            FarmerContractPaymentType.GRAIN,
        }
        if payload.payment_type not in allowed_for_person:
            raise HTTPException(
                status_code=400,
                detail="Для контракту з людиною доступні лише оплата готівкою/зерном або видача товару"
            )

    amount_uah = 0.0
    payment = None

    # ─── GOODS_ISSUE: ми видаємо товар фермеру ───
    if payload.payment_type == FarmerContractPaymentType.GOODS_ISSUE:
        if not payload.contract_item_id:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Оберіть позицію контракту")
        if not payload.quantity_kg or payload.quantity_kg <= 0:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Вкажіть кількість")

        item = session.get(FarmerContractItem, payload.contract_item_id)
        if not item or item.contract_id != contract_id:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Позицію не знайдено")
        if item.direction != FarmerContractItemDirection.FROM_COMPANY.value:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Ця позиція не від компанії")

        remaining = item.quantity_kg - item.delivered_kg
        if payload.quantity_kg > remaining + 0.01:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Кількість перевищує залишок. Залишок: {remaining:.2f}"
            )

        amount_uah = payload.quantity_kg * item.price_per_kg

        # Списуємо зі складу
        if item.item_type == FarmerContractItemType.GRAIN.value and item.culture_id:
            stock = _get_or_create_grain_stock(session, item.culture_id)
            stock.reserved_kg = max(0.0, stock.reserved_kg - payload.quantity_kg)
            stock.own_quantity_kg = max(0.0, stock.own_quantity_kg - payload.quantity_kg)
            stock.quantity_kg = max(0.0, stock.quantity_kg - payload.quantity_kg)
            session.add(stock)
        elif item.item_type == FarmerContractItemType.PURCHASE.value and item.purchase_stock_id:
            pstock = session.get(PurchaseStock, item.purchase_stock_id)
            if pstock:
                pstock.reserved_kg = max(0.0, pstock.reserved_kg - payload.quantity_kg)
                pstock.quantity_kg = max(0.0, pstock.quantity_kg - payload.quantity_kg)
                session.add(pstock)
        elif item.item_type == FarmerContractItemType.CASH.value:
            cash_register = _get_cash_register(session)
            currency_str = getattr(item, "currency", None) or "UAH"
            try:
                currency_enum = Currency(currency_str)
            except ValueError:
                currency_enum = Currency.UAH
            balance_field_map = {Currency.UAH: "uah_balance", Currency.USD: "usd_balance", Currency.EUR: "eur_balance"}
            field = balance_field_map[currency_enum]
            current_bal = getattr(cash_register, field)
            # Дозволяємо касі йти в мінус при виплаті
            setattr(cash_register, field, current_bal - payload.quantity_kg)
            session.add(cash_register)
            session.add(Transaction(
                currency=currency_enum,
                amount=payload.quantity_kg,
                transaction_type=TransactionType.SUBTRACT,
                user_id=current_user.id,
                description=f"Видача за контрактом фермера #{contract_id}",
                uah_balance_after=cash_register.uah_balance,
                usd_balance_after=cash_register.usd_balance,
                eur_balance_after=cash_register.eur_balance
            ))

        item.delivered_kg += payload.quantity_kg
        session.add(item)

        pay_currency = Currency.UAH
        pay_amount = 0.0
        if item.item_type == FarmerContractItemType.CASH.value:
            try:
                pay_currency = Currency(getattr(item, "currency", None) or "UAH")
            except ValueError:
                pay_currency = Currency.UAH
            pay_amount = payload.quantity_kg

        payment = FarmerContractPayment(
            contract_id=contract_id,
            contract_item_id=item.id,
            payment_type=payload.payment_type.value if hasattr(payload.payment_type, 'value') else payload.payment_type,
            item_name=item.item_name,
            amount=pay_amount,
            currency=pay_currency,
            exchange_rate=item.price_per_kg if item.item_type == FarmerContractItemType.CASH.value and getattr(item, "currency", "UAH") != "UAH" else None,
            amount_uah=amount_uah,
            quantity_kg=payload.quantity_kg,
            culture_id=item.culture_id,
            created_by_user_id=current_user.id
        )
        session.add(payment)
        # Видача НЕ зменшує balance_uah

    # ─── GOODS_RECEIVE: фермер передає товар нам (обмін) ───
    elif payload.payment_type == FarmerContractPaymentType.GOODS_RECEIVE:
        if not payload.contract_item_id:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Оберіть позицію контракту")
        if not payload.quantity_kg or payload.quantity_kg <= 0:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Вкажіть кількість")

        item = session.get(FarmerContractItem, payload.contract_item_id)
        if not item or item.contract_id != contract_id:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Позицію не знайдено")
        if item.direction != FarmerContractItemDirection.FROM_FARMER.value:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Ця позиція не від фермера")

        remaining = item.quantity_kg - item.delivered_kg
        if payload.quantity_kg > remaining + 0.01:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Кількість перевищує залишок. Залишок: {remaining:.2f}"
            )

        amount_uah = payload.quantity_kg * item.price_per_kg

        # Приймаємо товар від фермера
        if item.item_type == FarmerContractItemType.GRAIN.value and item.culture_id:
            available = _get_farmer_balance(session, contract.owner_id, item.culture_id)
            if payload.quantity_kg > available + 0.01:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Недостатньо зерна на балансі фермера. Доступно: {available:.2f}"
                )
            stock = _get_or_create_grain_stock(session, item.culture_id)
            stock.farmer_quantity_kg -= payload.quantity_kg
            stock.own_quantity_kg += payload.quantity_kg
            session.add(stock)
        elif item.item_type == FarmerContractItemType.CASH.value:
            cash_register = _get_cash_register(session)
            cash_register.uah_balance += payload.quantity_kg
            session.add(cash_register)
            session.add(Transaction(
                currency=Currency.UAH,
                amount=payload.quantity_kg,
                transaction_type=TransactionType.ADD,
                user_id=current_user.id,
                description=f"Прийом за контрактом фермера #{contract_id}",
                uah_balance_after=cash_register.uah_balance,
                usd_balance_after=cash_register.usd_balance,
                eur_balance_after=cash_register.eur_balance
            ))

        item.delivered_kg += payload.quantity_kg
        session.add(item)

        payment = FarmerContractPayment(
            contract_id=contract_id,
            contract_item_id=item.id,
            payment_type=payload.payment_type.value if hasattr(payload.payment_type, 'value') else payload.payment_type,
            item_name=item.item_name,
            amount=0.0,
            currency=Currency.UAH,
            amount_uah=amount_uah,
            quantity_kg=payload.quantity_kg,
            culture_id=item.culture_id,
            created_by_user_id=current_user.id
        )
        session.add(payment)

        # Для зерна — запис списання з балансу фермера
        if item.item_type == FarmerContractItemType.GRAIN.value and item.culture_id:
            session.flush()
            session.add(FarmerGrainDeduction(
                owner_id=contract.owner_id,
                culture_id=item.culture_id,
                quantity_kg=payload.quantity_kg,
                payment_id=payment.id
            ))
        # Прийом НЕ зменшує balance_uah

    # ─── CASH: фермер платить грошима (зменшує борг) ───
    elif payload.payment_type == FarmerContractPaymentType.CASH:
        if payload.amount is None or payload.amount <= 0:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Вкажіть суму")
        if payload.currency != Currency.UAH:
            if not payload.exchange_rate or payload.exchange_rate <= 0:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Вкажіть курс")
            amount_uah = payload.amount * payload.exchange_rate
        else:
            amount_uah = payload.amount

        if amount_uah > contract.balance_uah + 0.01:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Сума перевищує залишок боргу")

        cash_register = _get_cash_register(session)
        balance_field_map = {
            Currency.UAH: "uah_balance",
            Currency.USD: "usd_balance",
            Currency.EUR: "eur_balance"
        }
        balance_field = balance_field_map[payload.currency]
        new_balance = getattr(cash_register, balance_field) + payload.amount
        setattr(cash_register, balance_field, new_balance)
        session.add(Transaction(
            currency=payload.currency,
            amount=payload.amount,
            transaction_type=TransactionType.ADD,
            user_id=current_user.id,
            description=f"Оплата боргу фермера #{contract_id}",
            uah_balance_after=cash_register.uah_balance,
            usd_balance_after=cash_register.usd_balance,
            eur_balance_after=cash_register.eur_balance
        ))
        session.add(cash_register)

        payment = FarmerContractPayment(
            contract_id=contract_id,
            payment_type=payload.payment_type.value if hasattr(payload.payment_type, 'value') else payload.payment_type,
            item_name="Грошова оплата",
            amount=payload.amount,
            currency=payload.currency,
            exchange_rate=payload.exchange_rate,
            amount_uah=amount_uah,
            created_by_user_id=current_user.id
        )
        session.add(payment)
        contract.balance_uah = max(0.0, contract.balance_uah - amount_uah)

    # ─── GRAIN: контрагент платить зерном з балансу (зменшує борг) ───
    # Для фермера: списується з його балансу через FarmerGrainDeduction.
    # Для людини: списується з person_quantity_kg на складі (зерно, переказане їй фермером).
    # В обох випадках зерно фізично переходить у наш бакет own_quantity_kg.
    elif payload.payment_type == FarmerContractPaymentType.GRAIN:
        if not payload.culture_id or not payload.quantity_kg or payload.quantity_kg <= 0:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Вкажіть культуру та кількість")
        culture = session.get(GrainCulture, payload.culture_id)
        if not culture:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Культуру не знайдено")

        if person:
            available = _get_person_balance(session, contract.person_id, payload.culture_id)
        else:
            available = _get_farmer_balance(session, contract.owner_id, payload.culture_id)
        if payload.quantity_kg > available + 0.01:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Недостатньо зерна на балансі. Доступно: {available:.2f}"
            )

        amount_uah = payload.quantity_kg * culture.price_per_kg
        if amount_uah > contract.balance_uah + 0.01:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Сума перевищує залишок боргу")

        stock = _get_or_create_grain_stock(session, payload.culture_id)
        if person:
            # У людини зерно лежить у person_quantity_kg — переносимо у own.
            stock.person_quantity_kg = max(0.0, (stock.person_quantity_kg or 0.0) - payload.quantity_kg)
        else:
            # У фермера — у farmer_quantity_kg.
            stock.farmer_quantity_kg -= payload.quantity_kg
        stock.own_quantity_kg += payload.quantity_kg
        session.add(stock)

        payment = FarmerContractPayment(
            contract_id=contract_id,
            payment_type=payload.payment_type.value if hasattr(payload.payment_type, 'value') else payload.payment_type,
            item_name=culture.name,
            amount=0.0,
            currency=Currency.UAH,
            amount_uah=amount_uah,
            culture_id=payload.culture_id,
            quantity_kg=payload.quantity_kg,
            created_by_user_id=current_user.id
        )
        session.add(payment)
        session.flush()
        if not person:
            # FarmerGrainDeduction — лише для фермерів (їхній баланс рахується через ці записи).
            # Баланс людини рахується з FarmerContractPayment(GRAIN), деduction не потрібен.
            session.add(FarmerGrainDeduction(
                owner_id=contract.owner_id,
                culture_id=payload.culture_id,
                quantity_kg=payload.quantity_kg,
                payment_id=payment.id
            ))
        contract.balance_uah = max(0.0, contract.balance_uah - amount_uah)

    # ─── VOUCHER: талон на зерно (хлібний завод) ───
    elif payload.payment_type == FarmerContractPaymentType.VOUCHER:
        if not payload.culture_id:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Вкажіть культуру")
        culture = session.get(GrainCulture, payload.culture_id)
        if not culture:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Культуру не знайдено")
        if culture.name != "Пшениця":
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Талони можна виписувати тільки на Пшеницю")

        # Визначаємо кількість з позиції контракту (талон виписується цілком)
        voucher_item = None
        voucher_contract_item_id = None
        if payload.contract_item_id:
            voucher_item = session.get(FarmerContractItem, payload.contract_item_id)
            if not voucher_item or voucher_item.contract_id != contract_id:
                raise HTTPException(status_code=400, detail="Позицію талону не знайдено")
            if voucher_item.item_type != FarmerContractItemType.VOUCHER.value:
                raise HTTPException(status_code=400, detail="Обрана позиція не є талоном")
            remaining = voucher_item.quantity_kg - voucher_item.delivered_kg
            if remaining < 0.01:
                raise HTTPException(status_code=400, detail="Цей талон вже виписано")
            # Талон виписується цілком
            payload.quantity_kg = remaining
        else:
            if not payload.quantity_kg or payload.quantity_kg <= 0:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Вкажіть кількість")
            voucher_item = session.exec(
                select(FarmerContractItem).where(
                    FarmerContractItem.contract_id == contract_id,
                    FarmerContractItem.item_type == FarmerContractItemType.VOUCHER.value,
                    FarmerContractItem.direction == FarmerContractItemDirection.FROM_COMPANY.value,
                    FarmerContractItem.culture_id == payload.culture_id
                )
            ).first()

        qty = payload.quantity_kg
        voucher_price = voucher_item.price_per_kg if voucher_item else culture.price_per_kg
        amount_uah = qty * voucher_price

        # Видача талону не торкається зерна на складі — лише запис у виплатах і талон у розділі «Хлібний завод»
        if voucher_item:
            voucher_item.delivered_kg += qty
            session.add(voucher_item)
            voucher_contract_item_id = voucher_item.id

        payment = FarmerContractPayment(
            contract_id=contract_id,
            contract_item_id=voucher_contract_item_id,
            payment_type=FarmerContractPaymentType.VOUCHER.value,
            item_name=f"Талон: {culture.name}",
            amount=0.0,
            currency=Currency.UAH,
            amount_uah=amount_uah,
            culture_id=payload.culture_id,
            quantity_kg=qty,
            created_by_user_id=current_user.id
        )
        session.add(payment)
        session.flush()

        # Создаём талон
        voucher = GrainVoucher(
            farmer_contract_id=contract_id,
            farmer_contract_payment_id=payment.id,
            owner_id=contract.owner_id,
            culture_id=payload.culture_id,
            quantity_kg=qty,
            price_per_kg=voucher_price,
            total_value_uah=amount_uah,
            paid_value_uah=0.0,
            remaining_value_uah=amount_uah,
            is_closed=False,
            created_by_user_id=current_user.id,
        )
        session.add(voucher)

        # Видача талону НЕ є оплатою боргу фермера, тому не змінює balance_uah.

    else:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Невідомий тип операції")

    # Перевіряємо чи контракт повністю виконаний
    _check_contract_completion(session, contract)
    session.add(contract)

    # Атомарно: stock + cash + transaction + contract balance + payment.
    # Один з найважливіших handler-ів — будь-яка часткова мутація = неконсистентний стан.
    try:
        session.commit()
    except Exception:
        session.rollback()
        raise
    session.refresh(payment)
    return payment


@router.post("/payments/{payment_id}/cancel", response_model=FarmerContractPaymentResponse)
async def cancel_farmer_contract_payment(
    payment_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_super_admin)
):
    """Скасування операції по контракту фермера з поверненням"""
    payment = session.get(FarmerContractPayment, payment_id)
    if not payment:
        raise HTTPException(status_code=404, detail="Операцію не знайдено")

    if payment.is_cancelled:
        raise HTTPException(status_code=400, detail="Операцію вже було скасовано")

    contract = session.get(FarmerContract, payment.contract_id)
    if not contract:
        raise HTTPException(status_code=404, detail="Контракт не знайдено")

    owner = session.get(GrainOwner, contract.owner_id)

    # ─── GOODS_ISSUE: ми видавали товар фермеру → повертаємо на склад ───
    if payment.payment_type == FarmerContractPaymentType.GOODS_ISSUE.value:
        item = session.get(FarmerContractItem, payment.contract_item_id) if payment.contract_item_id else None
        qty = payment.quantity_kg or 0.0

        if item and qty > 0:
            item.delivered_kg = max(0.0, item.delivered_kg - qty)
            session.add(item)

            if item.item_type == FarmerContractItemType.GRAIN.value and item.culture_id:
                stock = _get_or_create_grain_stock(session, item.culture_id)
                stock.own_quantity_kg += qty
                stock.quantity_kg += qty
                stock.reserved_kg += qty
                session.add(stock)
            elif item.item_type == FarmerContractItemType.PURCHASE.value and item.purchase_stock_id:
                pstock = session.get(PurchaseStock, item.purchase_stock_id)
                if pstock:
                    pstock.quantity_kg += qty
                    pstock.reserved_kg += qty
                    session.add(pstock)
            elif item.item_type == FarmerContractItemType.CASH.value:
                cash_register = _get_cash_register(session)
                cash_register.uah_balance += qty
                session.add(cash_register)
                session.add(Transaction(
                    currency=Currency.UAH,
                    amount=qty,
                    transaction_type=TransactionType.ADD,
                    user_id=current_user.id,
                    description=f"Скасування видачі за контрактом #{contract.id} ({owner.full_name if owner else '?'})",
                    uah_balance_after=cash_register.uah_balance,
                    usd_balance_after=cash_register.usd_balance,
                    eur_balance_after=cash_register.eur_balance
                ))

    # ─── GOODS_RECEIVE: фермер передавав товар нам → повертаємо фермеру ───
    elif payment.payment_type == FarmerContractPaymentType.GOODS_RECEIVE.value:
        item = session.get(FarmerContractItem, payment.contract_item_id) if payment.contract_item_id else None
        qty = payment.quantity_kg or 0.0

        if item and qty > 0:
            item.delivered_kg = max(0.0, item.delivered_kg - qty)
            session.add(item)

            if item.item_type == FarmerContractItemType.GRAIN.value and item.culture_id:
                stock = _get_or_create_grain_stock(session, item.culture_id)
                stock.own_quantity_kg = max(0.0, stock.own_quantity_kg - qty)
                stock.farmer_quantity_kg += qty
                session.add(stock)

                # Видаляємо списання з балансу фермера
                deduction = session.exec(
                    select(FarmerGrainDeduction).where(
                        FarmerGrainDeduction.payment_id == payment.id
                    )
                ).first()
                if deduction:
                    session.delete(deduction)

            elif item.item_type == FarmerContractItemType.CASH.value:
                cash_register = _get_cash_register(session)
                cash_register.uah_balance = max(0.0, cash_register.uah_balance - qty)
                session.add(cash_register)
                session.add(Transaction(
                    currency=Currency.UAH,
                    amount=qty,
                    transaction_type=TransactionType.SUBTRACT,
                    user_id=current_user.id,
                    description=f"Скасування прийому за контрактом #{contract.id} ({owner.full_name if owner else '?'})",
                    uah_balance_after=cash_register.uah_balance,
                    usd_balance_after=cash_register.usd_balance,
                    eur_balance_after=cash_register.eur_balance
                ))

    # ─── CASH: фермер платив грошима → повертаємо з каси ───
    elif payment.payment_type == FarmerContractPaymentType.CASH.value:
        if payment.amount and payment.amount > 0:
            cash_register = _get_cash_register(session)
            currency_str = payment.currency.value if hasattr(payment.currency, 'value') else (payment.currency or "UAH")
            balance_field_map = {"UAH": "uah_balance", "USD": "usd_balance", "EUR": "eur_balance"}
            balance_field = balance_field_map.get(currency_str, "uah_balance")

            current_balance = getattr(cash_register, balance_field)
            setattr(cash_register, balance_field, max(0.0, current_balance - payment.amount))
            session.add(cash_register)

            try:
                currency_enum = Currency(currency_str)
            except ValueError:
                currency_enum = Currency.UAH

            session.add(Transaction(
                currency=currency_enum,
                amount=payment.amount,
                transaction_type=TransactionType.SUBTRACT,
                user_id=current_user.id,
                description=f"Скасування оплати боргу фермера #{contract.id} ({owner.full_name if owner else '?'})",
                uah_balance_after=cash_register.uah_balance,
                usd_balance_after=cash_register.usd_balance,
                eur_balance_after=cash_register.eur_balance
            ))

        # Повертаємо борг
        contract.balance_uah += payment.amount_uah
        session.add(contract)

    # ─── GRAIN: контрагент платив зерном → повертаємо зерно йому ───
    elif payment.payment_type == FarmerContractPaymentType.GRAIN.value:
        qty = payment.quantity_kg or 0.0

        if qty > 0 and payment.culture_id:
            stock = _get_or_create_grain_stock(session, payment.culture_id)
            stock.own_quantity_kg = max(0.0, stock.own_quantity_kg - qty)
            if contract.person_id:
                # Повертаємо у бакет людини на складі.
                stock.person_quantity_kg = (stock.person_quantity_kg or 0.0) + qty
            else:
                # Фермеру повертаємо у farmer_quantity_kg + видаляємо deduction-запис.
                stock.farmer_quantity_kg += qty
                deduction = session.exec(
                    select(FarmerGrainDeduction).where(
                        FarmerGrainDeduction.payment_id == payment.id
                    )
                ).first()
                if deduction:
                    session.delete(deduction)
            session.add(stock)

        # Повертаємо борг
        contract.balance_uah += payment.amount_uah
        session.add(contract)

    # ─── VOUCHER: талон на зерно — видача не торкалась складу, при скасуванні лише відкочуємо delivered і борг ───
    elif payment.payment_type == FarmerContractPaymentType.VOUCHER.value:
        qty = payment.quantity_kg or 0.0

        # Повертаємо delivered_kg у позиції контракту
        if payment.contract_item_id:
            voucher_item = session.get(FarmerContractItem, payment.contract_item_id)
            if voucher_item:
                voucher_item.delivered_kg = max(0.0, voucher_item.delivered_kg - qty)
                session.add(voucher_item)

        # Видача талону не змінює balance_uah, тому й при скасуванні борг не коригуємо.

        # Видаляємо або закриваємо пов'язаний талон
        voucher = session.exec(
            select(GrainVoucher).where(
                GrainVoucher.farmer_contract_payment_id == payment.id
            )
        ).first()
        if voucher:
            # Если по талону уже были выплаты — нельзя отменить
            from backend.models import GrainVoucherPayment
            active_vp = session.exec(
                select(GrainVoucherPayment).where(
                    GrainVoucherPayment.voucher_id == voucher.id,
                    GrainVoucherPayment.is_cancelled == False
                )
            ).all()
            if active_vp:
                raise HTTPException(
                    status_code=400,
                    detail="Неможливо скасувати: по талону вже є активні виплати. Спершу скасуйте виплати по талону."
                )
            session.delete(voucher)

        session.add(contract)

    # ─── SETTLEMENT: розрахунок (контракт виплати) ───
    # Скасування дзеркалить creation: зерно з власного складу → на баланс
    # фермера/людини, видаляємо deductions, гроші повертаємо у касу,
    # контракт переводимо в OPEN з balance_uah=0 (нічого не винні —
    # повернули і зерно, і гроші).
    elif payment.payment_type == FarmerContractPaymentType.SETTLEMENT.value:
        items = session.exec(
            select(FarmerContractItem).where(FarmerContractItem.contract_id == contract.id)
        ).all()

        for item in items:
            if item.item_type != FarmerContractItemType.GRAIN.value or not item.culture_id:
                continue
            qty = item.quantity_kg or 0.0
            if qty <= 0:
                continue

            stock = _get_or_create_grain_stock(session, item.culture_id)
            stock.own_quantity_kg = max(0.0, stock.own_quantity_kg - qty)
            if contract.person_id:
                stock.person_quantity_kg = (stock.person_quantity_kg or 0.0) + qty
            else:
                stock.farmer_quantity_kg += qty
            session.add(stock)

            item.delivered_kg = max(0.0, (item.delivered_kg or 0.0) - qty)
            session.add(item)

            # Видаляємо запис списання з балансу фермера (для людини deduction-ів немає)
            if not contract.person_id and contract.owner_id:
                deduction = session.exec(
                    select(FarmerGrainDeduction).where(
                        FarmerGrainDeduction.owner_id == contract.owner_id,
                        FarmerGrainDeduction.culture_id == item.culture_id,
                        FarmerGrainDeduction.quantity_kg == qty,
                    ).order_by(FarmerGrainDeduction.created_at.desc()).limit(1)
                ).first()
                if deduction:
                    session.delete(deduction)

        # Повертаємо гроші у касу
        amount = payment.amount or 0.0
        if amount > 0:
            cash_register = _get_cash_register(session)
            currency_str = payment.currency.value if hasattr(payment.currency, 'value') else (payment.currency or "UAH")
            balance_field_map = {"UAH": "uah_balance", "USD": "usd_balance", "EUR": "eur_balance"}
            balance_field = balance_field_map.get(currency_str, "uah_balance")
            current_balance = getattr(cash_register, balance_field)
            setattr(cash_register, balance_field, current_balance + amount)
            session.add(cash_register)
            try:
                currency_enum = Currency(currency_str)
            except ValueError:
                currency_enum = Currency.UAH

            if contract.person_id:
                p = session.get(Person, contract.person_id)
                counterparty_label = p.full_name if p else "Людина"
            else:
                counterparty_label = owner.full_name if owner else "?"

            session.add(Transaction(
                currency=currency_enum,
                amount=amount,
                transaction_type=TransactionType.ADD,
                user_id=current_user.id,
                description=f"Скасування розрахунку за контрактом #{contract.id} ({counterparty_label})",
                uah_balance_after=cash_register.uah_balance,
                usd_balance_after=cash_register.usd_balance,
                eur_balance_after=cash_register.eur_balance
            ))

        # Контракт стає відкритим з нульовим балансом — зерно і гроші повернули,
        # ніхто нікому нічого не винен. Оператор може створити нову операцію.
        contract.balance_uah = 0.0
        session.add(contract)

    # Якщо контракт був закритий — відкриваємо назад
    if contract.status == FarmerContractStatus.CLOSED.value:
        contract.status = FarmerContractStatus.OPEN.value
        session.add(contract)

    payment.is_cancelled = True
    payment.updated_at = datetime.utcnow()
    session.add(payment)

    # Атомарно: повернення зі складу/каси + reopen контракту + cancel payment.
    try:
        session.commit()
    except Exception:
        session.rollback()
        raise
    session.refresh(payment)
    return payment

